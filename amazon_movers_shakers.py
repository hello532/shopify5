#!/usr/bin/env python3
"""
amazon_movers_shakers.py — Amazon Movers & Shakers P0 榜单抓取器

输出：/Users/doi/Desktop/Selection/output/amazon_movers_shakers_YYYY-MM-DD.xlsx
字段：Category | Rank | ASIN | Product Title | Price | Rating | Reviews | Change | URL

说明：Amazon 页面经常反爬；抓取失败时不伪造数据，直接输出错误。
"""
from __future__ import annotations

import random
import re
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path("/Users/doi/Desktop/Selection/output")
OUT_DIR.mkdir(parents=True, exist_ok=True)

BASE = "https://www.amazon.com"
CATEGORIES: Dict[str, str] = {
    "Beauty": "beauty",
    "Health": "hpc",
    "Home & Garden": "home-garden",
    "Kitchen": "kitchen",
    "Sports": "sporting-goods",
    "Pet Supplies": "pet-supplies",
    "Baby": "baby-products",
    "Patio & Garden": "lawn-garden",
    "Office": "office-products",
    "Automotive": "automotive",
    "Electronics": "electronics",
    "Wireless": "wireless",
    "Fashion": "fashion",
}

UA_POOL = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
]


def clean(text: str | None) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def asin_from_url(url: str) -> str:
    m = re.search(r"/(?:dp|gp/product)/([A-Z0-9]{10})", url or "")
    return m.group(1) if m else ""


def product_url(href: str) -> str:
    if not href:
        return ""
    if href.startswith("/"):
        href = BASE + href
    asin = asin_from_url(href)
    return f"{BASE}/dp/{asin}" if asin else href.split("?")[0]


def get_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": random.choice(UA_POOL),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    })
    return s


def fetch(session: requests.Session, url: str) -> str:
    r = session.get(url, timeout=30)
    r.raise_for_status()
    if "captcha" in r.text.lower() or "robot check" in r.text.lower():
        raise RuntimeError("Amazon CAPTCHA / robot check")
    return r.text


def parse_page(html: str, category: str) -> List[dict]:
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select("div.zg-grid-general-faceout") or soup.select("div.p13n-grid-content")
    rows: List[dict] = []
    for idx, card in enumerate(cards, 1):
        a = card.select_one("a.a-link-normal[href*='/dp/'], a.a-link-normal[href*='/gp/product/']")
        url = product_url(a.get("href", "") if a else "")
        asin = asin_from_url(url)

        title_el = card.select_one("div._cDEzb_p13n-sc-css-line-clamp-3_g3dy1, div._cDEzb_p13n-sc-css-line-clamp-4_2q2cc, span.a-truncate-cut, img[alt]")
        if title_el and title_el.name == "img":
            title = clean(title_el.get("alt"))
        else:
            title = clean(title_el.get_text(" ") if title_el else "")
        if not title and a:
            title = clean(a.get_text(" "))
        if not title or not asin:
            continue

        rank_el = card.select_one("span.zg-bdg-text")
        rank = clean(rank_el.get_text() if rank_el else f"#{idx}")
        change_el = card.select_one("span.zg-percent-change, span.a-color-price")
        change = clean(change_el.get_text() if change_el else rank)
        price_el = card.select_one("span.a-price span.a-offscreen")
        price = clean(price_el.get_text() if price_el else "")
        rating_el = card.select_one("span.a-icon-alt")
        rating = clean(rating_el.get_text() if rating_el else "")
        reviews_el = card.select_one("a.a-size-small.a-link-normal, span.a-size-small")
        reviews = clean(reviews_el.get_text() if reviews_el else "")

        rows.append({
            "Category": category,
            "Rank": rank,
            "ASIN": asin,
            "Product Title": title[:180],
            "Price": price,
            "Rating": rating,
            "Reviews": reviews,
            "Change": change,
            "URL": url,
        })
    return rows


def extract_keywords(rows: List[dict], top_n: int = 200) -> List[tuple[str, int]]:
    stop = set("the and for with from that this your you are not but all new pack set kit women men kids baby amazon best plus into out use home".split())
    words = []
    for r in rows:
        for w in re.findall(r"[A-Za-z][A-Za-z0-9\-]{2,}", r.get("Product Title", "").lower()):
            if w not in stop and not w.isdigit():
                words.append(w)
    return Counter(words).most_common(top_n)


def save_xlsx(rows: List[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movers & Shakers"
    headers = ["Category", "Rank", "ASIN", "Product Title", "Price", "Rating", "Reviews", "Change", "URL"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    widths = [18, 10, 14, 65, 12, 20, 12, 12, 42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    kw = wb.create_sheet("关键词提取")
    kw.append(["Keyword", "Count"])
    for word, cnt in extract_keywords(rows):
        kw.append([word, cnt])
    for cell in kw[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
    kw.column_dimensions["A"].width = 30
    kw.column_dimensions["B"].width = 12

    wb.save(out_path)


def main():
    session = get_session()
    rows: List[dict] = []
    errors = []
    for category, slug in CATEGORIES.items():
        url = f"{BASE}/gp/movers-and-shakers/{slug}/"
        try:
            print(f"抓取 {category}: {url}")
            html = fetch(session, url)
            parsed = parse_page(html, category)
            print(f"  -> {len(parsed)} products")
            rows.extend(parsed[:30])
        except Exception as e:
            print(f"  !! {category} failed: {e}")
            errors.append({"category": category, "error": str(e)})
        time.sleep(random.uniform(2.0, 4.0))

    if not rows:
        raise SystemExit(f"没有抓到产品。errors={errors[:3]}")

    out_path = OUT_DIR / f"amazon_movers_shakers_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    save_xlsx(rows, out_path)
    print(f"✅ Saved {len(rows)} products -> {out_path}")
    if errors:
        print(f"⚠️ 部分品类失败：{len(errors)}")


if __name__ == "__main__":
    main()
