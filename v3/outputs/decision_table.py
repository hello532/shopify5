"""Decision table report — Excel + HTML + JSON, same data three formats."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from .. import config, db

DECISION_ICONS = {"GO_TEST": "🟢", "WATCH": "🟡", "KILL": "🔴"}


def _query_rows(since: datetime | None = None) -> list[dict[str, Any]]:
    sql = """
    SELECT p.id, p.title, p.shop_domain, p.handle, p.price_usd, p.image_url,
           p.product_url, p.landing_url, p.category,
           sc.fb_score, sc.profit_score, sc.trend_score, sc.lp_score, sc.composite_score,
           ad.days_active, ad.impressions_total, ad.distinct_entity_ids,
           ad.creative_count_raw, ad.countries_running, ad.homogeneity_flag,
           pr.gross_margin_pct, pr.markup_multiplier, pr.beroas, pr.target_roas,
           pr.source_cost_usd, pr.price_band,
           tr.score_7d, tr.score_90d, tr.yoy_growth, tr.seasonality_phase,
           lp.has_shopify, lp.has_klaviyo, lp.has_reviews_app, lp.has_pixel,
           lp.payment_methods_count, lp.awareness_match_layer,
           d.decision, d.kill_reason, d.watch_reason, d.watch_recheck_at, d.test_plan_json,
           d.decided_at
    FROM products p
    LEFT JOIN (SELECT product_id, MAX(id) AS mid FROM scores GROUP BY product_id) sm ON sm.product_id = p.id
    LEFT JOIN scores sc ON sc.id = sm.mid
    LEFT JOIN (SELECT product_id, MAX(id) AS mid FROM ad_signals GROUP BY product_id) am ON am.product_id = p.id
    LEFT JOIN ad_signals ad ON ad.id = am.mid
    LEFT JOIN (SELECT product_id, MAX(id) AS mid FROM profit_signals GROUP BY product_id) pm ON pm.product_id = p.id
    LEFT JOIN profit_signals pr ON pr.id = pm.mid
    LEFT JOIN (SELECT product_id, MAX(id) AS mid FROM trend_signals GROUP BY product_id) tm ON tm.product_id = p.id
    LEFT JOIN trend_signals tr ON tr.id = tm.mid
    LEFT JOIN (SELECT product_id, MAX(id) AS mid FROM lp_signals GROUP BY product_id) lm ON lm.product_id = p.id
    LEFT JOIN lp_signals lp ON lp.id = lm.mid
    LEFT JOIN (SELECT product_id, MAX(id) AS mid FROM decisions GROUP BY product_id) dm ON dm.product_id = p.id
    LEFT JOIN decisions d ON d.id = dm.mid
    WHERE d.id IS NOT NULL
    """
    params: list[Any] = []
    if since:
        sql += " AND d.decided_at >= ?"
        params.append(since.isoformat(timespec="seconds") + "Z")
    sql += " ORDER BY CASE d.decision WHEN 'GO_TEST' THEN 0 WHEN 'WATCH' THEN 1 ELSE 2 END, sc.composite_score DESC NULLS LAST"
    with db.conn() as c:
        rows = c.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def generate(since: datetime | None = None, fmt: str = "xlsx") -> Path:
    rows = _query_rows(since)
    out_dir = config.resolve_path("paths.reports")
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    if fmt == "json":
        p = out_dir / f"decisions_{ts}.json"
        p.write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return p
    if fmt == "html":
        return _write_html(rows, out_dir / f"decisions_{ts}.html")
    return _write_xlsx(rows, out_dir / f"decisions_{ts}.xlsx")


def _write_xlsx(rows: list[dict[str, Any]], path: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to JSON when openpyxl missing
        path = path.with_suffix(".json")
        path.write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "decisions"
    headers = [
        "决策", "ID", "标题", "店铺", "卖价",
        "成本/毛利%", "BEROAS/TgROAS", "Markup",
        "广告天数", "展示量", "Distinct创意/原始",
        "Trends 7d/90d", "YoY%", "阶段",
        "LP信号", "Awareness",
        "Composite", "fb/profit/trend/lp",
        "Kill原因", "Watch原因/复查日",
        "测试预算", "TargetROAS", "Kill D3/D7", "Scale D7≥",
        "落地页", "图",
    ]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(bold=True, color="FFFFFF")
    fills = {
        "GO_TEST": PatternFill("solid", fgColor="DCFCE7"),
        "WATCH": PatternFill("solid", fgColor="FEF9C3"),
        "KILL": PatternFill("solid", fgColor="FEE2E2"),
    }
    for r in rows:
        tp = json.loads(r["test_plan_json"]) if r.get("test_plan_json") else {}
        icon = DECISION_ICONS.get(r.get("decision"), "")
        ws.append([
            f"{icon} {r.get('decision') or ''}",
            r["id"],
            r.get("title") or "",
            r.get("shop_domain") or "",
            r.get("price_usd"),
            _fmt_cost_margin(r),
            _fmt_beroas(r),
            r.get("markup_multiplier"),
            r.get("days_active"),
            r.get("impressions_total"),
            f"{r.get('distinct_entity_ids') or 0}/{r.get('creative_count_raw') or 0}",
            f"{_fmt(r.get('score_7d'))}/{_fmt(r.get('score_90d'))}",
            r.get("yoy_growth"),
            r.get("seasonality_phase"),
            _fmt_lp(r),
            r.get("awareness_match_layer") or "",
            r.get("composite_score"),
            f"{_fmt(r.get('fb_score'))}/{_fmt(r.get('profit_score'))}/{_fmt(r.get('trend_score'))}/{_fmt(r.get('lp_score'))}",
            r.get("kill_reason") or "",
            f"{r.get('watch_reason') or ''} {r.get('watch_recheck_at') or ''}".strip(),
            tp.get("test_budget_usd"),
            tp.get("target_roas"),
            f"D3<{tp.get('kill_rule_day3') or ''}/D7<{tp.get('kill_rule_day7') or ''}",
            tp.get("scale_rule"),
            r.get("landing_url") or r.get("product_url"),
            r.get("image_url"),
        ])
        fill = fills.get(r.get("decision"))
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = fill
    # Column widths
    widths = [12, 6, 30, 22, 8, 14, 14, 8, 10, 12, 14, 14, 8, 10, 22, 14, 11, 22, 16, 24, 10, 12, 16, 12, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w
    wb.save(path)
    return path


def _write_html(rows: list[dict[str, Any]], path: Path) -> Path:
    head = """<!doctype html><html><head><meta charset='utf-8'><title>v3 Decisions</title>
    <style>
    body{font:14px -apple-system,sans-serif;margin:20px;background:#0b0f17;color:#e5e7eb}
    table{border-collapse:collapse;width:100%;background:#111827}
    th,td{padding:8px 10px;border-bottom:1px solid #1f2937;text-align:left;vertical-align:top}
    th{background:#1f2937;position:sticky;top:0}
    tr.GO_TEST{background:#052e16}
    tr.WATCH{background:#3b2f10}
    tr.KILL{background:#3b0a0a;opacity:0.55}
    img{height:38px;border-radius:4px}
    a{color:#93c5fd}
    .pill{padding:2px 6px;border-radius:6px;font-weight:600}
    .pill.GO_TEST{background:#16a34a;color:#fff}
    .pill.WATCH{background:#eab308;color:#000}
    .pill.KILL{background:#dc2626;color:#fff}
    </style></head><body>
    <h1>v3 决策表</h1>
    <p>Generated %s · Click 落地页 to open; right-click 图 to save.</p>
    """ % datetime.utcnow().isoformat(timespec="seconds")
    rows_html = []
    for r in rows:
        tp = json.loads(r["test_plan_json"]) if r.get("test_plan_json") else {}
        dec = r.get("decision") or ""
        rows_html.append(f"""<tr class='{dec}'>
            <td><span class='pill {dec}'>{DECISION_ICONS.get(dec,'')} {dec}</span></td>
            <td>{r['id']}</td>
            <td>{(r.get('title') or '')[:80]}</td>
            <td>{r.get('shop_domain') or ''}</td>
            <td>${r.get('price_usd') or ''}</td>
            <td>{_fmt_cost_margin(r)}</td>
            <td>{_fmt_beroas(r)}</td>
            <td>{_fmt(r.get('markup_multiplier'))}x</td>
            <td>{r.get('days_active') or '–'}d</td>
            <td>{r.get('impressions_total') or '–'}</td>
            <td>{r.get('distinct_entity_ids') or 0}/{r.get('creative_count_raw') or 0}</td>
            <td>{_fmt(r.get('score_7d'))}/{_fmt(r.get('score_90d'))}</td>
            <td>{_fmt(r.get('yoy_growth'))}%</td>
            <td>{_fmt_lp(r)}</td>
            <td>{r.get('awareness_match_layer') or ''}</td>
            <td>{_fmt(r.get('composite_score'))}</td>
            <td>{r.get('kill_reason') or ''}{r.get('watch_reason') or ''}</td>
            <td>${tp.get('test_budget_usd') or ''}</td>
            <td>{tp.get('target_roas') or ''}</td>
            <td>D3<{tp.get('kill_rule_day3') or ''}<br>D7<{tp.get('kill_rule_day7') or ''}</td>
            <td>{tp.get('scale_rule') or ''}</td>
            <td><a href='{r.get('landing_url') or r.get('product_url') or ''}' target='_blank'>open</a></td>
            <td>{f'<img src=\"{r["image_url"]}\">' if r.get('image_url') else ''}</td>
        </tr>""")
    html = head + "<table><thead><tr>" + "".join(
        f"<th>{h}</th>" for h in [
            "决策", "ID", "标题", "店铺", "卖价", "成本/毛利", "BEROAS/Tg", "Markup",
            "广告天", "展示", "Distinct/Raw", "7d/90d", "YoY", "LP", "Awareness",
            "Composite", "原因", "测试$", "TgROAS", "Kill", "Scale", "LP", "图",
        ]
    ) + "</tr></thead><tbody>" + "".join(rows_html) + "</tbody></table></body></html>"
    path.write_text(html, encoding="utf-8")
    return path


def _fmt(v: Any, nd: int = 1) -> str:
    if v is None:
        return "–"
    try:
        return f"{float(v):.{nd}f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_cost_margin(r: dict[str, Any]) -> str:
    cost = r.get("source_cost_usd")
    margin = r.get("gross_margin_pct")
    if cost is None and margin is None:
        return "–"
    return f"${cost}/{_fmt(margin)}%"


def _fmt_beroas(r: dict[str, Any]) -> str:
    a = r.get("beroas")
    b = r.get("target_roas")
    return f"{_fmt(a,2)}/{_fmt(b,2)}"


def _fmt_lp(r: dict[str, Any]) -> str:
    flags = []
    if r.get("has_shopify"): flags.append("S")
    if r.get("has_klaviyo"): flags.append("K")
    if r.get("has_reviews_app"): flags.append("R")
    if r.get("has_pixel"): flags.append("P")
    pm = r.get("payment_methods_count") or 0
    return f"{''.join(flags)} ·{pm}pm"
