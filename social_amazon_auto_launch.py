#!/usr/bin/env python3
"""Social + Amazon product selection agent.

The module builds a conservative auto-launch queue from local social trend files
and Amazon demand files, then writes a Shopify DRAFT catalog. It never publishes
products and never copies competitor media.
"""
from __future__ import annotations

import csv
import glob
import json
import math
import os
import re
import subprocess
import sys
import time
import urllib.request
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

BASE_DIR = Path("/Users/doi/Desktop/amazon")
SELECTION_DIR = Path("/Users/doi/Desktop/Selection")
OUT_DIR = BASE_DIR / "output" / "auto_launch"
RANK_MONITOR_LATEST = BASE_DIR / "output" / "ecommerce_rank_monitor" / "latest.json"
PROFIT_PIPELINE_LATEST = BASE_DIR / "output" / "auto_intelligence" / "profit_pipeline_latest.json"
AMAZON_MOVERS_SCRIPT = BASE_DIR / "amazon_movers_shakers.py"

LOCAL_OPENER = urllib.request.build_opener(urllib.request.ProxyHandler({}))

STOPWORDS = {
    "the", "and", "for", "with", "from", "that", "this", "your", "you", "are",
    "not", "but", "all", "new", "pack", "set", "kit", "amazon", "best", "plus",
    "into", "out", "use", "home", "women", "woman", "men", "man", "kids", "baby",
    "gift", "gifts", "mom", "mothers", "mother", "day", "ideas", "product",
    "shop", "store", "official", "sale", "free", "shipping", "get", "make",
    "piece", "pieces", "count", "counts", "basic", "basics", "double", "tipped",
    "large", "small", "medium", "brand", "refill", "replacement", "value",
}

BIG_BRAND_TERMS = {
    "amazon", "kindle", "apple", "nike", "adidas", "samsung", "sony", "lego", "disney", "pokemon",
    "nintendo", "xbox", "playstation", "tesla", "toyota", "honda", "dyson",
    "kitchenaid", "stanley", "lululemon", "patagonia", "north face", "costco",
    "walmart", "target", "ikea", "sephora", "loreal", "maybelline", "cerave",
    "sensodyne", "colgate", "crest", "oral-b", "laura geller", "kitsch",
    "skintific", "ninja", "foodi", "amazon basics", "blue lizard", "pronamel",
    "revlon", "philips", "braun", "conair", "neutrogena", "dove", "olay",
    "cetaphil", "keurig", "instant pot", "shark", "bissell",
}

RISK_TERMS = {
    "cbd", "hemp", "nicotine", "vape", "fda", "ozempic", "diabetes", "insulin",
    "cancer", "pregnancy", "medical device", "prescription", "pokemon", "disney",
}

HARD_FILTER_TERMS = {
    "toothpaste", "sunscreen", "cotton swab", "cotton swabs", "q-tip", "q tips",
    "deodorant", "medicine", "supplement facts", "spf ", "food processor",
}

CATEGORY_PRICE_HINTS = [
    (("shower", "steamer", "aromatherapy", "gift"), 34.99),
    (("heating", "pad", "heat"), 49.99),
    (("pillowcase", "satin", "silk", "cooling"), 29.99),
    (("teeth", "whitening"), 59.99),
    (("hair", "removal", "ipl"), 89.99),
    (("massage", "massager"), 69.99),
    (("red", "light", "therapy"), 99.99),
    (("organizer", "storage"), 39.99),
    (("pet", "dog", "cat"), 34.99),
    (("fitness", "resistance", "bands"), 39.99),
]

CONCEPT_RULES = [
    (("shower", "steamer", "aromatherapy"), "Aromatherapy Shower Steamers Gift Set", "Beauty & Wellness"),
    (("heating", "pad"), "Portable Heat Therapy Pad", "Pain Relief"),
    (("mugwort", "clay"), "Mugwort Clay Mask Set", "Skincare"),
    (("african", "bath"), "Exfoliating Bath Sponge Set", "Bath & Body"),
    (("pillowcase", "satin"), "Cooling Satin Pillowcase Set", "Sleep & Beauty"),
    (("teeth", "whitening"), "LED Teeth Whitening Kit", "Beauty & Personal Care"),
    (("hair", "removal"), "At-Home Hair Removal Device", "Beauty Device"),
    (("massage", "massager"), "Portable Massage Relief Pad", "Massage & Recovery"),
    (("red", "light"), "Red Light Therapy Device", "Light Therapy"),
    (("resistance", "bands"), "Resistance Training Bundle", "Fitness & Recovery"),
    (("weighted", "vest"), "Adjustable Weighted Vest", "Fitness & Recovery"),
    (("organizer", "storage"), "Space-Saving Organizer Kit", "Home Organization"),
    (("pet",), "Pet Comfort Accessory", "Pet Supplies"),
]


@dataclass
class SourceStatus:
    name: str
    count: int = 0
    status: str = "missing"
    note: str = ""
    latest_file: str = ""
    freshness_days: int | None = None


@dataclass
class SignalRecord:
    source: str
    platform: str
    title: str
    url: str = ""
    category: str = ""
    brand: str = ""
    image: str = ""
    price: float | None = None
    rank: str = ""
    rating: float | None = None
    reviews: int | None = None
    sales_estimate: int | None = None
    score: float = 0.0
    keywords: list[str] = field(default_factory=list)
    evidence: str = ""
    metrics: dict[str, Any] = field(default_factory=dict)
    data_file: str = ""
    data_date: str = ""
    freshness_days: int | None = None
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class Opportunity:
    rank: int
    title: str
    concept_title: str
    product_type: str
    score: float
    decision: str
    can_create_draft: bool
    match_score: float
    suggested_price: float
    evidence_sources: list[str]
    proof_source_count: int
    keywords: list[str]
    risk_flags: list[str]
    amazon: dict[str, Any]
    social_matches: list[dict[str, Any]]
    why: list[str]
    next_actions: list[str]
    shopify_product: dict[str, Any] | None = None


def _text(value: Any, default: str = "") -> str:
    out = re.sub(r"\s+", " ", str(value or "")).strip()
    return out or default


def _float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", str(value).replace(",", ""))
    return float(match.group(1)) if match else None


def _int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    match = re.search(r"([0-9]+)", str(value).replace(",", ""))
    return int(match.group(1)) if match else None


def _slug(text: str, fallback: str = "product") -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", _text(text).lower()).strip("-")
    return slug[:80] or fallback


def _tokens(*parts: Any) -> list[str]:
    raw = " ".join(_text(part) for part in parts)
    words = []
    for word in re.findall(r"[a-zA-Z][a-zA-Z0-9\-]{2,}", raw.lower()):
        word = word.strip("-")
        if len(word) < 3 or word in STOPWORDS or word.isdigit():
            continue
        if any(word == bad or word in bad.split() for bad in BIG_BRAND_TERMS):
            continue
        words.append(word)
    seen: set[str] = set()
    out: list[str] = []
    for word in words:
        if word not in seen:
            seen.add(word)
            out.append(word)
    return out[:18]


def _file_age_days(path: Path | None) -> int | None:
    if not path or not path.exists():
        return None
    return max(0, int((time.time() - path.stat().st_mtime) // 86400))


def _date_from_name(path: Path | None) -> str:
    if not path:
        return ""
    match = re.search(r"(20\d{2})[-_]?(\d{2})[-_]?(\d{2})", path.name)
    if not match:
        return ""
    return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"


def _newest_file(patterns: list[str]) -> Path | None:
    files: list[Path] = []
    for pattern in patterns:
        files.extend(Path(p) for p in glob.glob(pattern, recursive=True))
    files = [p for p in files if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def _is_blocked(text: str) -> bool:
    lower = _text(text).lower()
    return any(term in lower for term in BIG_BRAND_TERMS) or any(term in lower for term in HARD_FILTER_TERMS)


def _risk_flags(*parts: Any) -> list[str]:
    lower = " ".join(_text(part).lower() for part in parts)
    flags = [term for term in RISK_TERMS if term in lower]
    flags.extend([term for term in HARD_FILTER_TERMS if term in lower])
    return sorted(set(flags))


def _estimate_monthly_sales(rank: Any, reviews: Any, change: Any = "") -> int | None:
    rank_int = _int(rank)
    review_count = _int(reviews) or 0
    change_text = _text(change)
    base = 0
    if rank_int:
        base = max(12, int(2200 / math.sqrt(max(rank_int, 1))))
    if review_count:
        base = max(base, int(review_count * 0.06))
    if "#" in change_text or "%" in change_text:
        base = int(base * 1.12) if base else 20
    return base or None


def _source_meta(path: Path | None) -> tuple[str, str, int | None]:
    return (path.name if path else "", _date_from_name(path), _file_age_days(path))


def _record(source: str, platform: str, title: str, path: Path | None, **kwargs: Any) -> SignalRecord | None:
    title = _text(title)
    if not title or _is_blocked(title):
        return None
    data_file, data_date, age = _source_meta(path)
    keywords = kwargs.pop("keywords", None) or _tokens(
        title,
        kwargs.get("category", ""),
        kwargs.get("evidence", ""),
        kwargs.get("metrics", {}),
    )
    if not keywords:
        return None
    return SignalRecord(
        source=source,
        platform=platform,
        title=title[:220],
        keywords=keywords,
        data_file=data_file,
        data_date=data_date,
        freshness_days=age,
        **kwargs,
    )


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _row_get(mapping: dict[str, Any], *keys: str) -> Any:
    lower = {str(k).strip().lower(): v for k, v in mapping.items()}
    for key in keys:
        if key in mapping and mapping[key] not in (None, ""):
            return mapping[key]
        lk = key.lower()
        if lk in lower and lower[lk] not in (None, ""):
            return lower[lk]
    return None


def refresh_source_files() -> list[dict[str, Any]]:
    """Best-effort source refresh. Failures are reported, not hidden."""
    logs: list[dict[str, Any]] = []
    if AMAZON_MOVERS_SCRIPT.exists():
        started = time.time()
        try:
            proc = subprocess.run(
                [sys.executable, str(AMAZON_MOVERS_SCRIPT)],
                cwd=str(BASE_DIR),
                capture_output=True,
                text=True,
                timeout=int(os.getenv("AUTO_LAUNCH_AMAZON_REFRESH_TIMEOUT", "140")),
            )
            logs.append(
                {
                    "source": "Amazon Movers & Shakers",
                    "ok": proc.returncode == 0,
                    "elapsed_seconds": round(time.time() - started, 1),
                    "stdout": (proc.stdout or "")[-1200:],
                    "stderr": (proc.stderr or "")[-1200:],
                }
            )
        except Exception as exc:
            logs.append({"source": "Amazon Movers & Shakers", "ok": False, "error": str(exc)})

    # TikTok local project has its own daily runner; only call it when explicitly enabled.
    if os.getenv("AUTO_LAUNCH_REFRESH_TIKTOK", "").lower() in {"1", "true", "yes", "on"}:
        daily = SELECTION_DIR / "tiktok" / "daily.sh"
        if daily.exists():
            started = time.time()
            try:
                proc = subprocess.run(
                    [str(daily)],
                    cwd=str(daily.parent),
                    capture_output=True,
                    text=True,
                    timeout=int(os.getenv("AUTO_LAUNCH_TIKTOK_REFRESH_TIMEOUT", "240")),
                )
                logs.append(
                    {
                        "source": "TikTok Top Ads",
                        "ok": proc.returncode == 0,
                        "elapsed_seconds": round(time.time() - started, 1),
                        "stdout": (proc.stdout or "")[-1200:],
                        "stderr": (proc.stderr or "")[-1200:],
                    }
                )
            except Exception as exc:
                logs.append({"source": "TikTok Top Ads", "ok": False, "error": str(exc)})
    return logs


def load_tiktok(limit: int = 240) -> tuple[list[SignalRecord], SourceStatus]:
    patterns = [
        str(SELECTION_DIR / "tiktok" / "output" / "**" / "*.csv"),
        str(SELECTION_DIR / "tiktok" / "output" / "**" / "*.json"),
        str(SELECTION_DIR / "V7" / "**" / "*tiktok*.json"),
        str(SELECTION_DIR / "V7" / "**" / "*tiktok*.csv"),
    ]
    files: list[Path] = []
    for pattern in patterns:
        files.extend(Path(p) for p in glob.glob(pattern, recursive=True))
    files = [
        f for f in files
        if f.is_file()
        and any(mark in f.name.lower() for mark in ("top_ads", "tiktok", "viral", "product"))
        and "audit" not in f.name.lower()
    ]
    if not files:
        return [], SourceStatus("TikTok", 0, "missing", "未找到 TikTok 本地输出")
    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:8]
    out: list[SignalRecord] = []
    for path in files:
        try:
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
                    for row in csv.DictReader(handle):
                        title = _text(_row_get(row, "ad_title", "title", "product_name", "highlight_text"))
                        if not title:
                            continue
                        category = _text(_row_get(row, "industry_name", "category", "first_ecom_category"))
                        metrics = {
                            "likes": _int(_row_get(row, "like", "likes")),
                            "shares": _int(_row_get(row, "share", "shares")),
                            "comments": _int(_row_get(row, "comment", "comments")),
                            "ctr": _float(_row_get(row, "ctr")),
                            "selection_score": _float(_row_get(row, "selection_score")),
                            "commerce_type": _text(_row_get(row, "commerce_type")),
                        }
                        score = max(
                            _float(metrics.get("selection_score")) or 0,
                            min(100, 35 + math.log1p(metrics.get("likes") or 0) * 8 + (_float(metrics.get("ctr")) or 0) * 15),
                        )
                        rec = _record(
                            "TikTok Top Ads",
                            "TikTok",
                            title,
                            path,
                            url=_text(_row_get(row, "tt_url", "landing_page", "url")),
                            category=category,
                            brand=_text(_row_get(row, "brand_name", "store_name", "source_brand")),
                            image=_text(_row_get(row, "video_cover")),
                            score=score,
                            metrics=metrics,
                            evidence=f"TikTok ad/product signal; {category}; score={score:.1f}",
                            raw=dict(row),
                        )
                        if rec:
                            out.append(rec)
                        if len(out) >= limit:
                            break
            elif path.suffix.lower() == ".json":
                data = _read_json(path)
                rows = data if isinstance(data, list) else data.get("results") or data.get("items") or data.get("data") or []
                if isinstance(rows, dict):
                    rows = rows.get("list") or rows.get("items") or []
                if not isinstance(rows, list):
                    continue
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    title = _text(_row_get(row, "ad_title", "title", "product_name", "highlight_text"))
                    if not title:
                        continue
                    category = _text(_row_get(row, "industry_name", "category", "first_ecom_category"))
                    metrics = {
                        "likes": _int(_row_get(row, "like", "likes")),
                        "shares": _int(_row_get(row, "share", "shares")),
                        "comments": _int(_row_get(row, "comment", "comments")),
                        "ctr": _float(_row_get(row, "ctr")),
                        "selection_score": _float(_row_get(row, "selection_score")),
                        "commerce_type": _text(_row_get(row, "commerce_type")),
                    }
                    score = max(
                        _float(metrics.get("selection_score")) or 0,
                        min(100, 35 + math.log1p(metrics.get("likes") or 0) * 8 + (_float(metrics.get("ctr")) or 0) * 15),
                    )
                    rec = _record(
                        "TikTok Top Ads",
                        "TikTok",
                        title,
                        path,
                        url=_text(_row_get(row, "tt_url", "landing_page", "url")),
                        category=category,
                        brand=_text(_row_get(row, "brand_name", "store_name", "source_brand")),
                        image=_text(_row_get(row, "video_cover")),
                        score=score,
                        metrics=metrics,
                        evidence=f"TikTok ad/product signal; {category}; score={score:.1f}",
                        raw=dict(row),
                    )
                    if rec:
                        out.append(rec)
                    if len(out) >= limit:
                        break
        except Exception:
            continue
        if len(out) >= limit:
            break
    latest = files[0] if files else None
    return out[:limit], SourceStatus("TikTok", len(out[:limit]), "ok", "本地 Top Ads 输出", str(latest or ""), _file_age_days(latest))


def load_instagram(limit: int = 160) -> tuple[list[SignalRecord], SourceStatus]:
    patterns = [
        str(SELECTION_DIR / "instagram" / "**" / "*.json"),
        str(SELECTION_DIR / "instagram" / "**" / "*.csv"),
        str(BASE_DIR / "output" / "social" / "**" / "*instagram*.json"),
        str(BASE_DIR / "output" / "social" / "**" / "*instagram*.csv"),
    ]
    files: list[Path] = []
    for pattern in patterns:
        files.extend(Path(p) for p in glob.glob(pattern, recursive=True))
    files = [
        f for f in files
        if f.is_file()
        and "shopify_monitor" not in str(f)
        and "_snapshot" not in f.name.lower()
    ]
    if not files:
        return [], SourceStatus("Instagram", 0, "missing", "未找到 Instagram 本地输出；Meta/FB 线索会单独纳入")
    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:6]
    out: list[SignalRecord] = []
    for path in files:
        try:
            rows: list[dict[str, Any]] = []
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
                    rows = list(csv.DictReader(handle))
            elif path.suffix.lower() == ".json":
                data = _read_json(path)
                raw_rows = data if isinstance(data, list) else data.get("items") or data.get("results") or data.get("posts") or []
                rows = [r for r in raw_rows if isinstance(r, dict)]
            for row in rows:
                title = _text(_row_get(row, "caption", "title", "text", "product_name", "name"))
                if not title:
                    continue
                metrics = {
                    "likes": _int(_row_get(row, "likes", "like_count")),
                    "comments": _int(_row_get(row, "comments", "comment_count")),
                    "views": _int(_row_get(row, "views", "play_count")),
                }
                score = min(100, 35 + math.log1p(metrics.get("likes") or 0) * 7 + math.log1p(metrics.get("views") or 0) * 3)
                rec = _record(
                    "Instagram trend file",
                    "Instagram",
                    title,
                    path,
                    url=_text(_row_get(row, "url", "permalink", "link")),
                    category=_text(_row_get(row, "category", "industry")),
                    brand=_text(_row_get(row, "brand", "username", "account")),
                    image=_text(_row_get(row, "image", "thumbnail", "cover_url")),
                    score=score,
                    metrics=metrics,
                    evidence=f"Instagram local signal; score={score:.1f}",
                    raw=dict(row),
                )
                if rec:
                    out.append(rec)
                if len(out) >= limit:
                    break
        except Exception:
            continue
        if len(out) >= limit:
            break
    latest = files[0] if files else None
    return out[:limit], SourceStatus("Instagram", len(out[:limit]), "ok", "本地 Instagram 输出", str(latest or ""), _file_age_days(latest))


def load_meta_orbits(limit: int = 120) -> tuple[list[SignalRecord], SourceStatus]:
    url = f"http://127.0.0.1:8000/orbits?limit={limit}"
    try:
        with LOCAL_OPENER.open(url, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as exc:
        return [], SourceStatus("Meta/FB/IG orbit", 0, "missing", f"8000 端口不可用或无响应: {exc}")
    rows = data if isinstance(data, list) else data.get("items") or data.get("results") or []
    out: list[SignalRecord] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        domain = _text(row.get("domain"))
        name = _text(row.get("canonical_name") or domain)
        if not name or _is_blocked(name + " " + domain):
            continue
        score = _float(row.get("orbit_score")) or 0
        count = _int(row.get("appearance_count")) or 0
        rec = _record(
            "Meta/Facebook/Instagram orbit",
            "Meta",
            name,
            None,
            url=f"https://{domain}" if domain else "",
            brand=name,
            score=score,
            metrics={"orbit_score": score, "appearance_count": count, "domain": domain},
            evidence=f"Meta orbit active brand/domain signal; appearances={count}; score={score}",
            raw=row,
        )
        if rec:
            out.append(rec)
    return out, SourceStatus("Meta/FB/IG orbit", len(out), "ok", "来自 8000 已过滤大品牌的 orbits", url, None)


def load_amazon_xlsx(limit: int = 260) -> tuple[list[SignalRecord], SourceStatus]:
    path = _newest_file([str(SELECTION_DIR / "output" / "amazon_movers_shakers_*.xlsx")])
    if not path:
        return [], SourceStatus("Amazon Movers & Shakers", 0, "missing", "未找到 amazon_movers_shakers_*.xlsx")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return [], SourceStatus("Amazon Movers & Shakers", 0, "error", f"读取失败: {exc}", str(path), _file_age_days(path))
    if not rows:
        return [], SourceStatus("Amazon Movers & Shakers", 0, "empty", "空文件", str(path), _file_age_days(path))
    headers = [_text(h) for h in rows[0]]
    out: list[SignalRecord] = []
    for row in rows[1:limit + 1]:
        d = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        title = _text(_row_get(d, "Product Title", "产品名称", "title"))
        if not title:
            continue
        rank = _text(_row_get(d, "Rank", "排名"))
        reviews = _int(_row_get(d, "Reviews", "评论数"))
        rec = _record(
            "Amazon Movers & Shakers",
            "Amazon",
            title,
            path,
            url=_text(_row_get(d, "URL", "链接")),
            category=_text(_row_get(d, "Category", "品类")),
            price=_float(_row_get(d, "Price", "价格")),
            rank=rank,
            rating=_float(_row_get(d, "Rating", "评分")),
            reviews=reviews,
            sales_estimate=_estimate_monthly_sales(rank, reviews, _row_get(d, "Change", "变化")),
            score=max(35, 75 - min(_int(rank) or 200, 200) * 0.18),
            evidence=f"Amazon Movers rank={rank}; change={_text(_row_get(d, 'Change', '变化'))}",
            metrics={"change": _text(_row_get(d, "Change", "变化")), "asin": _text(_row_get(d, "ASIN"))},
            raw=d,
        )
        if rec:
            out.append(rec)
    return out, SourceStatus("Amazon Movers & Shakers", len(out), "ok", "Amazon 榜单本地抓取", str(path), _file_age_days(path))


def load_rank_monitor(limit: int = 360) -> tuple[list[SignalRecord], list[SignalRecord], SourceStatus]:
    if not RANK_MONITOR_LATEST.exists():
        status = SourceStatus("Rank monitor latest", 0, "missing", "未找到 ecommerce_rank_monitor/latest.json")
        return [], [], status
    try:
        data = _read_json(RANK_MONITOR_LATEST)
    except Exception as exc:
        status = SourceStatus("Rank monitor latest", 0, "error", f"读取失败: {exc}", str(RANK_MONITOR_LATEST))
        return [], [], status
    items = data.get("items") if isinstance(data, dict) else []
    if not isinstance(items, list):
        items = []
    amazon: list[SignalRecord] = []
    social: list[SignalRecord] = []
    for item in items[:limit]:
        if not isinstance(item, dict):
            continue
        title = _text(item.get("title"))
        source = _text(item.get("source"))
        platform = _text(item.get("platform"))
        evidence = _text(item.get("evidence")) + "; " + _text(item.get("trend_metrics"))
        raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
        keywords = _tokens(title, item.get("category"), evidence, raw.get("匹配FB关键词"), raw.get("keyword_list"))
        if "amazon" in (source + " " + platform).lower():
            rank = _text(item.get("rank") or raw.get("Amazon排名") or raw.get("Rank"))
            reviews = _int(item.get("reviews") or raw.get("Reviews") or raw.get("评论数"))
            rec = _record(
                source or "Amazon rank monitor",
                "Amazon",
                title,
                RANK_MONITOR_LATEST,
                url=_text(item.get("url")),
                category=_text(item.get("category")),
                price=_float(item.get("price")),
                rank=rank,
                rating=_float(item.get("rating") or raw.get("评分")),
                reviews=reviews,
                sales_estimate=_estimate_monthly_sales(rank, reviews, item.get("change")),
                score=_float(item.get("score")) or 45,
                keywords=keywords,
                evidence=evidence,
                metrics={"decision": item.get("decision"), "trend_metrics": item.get("trend_metrics")},
                raw=item,
            )
            if rec:
                amazon.append(rec)
            fb_keywords = _text(raw.get("匹配FB关键词"))
            if fb_keywords:
                srec = _record(
                    "Facebook keyword demand from Amazon dual-signal",
                    "Facebook/Meta",
                    fb_keywords[:180],
                    RANK_MONITOR_LATEST,
                    url="",
                    category=_text(item.get("category")),
                    score=_float(raw.get("FB广告分(45%)") or item.get("score")) or 55,
                    keywords=_tokens(fb_keywords, title, item.get("category")),
                    evidence=f"FB keywords matched to Amazon ASIN; {fb_keywords[:220]}",
                    metrics={
                        "active_ads": _int(raw.get("活跃广告数")),
                        "total_ads": _int(raw.get("FB广告总量")),
                        "keyword_count": _int(raw.get("关键词数")),
                        "linked_amazon_title": title,
                        "linked_amazon_url": _text(item.get("url")),
                    },
                    raw=raw,
                )
                if srec:
                    social.append(srec)
        elif any(mark in (source + " " + platform).lower() for mark in ("tiktok", "facebook", "meta", "instagram", "fb keyword")):
            rec = _record(
                source or "Social rank monitor",
                platform or "Social",
                title,
                RANK_MONITOR_LATEST,
                url=_text(item.get("url")),
                category=_text(item.get("category")),
                brand=_text(item.get("brand")),
                score=_float(item.get("score")) or 40,
                keywords=keywords,
                evidence=evidence,
                metrics={"decision": item.get("decision"), "trend_metrics": item.get("trend_metrics")},
                raw=item,
            )
            if rec:
                social.append(rec)
    status = SourceStatus(
        "Rank monitor latest",
        len(amazon) + len(social),
        "ok",
        f"Amazon {len(amazon)} / Social {len(social)}",
        str(RANK_MONITOR_LATEST),
        _file_age_days(RANK_MONITOR_LATEST),
    )
    return amazon, social, status


def load_profit_pipeline(limit: int = 120) -> tuple[list[SignalRecord], SourceStatus]:
    if not PROFIT_PIPELINE_LATEST.exists():
        return [], SourceStatus("V2 28-follow evidence", 0, "missing", "未找到 profit_pipeline_latest.json")
    try:
        data = _read_json(PROFIT_PIPELINE_LATEST)
    except Exception as exc:
        return [], SourceStatus("V2 28-follow evidence", 0, "error", f"读取失败: {exc}", str(PROFIT_PIPELINE_LATEST))
    actions = data.get("actions") if isinstance(data, dict) else []
    if not isinstance(actions, list):
        actions = []
    out: list[SignalRecord] = []
    for action in actions[:limit]:
        if not isinstance(action, dict):
            continue
        title = _text(action.get("title"))
        validation = action.get("validation") if isinstance(action.get("validation"), dict) else {}
        proof = action.get("proof_sources") or []
        metrics = {
            "proof_sources": proof,
            "fb_creative_count": ((validation.get("fb_ads") or {}).get("creative_count") if isinstance(validation.get("fb_ads"), dict) else None),
            "trend_status": ((validation.get("google_trends") or {}).get("status") if isinstance(validation.get("google_trends"), dict) else None),
            "operator_gate": ((action.get("operator_gate") or {}).get("status") if isinstance(action.get("operator_gate"), dict) else None),
        }
        rec = _record(
            "V2 follow-product market proof",
            "DTC+Meta+Trends",
            title,
            PROFIT_PIPELINE_LATEST,
            url=_text(action.get("product_url")),
            category=_text(action.get("product_type")),
            brand=_text(action.get("domain")),
            price=_float(action.get("price")),
            score=_float(action.get("score")) or 50,
            keywords=_tokens(title, action.get("product_type"), metrics),
            evidence=f"V2 proof_sources={','.join(map(str, proof))}; gate={metrics.get('operator_gate')}",
            metrics=metrics,
            raw=action,
        )
        if rec:
            out.append(rec)
    return out, SourceStatus("V2 28-follow evidence", len(out), "ok", "已有竞品/趋势/小垂直站验真", str(PROFIT_PIPELINE_LATEST), _file_age_days(PROFIT_PIPELINE_LATEST))


def collect_sources() -> tuple[list[SignalRecord], list[SignalRecord], list[SourceStatus]]:
    statuses: list[SourceStatus] = []
    amazon: list[SignalRecord] = []
    social: list[SignalRecord] = []

    amz, st = load_amazon_xlsx()
    amazon.extend(amz)
    statuses.append(st)

    amz2, soc2, st = load_rank_monitor()
    amazon.extend(amz2)
    social.extend(soc2)
    statuses.append(st)

    for loader in (load_tiktok, load_instagram, load_meta_orbits, load_profit_pipeline):
        rows, status = loader()
        social.extend(rows)
        statuses.append(status)

    amazon = _dedupe_records(amazon)
    social = _dedupe_records(social)
    return amazon, social, statuses


def _dedupe_records(records: list[SignalRecord]) -> list[SignalRecord]:
    seen: dict[str, SignalRecord] = {}
    for rec in records:
        key = " ".join(rec.keywords[:8]) or _slug(rec.title)
        if key not in seen or rec.score > seen[key].score:
            seen[key] = rec
    return list(seen.values())


def _match_score(amazon: SignalRecord, social: SignalRecord) -> float:
    a = set(amazon.keywords)
    b = set(social.keywords)
    if not a or not b:
        return 0.0
    linked_title = _text(social.metrics.get("linked_amazon_title") if isinstance(social.metrics, dict) else "")
    if social.source.startswith("Facebook keyword demand") and linked_title:
        linked_tokens = set(_tokens(linked_title))
        title_overlap = len(a & linked_tokens) / max(1, min(len(a), len(linked_tokens)))
        linked_url = _text(social.metrics.get("linked_amazon_url") if isinstance(social.metrics, dict) else "").rstrip("/")
        same_url = bool(_text(amazon.url).rstrip("/") and _text(amazon.url).rstrip("/") == linked_url)
        if not same_url and title_overlap < 0.48:
            return 0.0
    overlap = len(a & b)
    coefficient = overlap / max(1, min(len(a), len(b)))
    jaccard = overlap / max(1, len(a | b))
    score = coefficient * 0.72 + jaccard * 0.28
    cat_a = _text(amazon.category).lower()
    cat_b = _text(social.category).lower()
    if cat_a and cat_b and (cat_a in cat_b or cat_b in cat_a):
        score += 0.12
    if social.platform == "Meta" and overlap:
        score *= 0.75
    return min(score, 1.0)


def _infer_concept(title: str, keywords: list[str], category: str) -> tuple[str, str]:
    hay = " ".join([title.lower(), category.lower(), " ".join(keywords)])
    for keys, concept, product_type in CONCEPT_RULES:
        if all(key in hay for key in keys):
            return concept, product_type
    clean_words = [w for w in keywords if w not in STOPWORDS][:4]
    if clean_words:
        readable = " ".join(w.capitalize() for w in clean_words[:3])
        return f"{readable} Kit", category or "General Merchandise"
    if category:
        return f"{category[:42]} Product Concept", category
    return "Validated Product Concept", "General Merchandise"


def _suggest_price(amazon: SignalRecord, concept_keywords: list[str]) -> float:
    if amazon.price and amazon.price >= 9:
        price = amazon.price * 1.35
        if price < 25:
            price = 29.99
        return round(price - 0.01 if abs(price - round(price)) < 0.05 else price, 2)
    hay = set(concept_keywords)
    for keys, price in CATEGORY_PRICE_HINTS:
        if any(key in hay for key in keys):
            return price
    return 49.99


def _opportunity_score(amazon: SignalRecord, matches: list[tuple[float, SignalRecord]]) -> float:
    best_match = matches[0][0] if matches else 0
    platform_count = len({m.platform for _, m in matches})
    freshness_bonus = 0
    for rec in [amazon] + [m for _, m in matches[:3]]:
        if rec.freshness_days is None:
            continue
        if rec.freshness_days <= 1:
            freshness_bonus += 4
        elif rec.freshness_days <= 7:
            freshness_bonus += 2
    amazon_part = min(42, amazon.score * 0.45)
    social_part = min(34, sum(min(100, m.score) for _, m in matches[:3]) / max(1, len(matches[:3])) * 0.32 if matches else 0)
    return round(min(100, amazon_part + social_part + best_match * 18 + platform_count * 3 + freshness_bonus), 1)


def _source_name(record: SignalRecord) -> str:
    return f"{record.platform}:{record.source}"


def _build_description_html(opp: Opportunity) -> str:
    why = "".join(f"<li>{_html(item)}</li>" for item in opp.why[:8])
    actions = "".join(f"<li>{_html(item)}</li>" for item in opp.next_actions[:8])
    sources = "".join(f"<li>{_html(src)}</li>" for src in opp.evidence_sources[:10])
    kw = ", ".join(opp.keywords[:10])
    risk = ", ".join(opp.risk_flags) if opp.risk_flags else "No hard compliance flags detected by local rules."
    return (
        "<h2>Product Concept</h2>"
        f"<p>{_html(opp.concept_title)} is a draft-only product scaffold generated from cross-validated social and Amazon demand signals.</p>"
        "<h3>Why This Is In The Queue</h3>"
        f"<ul>{why}</ul>"
        "<h3>Evidence Sources</h3>"
        f"<ul>{sources}</ul>"
        "<h3>Original Launch Work Required</h3>"
        f"<ul>{actions}</ul>"
        "<h3>Safety Notes</h3>"
        "<ul>"
        "<li>Use original media only. Do not copy competitor images, videos, reviews, creator clips, logos, or brand claims.</li>"
        "<li>Confirm supplier cost, certifications, shipping time, refund reserve, and claim compliance before publishing.</li>"
        f"<li>Seed keywords: {_html(kw)}</li>"
        f"<li>Risk screen: {_html(risk)}</li>"
        "</ul>"
    )


def _html(value: Any) -> str:
    text = _text(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _build_shopify_product(opp: Opportunity) -> dict[str, Any]:
    handle = f"auto-launch-{datetime.now().strftime('%Y%m%d')}-{opp.rank}-{_slug(opp.concept_title)}"
    price = f"{opp.suggested_price:.2f}"
    compare = f"{max(opp.suggested_price * 1.8, opp.suggested_price + 20):.2f}"
    evidence_json = json.dumps(
        {
            "score": opp.score,
            "proof_source_count": opp.proof_source_count,
            "match_score": opp.match_score,
            "sources": opp.evidence_sources,
            "amazon": {
                "title": opp.amazon.get("title"),
                "url": opp.amazon.get("url"),
                "rank": opp.amazon.get("rank"),
                "sales_estimate": opp.amazon.get("sales_estimate"),
            },
        },
        ensure_ascii=False,
    )
    tags = [
        "auto-launch-draft",
        "social-amazon-validated",
        "needs-human-review",
        "needs-original-media",
        "needs-supplier",
        "needs-compliance",
        f"score-{int(opp.score)}",
    ]
    for source in opp.evidence_sources:
        if "TikTok" in source:
            tags.append("tiktok-signal")
        if "Instagram" in source or "Meta" in source:
            tags.append("meta-signal")
        if "Amazon" in source:
            tags.append("amazon-demand")
    return {
        "title": opp.concept_title,
        "handle": handle,
        "descriptionHtml": _build_description_html(opp),
        "vendor": os.getenv("SHOPIFY_VENDOR", "Auto Launch Agent"),
        "productType": opp.product_type,
        "status": "DRAFT",
        "tags": sorted(set(tags))[:20],
        "seo": {
            "title": f"{opp.concept_title} | Draft",
            "description": "Draft-only product scaffold generated from social trend and Amazon demand validation. Original media and compliance review required.",
        },
        "metafields": [
            {"namespace": "auto_launch", "key": "score", "type": "number_decimal", "value": str(opp.score)},
            {"namespace": "auto_launch", "key": "proof_source_count", "type": "number_integer", "value": str(opp.proof_source_count)},
            {"namespace": "auto_launch", "key": "keywords", "type": "multi_line_text_field", "value": "\n".join(opp.keywords[:20])},
            {"namespace": "auto_launch", "key": "amazon_source_url", "type": "url", "value": _text(opp.amazon.get("url"))[:255] or "https://www.amazon.com"},
            {"namespace": "auto_launch", "key": "evidence_json", "type": "json", "value": evidence_json[:9000]},
        ],
        "productOptions": [{"name": "Title", "values": [{"name": "Default Title"}]}],
        "variants": [
            {
                "sku": f"AUTO-{datetime.now().strftime('%y%m%d')}-{opp.rank:03d}",
                "price": price,
                "compareAtPrice": compare,
                "inventoryPolicy": "DENY",
                "taxable": True,
                "optionValues": [{"optionName": "Title", "name": "Default Title"}],
            }
        ],
    }


def build_opportunities(
    amazon_records: list[SignalRecord],
    social_records: list[SignalRecord],
    *,
    limit: int = 20,
    min_score: float = 62,
    min_match_score: float = 0.18,
) -> list[Opportunity]:
    opportunities: list[Opportunity] = []
    for amazon in amazon_records:
        risks = _risk_flags(amazon.title, amazon.category)
        if risks:
            continue
        matches = [
            (score, social)
            for social in social_records
            for score in [_match_score(amazon, social)]
            if score >= min_match_score
        ]
        matches.sort(key=lambda item: (item[0], item[1].score), reverse=True)
        if not matches:
            continue
        matches = matches[:5]
        all_keywords = []
        for word in amazon.keywords + [w for _, social in matches for w in social.keywords]:
            if word not in all_keywords:
                all_keywords.append(word)
        concept_title, product_type = _infer_concept(amazon.title, amazon.keywords, amazon.category)
        score = _opportunity_score(amazon, matches)
        evidence_sources = []
        for source in [_source_name(amazon)] + [_source_name(m) for _, m in matches]:
            if source not in evidence_sources:
                evidence_sources.append(source)
        proof_count = len(evidence_sources)
        can_create = score >= min_score and proof_count >= 2 and matches[0][0] >= min_match_score
        decision = "CREATE_DRAFT" if can_create else "REPORT_ONLY"
        price = _suggest_price(amazon, all_keywords)
        why = [
            f"Amazon demand signal: {amazon.source} rank {amazon.rank or 'n/a'}, estimated monthly sales {amazon.sales_estimate or 'n/a'}.",
            f"Best social match: {matches[0][1].platform} / {matches[0][1].source}, keyword match {matches[0][0]:.2f}.",
            f"Evidence source count: {proof_count}; candidate score {score}.",
        ]
        if amazon.freshness_days is not None:
            why.append(f"Amazon source freshness: {amazon.freshness_days} day(s).")
        if matches[0][1].freshness_days is not None:
            why.append(f"Social source freshness: {matches[0][1].freshness_days} day(s).")
        next_actions = [
            "Find 2-3 suppliers and confirm landed COGS, MOQ, shipping time, and refund reserve.",
            "Create original PDP images/video and 6 short-form ad scripts; do not reuse competitor media.",
            "Verify claims/compliance wording before publishing the Shopify draft.",
            "Run a 48-hour small-budget test only after pixel/CAPI and product feed are clean.",
        ]
        opp = Opportunity(
            rank=0,
            title=amazon.title,
            concept_title=concept_title,
            product_type=product_type,
            score=score,
            decision=decision,
            can_create_draft=can_create,
            match_score=round(matches[0][0], 3),
            suggested_price=price,
            evidence_sources=evidence_sources,
            proof_source_count=proof_count,
            keywords=all_keywords[:20],
            risk_flags=_risk_flags(amazon.title, concept_title, product_type),
            amazon=asdict(amazon),
            social_matches=[
                {
                    "match_score": round(match_score, 3),
                    **asdict(social),
                }
                for match_score, social in matches
            ],
            why=why,
            next_actions=next_actions,
        )
        opportunities.append(opp)
    opportunities.sort(key=lambda item: (item.can_create_draft, item.score, item.match_score), reverse=True)
    filtered = []
    seen_concepts: set[str] = set()
    for opp in opportunities:
        concept_key = _slug(opp.concept_title)
        if concept_key in seen_concepts:
            continue
        seen_concepts.add(concept_key)
        filtered.append(opp)
    for idx, opp in enumerate(filtered[:limit], start=1):
        opp.rank = idx
        if opp.can_create_draft:
            opp.shopify_product = _build_shopify_product(opp)
    return filtered[:limit]


def _write_markdown(payload: dict[str, Any], path: Path) -> None:
    lines = [
        "# 社媒 + Amazon 自动选品与 Shopify 草稿报告",
        "",
        f"- 生成时间: {payload.get('generated_at')}",
        f"- 候选数: {payload.get('stats', {}).get('opportunities')}",
        f"- 可创建 DRAFT: {payload.get('stats', {}).get('draft_products')}",
        f"- 安全规则: 只生成 Shopify DRAFT；不复制竞品素材、评价、Logo、品牌词；发布前必须人工审核。",
        "",
        "## 数据源",
    ]
    for source in payload.get("sources", []):
        lines.append(f"- {source.get('name')}: {source.get('count')} · {source.get('status')} · {source.get('note')} · {source.get('latest_file', '')}")
    lines.extend(["", "## Top 机会"])
    for opp in payload.get("opportunities", [])[:30]:
        lines.extend(
            [
                "",
                f"### #{opp.get('rank')} {opp.get('concept_title')}",
                f"- 决策: {opp.get('decision')} · 分数 {opp.get('score')} · 匹配 {opp.get('match_score')}",
                f"- Amazon: {opp.get('amazon', {}).get('title')} ({opp.get('amazon', {}).get('url', '')})",
                f"- 证据: {', '.join(opp.get('evidence_sources') or [])}",
                f"- 关键词: {', '.join(opp.get('keywords') or [])}",
                "- 下一步: " + " / ".join(opp.get("next_actions") or []),
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_artifacts(
    opportunities: list[Opportunity],
    statuses: list[SourceStatus],
    refresh_logs: list[dict[str, Any]] | None = None,
    *,
    min_score: float,
    min_match_score: float,
) -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    products = [opp.shopify_product for opp in opportunities if opp.shopify_product]
    catalog = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "mode": "social_amazon_auto_launch",
        "safety": {
            "status": "DRAFT",
            "human_review_required": True,
            "rule": "Cross-validated product concepts only. Competitor media, copied text, reviews, logos, and brand terms are excluded.",
        },
        "products": products,
    }
    catalog_path = OUT_DIR / f"auto_launch_catalog_{ts}.json"
    normalized_path = OUT_DIR / f"auto_launch_catalog_{ts}.normalized.json"
    upload_report_path = OUT_DIR / f"auto_launch_upload_report_{ts}.json"
    report_path = OUT_DIR / f"auto_launch_report_{ts}.md"
    payload_path = OUT_DIR / f"auto_launch_{ts}.json"

    payload = {
        "ok": True,
        "generated_at": catalog["generated_at"],
        "version": "social-amazon-auto-launch-v1",
        "safety": catalog["safety"],
        "config": {
            "min_score": min_score,
            "min_match_score": min_match_score,
        },
        "stats": {
            "opportunities": len(opportunities),
            "draft_products": len(products),
            "blocked_report_only": len([opp for opp in opportunities if not opp.can_create_draft]),
            "sources_with_data": len([s for s in statuses if s.count > 0]),
        },
        "sources": [asdict(s) for s in statuses],
        "refresh_logs": refresh_logs or [],
        "opportunities": [asdict(opp) for opp in opportunities],
        "artifact": {
            "catalog_path": str(catalog_path),
            "normalized_path": str(normalized_path),
            "upload_report_path": str(upload_report_path),
            "report_path": str(report_path),
            "payload_path": str(payload_path),
            "products": [
                {
                    "title": p.get("title"),
                    "handle": p.get("handle"),
                    "status": p.get("status"),
                    "price": ((p.get("variants") or [{}])[0] or {}).get("price"),
                }
                for p in products
            ],
        },
    }
    catalog_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    _write_markdown(payload, report_path)
    (OUT_DIR / "latest.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (OUT_DIR / "latest_catalog.json").write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def run_auto_launch(
    *,
    limit: int = 20,
    min_score: float = 62,
    min_match_score: float = 0.18,
    refresh_sources: bool = False,
) -> dict[str, Any]:
    refresh_logs = refresh_source_files() if refresh_sources else []
    amazon_records, social_records, statuses = collect_sources()
    opportunities = build_opportunities(
        amazon_records,
        social_records,
        limit=limit,
        min_score=min_score,
        min_match_score=min_match_score,
    )
    payload = write_artifacts(
        opportunities,
        statuses,
        refresh_logs,
        min_score=min_score,
        min_match_score=min_match_score,
    )
    payload["raw_counts"] = {"amazon_records": len(amazon_records), "social_records": len(social_records)}
    return payload


def latest_auto_launch() -> dict[str, Any]:
    path = OUT_DIR / "latest.json"
    if not path.exists():
        return {"ok": False, "error": "No auto-launch result yet"}
    try:
        return _read_json(path)
    except Exception as exc:
        return {"ok": False, "error": str(exc), "path": str(path)}


if __name__ == "__main__":
    result = run_auto_launch(
        limit=int(os.getenv("AUTO_LAUNCH_LIMIT", "20")),
        min_score=float(os.getenv("AUTO_LAUNCH_MIN_SCORE", "62")),
        refresh_sources=os.getenv("AUTO_LAUNCH_REFRESH", "").lower() in {"1", "true", "yes", "on"},
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
