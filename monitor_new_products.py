#!/usr/bin/env python3
"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 monitor_new_products.py — 全量竞品新品监控器
 2107 个 Shopify 独立站，只监控新品，结果写 Excel
 
 用法:
   python3 monitor_new_products.py                  # 全量跑
   python3 monitor_new_products.py --limit 100      # 只跑前100个
   python3 monitor_new_products.py --workers 10     # 10并发
   python3 monitor_new_products.py --geo US         # 趋势验证地区
   python3 monitor_new_products.py --trend-workers 4
   python3 monitor_new_products.py --no-trend       # 只监控新品，不做趋势验证
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
import json
import os
import re
import sys
import time
import hashlib
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from xml.etree import ElementTree as ET
from urllib.parse import quote

try:
    from pytrends.request import TrendReq
    HAS_TRENDS = True
except Exception:
    HAS_TRENDS = False

# ─── 配置 ───
BASE_DIR = Path.home() / "Desktop/amazon"
DATA_DIR = BASE_DIR / "output/shopify_monitor"
SNAPSHOT_DIR = DATA_DIR / "snapshots"
WATCHLIST_FILE = DATA_DIR / "competitors_watchlist.json"
OUTPUT_DIR = DATA_DIR / "new_products"
REPORT_DIR = DATA_DIR / "reports"
TREND_CACHE_FILE = DATA_DIR / "trend_cache.json"
PLAYBOOK_DIR = DATA_DIR / "playbooks"
QUEUE_DIR = DATA_DIR / "queues"
DAILY_DIR = DATA_DIR / "daily_checklists"
HERO_DIR = DATA_DIR / "hero_products"

SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)
PLAYBOOK_DIR.mkdir(parents=True, exist_ok=True)
QUEUE_DIR.mkdir(parents=True, exist_ok=True)
DAILY_DIR.mkdir(parents=True, exist_ok=True)
HERO_DIR.mkdir(parents=True, exist_ok=True)

TIMEOUT = 12
DELAY = 0.3
MAX_WORKERS = 20  # 并发数
MAX_TREND_WORKERS = 4
MAX_PRODUCT_PAGES = 10
MAX_SITEMAP_CANDIDATES = 40
MAX_HERO_SCAN = 50
HERO_CANDIDATES_PER_DOMAIN = 2
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36"
TREND_TIMEFRAME = "today 3-m"
EVERGREEN_TIMEFRAME = "today 12-m"

TREND_STOPWORDS = {
    "the", "and", "for", "with", "from", "your", "you", "this", "that", "new",
    "best", "sale", "shop", "store", "official", "premium", "limited", "edition",
    "kit", "set", "bundle", "pack", "plus", "pro", "ultra", "mini", "max",
}

CHANNEL_HINTS = {
    "beauty": "TikTok UGC + Meta Broad",
    "fashion": "TikTok Spark + Meta Reels",
    "pet": "Meta Video + TikTok Problem/Solution",
    "recovery": "Meta UGC/Testimonial + Google Search",
    "fitness": "Meta UGC/Testimonial + TikTok Demo",
    "home": "Meta Video + Pinterest/UGC",
    "tech": "Meta Video + Google Shopping",
    "baby": "Meta Social Proof + Google Search",
    "default": "Meta Broad + TikTok Test",
}

VERTICAL_KEYWORDS = {
    "beauty": {"serum", "skincare", "skin", "beauty", "mask", "facial", "hair", "lash", "lip", "nail"},
    "fashion": {"dress", "shirt", "leggings", "fashion", "jewelry", "ring", "bag", "shoe", "hoodie"},
    "pet": {"dog", "cat", "pet", "litter", "paw", "leash", "collar", "treat"},
    "recovery": {"recovery", "boots", "massager", "brace", "posture", "therapy", "pain", "muscle", "knee", "back"},
    "fitness": {"fitness", "gym", "workout", "training", "protein", "yoga", "resistance", "sport"},
    "home": {"kitchen", "organizer", "cleaner", "vacuum", "pillow", "mattress", "lamp", "chair", "desk"},
    "tech": {"charger", "camera", "projector", "led", "smart", "wireless", "bluetooth", "gadget"},
    "baby": {"baby", "newborn", "maternity", "nursing", "stroller", "crib"},
}


# ─── 工具函数 ───
def get_session():
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT, "Accept": "application/json"})
    retry = Retry(total=2, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://", HTTPAdapter(max_retries=retry))
    return s


def normalize_domain(d):
    d = d.strip().lower()
    d = d.replace("https://", "").replace("http://", "").rstrip("/")
    if d.startswith("www."):
        d = d[4:]
    return d


def get_base_url(domain):
    return f"https://{domain}"


def strip_html(text: str) -> str:
    text = re.sub(r"<[^>]+>", " ", str(text or ""))
    text = re.sub(r"\s+", " ", text).strip()
    return text


def safe_float(value, default=0.0):
    try:
        return float(value)
    except Exception:
        return default


def parse_dt(value: str):
    raw = str(value or "").strip()
    if not raw:
        return None
    raw = raw.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(raw)
    except Exception:
        try:
            return datetime.fromisoformat(raw.split("+")[0])
        except Exception:
            return None


def days_since(value: str):
    dt = parse_dt(value)
    if not dt:
        return None
    return max(0, (datetime.now(dt.tzinfo) - dt).days if dt.tzinfo else (datetime.now() - dt).days)


def days_between(start: str, end: str):
    start_dt = parse_dt(start)
    end_dt = parse_dt(end)
    if not start_dt or not end_dt:
        return None
    return max(0, (end_dt - start_dt).days)


def extract_image_urls(images) -> list[str]:
    urls = []
    for image in images or []:
        if isinstance(image, dict):
            src = image.get("src") or image.get("url") or ""
            if src:
                urls.append(str(src))
        elif isinstance(image, str) and image:
            urls.append(image)
    return urls


def load_trend_cache():
    if TREND_CACHE_FILE.exists():
        try:
            return json.loads(TREND_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_trend_cache(cache):
    TREND_CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_query(text: str) -> str:
    text = re.sub(r"<[^>]+>", " ", str(text or ""))
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"[^\w\s/&+]", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:80]


def build_trend_queries(product: dict) -> list[str]:
    title = normalize_query(product.get("title", ""))
    vendor = normalize_query(product.get("vendor", ""))
    product_type = normalize_query(product.get("product_type", ""))
    handle = normalize_query(product.get("handle", ""))

    candidates = []
    if title:
        candidates.append(title)

    if title and vendor:
        vendor_words = {w.lower() for w in vendor.split() if len(w) > 2}
        stripped = " ".join(
            word for word in title.split()
            if word.lower() not in vendor_words
        ).strip()
        stripped = normalize_query(stripped)
        if stripped and stripped != title:
            candidates.append(stripped)

    if product_type and product_type.lower() not in title.lower():
        candidates.append(product_type)

    if handle and handle.lower() not in {c.lower() for c in candidates}:
        candidates.append(handle)

    cleaned = []
    seen = set()
    for candidate in candidates:
        words = [w for w in candidate.split() if w.lower() not in TREND_STOPWORDS]
        rebuilt = normalize_query(" ".join(words) if words else candidate)
        if len(rebuilt) < 4:
            continue
        key = rebuilt.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(rebuilt)
    return cleaned[:4]


def fetch_google_related_queries(query: str, geo: str = "US", limit: int = 5) -> list[str]:
    encoded = requests.utils.quote(query)
    url = f"https://suggestqueries.google.com/complete/search?client=firefox&hl=en&gl={geo.lower()}&q={encoded}"
    try:
        r = requests.get(url, headers={"User-Agent": USER_AGENT, "Referer": "https://www.google.com/"}, timeout=10)
        r.raise_for_status()
        payload = r.json()
    except Exception:
        return []

    if not isinstance(payload, list) or len(payload) < 2:
        return []

    out = []
    seen = {query.lower()}
    for item in payload[1]:
        text = normalize_query(item)
        if not text or text.lower() in seen:
            continue
        seen.add(text.lower())
        out.append(text)
        if len(out) >= limit:
            break
    return out


def classify_trend(avg_interest: float, latest_interest: float, momentum_pct: float, related_count: int) -> tuple[str, bool]:
    if avg_interest >= 45 and latest_interest >= 50 and momentum_pct >= 8:
        return "hot", True
    if avg_interest >= 28 and momentum_pct >= 0:
        return "watch", True
    if avg_interest >= 18 or related_count >= 3:
        return "emerging", True
    if avg_interest > 0:
        return "weak", False
    return "unverified", False


def verify_query_trend(query: str, geo: str, cache: dict, timeframe: str = TREND_TIMEFRAME) -> dict:
    cache_key = f"{datetime.now().strftime('%Y-%m-%d')}|{geo}|{timeframe}|{query.lower()}"
    if cache_key in cache:
        return cache[cache_key]

    related_queries = fetch_google_related_queries(query, geo=geo, limit=5)
    result = {
        "query": query,
        "geo": geo,
        "source": "fallback",
        "status": "unverified",
        "trend_verified": False,
        "interest_avg": 0.0,
        "interest_latest": 0.0,
        "momentum_pct": 0.0,
        "direction": "flat",
        "related_queries": related_queries,
        "checked_at": datetime.now().isoformat(),
    }

    if HAS_TRENDS:
        try:
            pytrends = TrendReq(hl="en-US", tz=480)
            pytrends.build_payload([query], timeframe=timeframe, geo=geo)
            df = pytrends.interest_over_time()
            if not df.empty and query in df:
                values = [float(v) for v in df[query].tolist()]
                avg_interest = round(sum(values) / len(values), 2)
                latest_interest = round(sum(values[-4:]) / max(1, len(values[-4:])), 2)
                prev_window = values[-8:-4] if len(values) >= 8 else values[:-4]
                prev_interest = round(sum(prev_window) / len(prev_window), 2) if prev_window else avg_interest
                momentum_pct = round(((latest_interest - prev_interest) / prev_interest) * 100, 2) if prev_interest > 0 else 0.0
                direction = "up" if momentum_pct >= 8 else ("down" if momentum_pct <= -8 else "flat")

                try:
                    rq = pytrends.related_queries().get(query, {})
                    rising = rq.get("rising")
                    if rising is not None and "query" in rising:
                        related_queries = [normalize_query(q) for q in rising["query"].tolist()[:5] if normalize_query(q)]
                except Exception:
                    pass

                status, verified = classify_trend(avg_interest, latest_interest, momentum_pct, len(related_queries))
                result.update({
                    "source": "pytrends",
                    "status": status,
                    "trend_verified": verified,
                    "interest_avg": avg_interest,
                    "interest_latest": latest_interest,
                    "momentum_pct": momentum_pct,
                    "direction": direction,
                    "related_queries": related_queries,
                })
        except Exception as e:
            result["error"] = str(e)[:160]

    if result["source"] == "fallback" and related_queries:
        result["status"] = "emerging" if len(related_queries) >= 3 else "weak"
        result["trend_verified"] = len(related_queries) >= 3

    cache[cache_key] = result
    return result


def verify_product_trend(product: dict, geo: str, cache: dict) -> dict:
    queries = build_trend_queries(product)
    best = None
    for query in queries:
        candidate = verify_query_trend(query, geo, cache)
        score = (
            candidate.get("interest_avg", 0) * 1.0
            + max(candidate.get("momentum_pct", 0), 0) * 0.6
            + len(candidate.get("related_queries", [])) * 4
        )
        if best is None or score > best[0]:
            best = (score, candidate)

    if best is None:
        return {
            "trend_query": "",
            "trend_status": "unverified",
            "trend_verified": False,
            "trend_source": "none",
            "trend_interest_avg": 0.0,
            "trend_interest_latest": 0.0,
            "trend_momentum_pct": 0.0,
            "trend_direction": "flat",
            "trend_related_queries": [],
            "trend_checked_at": datetime.now().isoformat(),
        }

    candidate = best[1]
    return {
        "trend_query": candidate.get("query", ""),
        "trend_status": candidate.get("status", "unverified"),
        "trend_verified": candidate.get("trend_verified", False),
        "trend_source": candidate.get("source", "none"),
        "trend_interest_avg": candidate.get("interest_avg", 0.0),
        "trend_interest_latest": candidate.get("interest_latest", 0.0),
        "trend_momentum_pct": candidate.get("momentum_pct", 0.0),
        "trend_direction": candidate.get("direction", "flat"),
        "trend_related_queries": candidate.get("related_queries", []),
        "trend_checked_at": candidate.get("checked_at", datetime.now().isoformat()),
    }


def verify_evergreen_demand(product: dict, geo: str, cache: dict) -> dict:
    queries = build_trend_queries(product)
    best = None
    for query in queries:
        candidate = verify_query_trend(query, geo, cache, timeframe=EVERGREEN_TIMEFRAME)
        score = candidate.get("interest_avg", 0) + candidate.get("interest_latest", 0) * 0.4 + max(candidate.get("momentum_pct", 0), -20) * 0.2
        if best is None or score > best[0]:
            best = (score, candidate)
    if best is None:
        return {
            "evergreen_query": "",
            "evergreen_avg": 0.0,
            "evergreen_latest": 0.0,
            "evergreen_momentum_pct": 0.0,
            "evergreen_direction": "flat",
            "evergreen_status": "unverified",
        }
    candidate = best[1]
    return {
        "evergreen_query": candidate.get("query", ""),
        "evergreen_avg": candidate.get("interest_avg", 0.0),
        "evergreen_latest": candidate.get("interest_latest", 0.0),
        "evergreen_momentum_pct": candidate.get("momentum_pct", 0.0),
        "evergreen_direction": candidate.get("direction", "flat"),
        "evergreen_status": candidate.get("status", "unverified"),
    }


def trend_summary(products: list[dict]) -> dict:
    buckets = {"hot": 0, "watch": 0, "emerging": 0, "weak": 0, "unverified": 0}
    for product in products:
        buckets[product.get("trend_status", "unverified")] = buckets.get(product.get("trend_status", "unverified"), 0) + 1
    return {
        "verified_count": sum(1 for p in products if p.get("trend_verified")),
        "status_breakdown": buckets,
        "top_trending": sorted(
            products,
            key=lambda x: (
                1 if x.get("trend_verified") else 0,
                x.get("trend_interest_avg", 0),
                x.get("trend_momentum_pct", 0),
            ),
            reverse=True,
        )[:20],
    }


def detect_vertical(product: dict) -> str:
    text = " ".join(
        [
            str(product.get("title", "")),
            str(product.get("product_type", "")),
            str(product.get("vendor", "")),
            str(product.get("handle", "")),
        ]
    ).lower()
    for vertical, keywords in VERTICAL_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return vertical
    return "default"


def discount_pct(product: dict) -> float:
    price = safe_float(product.get("price_min"), 0.0)
    compare_at = safe_float(product.get("compare_at_max"), 0.0)
    if compare_at > price > 0:
        return round((compare_at - price) / compare_at * 100, 2)
    return 0.0


def make_meta_urls(product: dict) -> dict:
    domain = normalize_domain(product.get("domain", ""))
    brand = normalize_query(product.get("vendor") or domain.split(".")[0] or "")
    trend_query = normalize_query(product.get("trend_query") or product.get("title") or "")
    return {
        "meta_brand_url": f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={requests.utils.quote(brand)}&search_type=keyword_unordered",
        "meta_domain_url": f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={requests.utils.quote(domain)}&search_type=keyword_unordered",
        "google_ads_transparency_url": f"https://adstransparency.google.com/?region=US&domain={requests.utils.quote(domain)}",
        "tiktok_creative_center_url": "https://ads.tiktok.com/business/creativecenter/inspiration/topads/pc/en",
        "tiktok_query": trend_query,
    }


def suggest_offer_strategy(product: dict) -> str:
    price = safe_float(product.get("price_min"), 0.0)
    variants = int(product.get("variants_count") or 0)
    discount = discount_pct(product)
    vertical = product.get("vertical", "default")
    if discount >= 15:
        return "保留折扣锚点，叠加 bundle / gift，优先抢低 CPA"
    if price <= 39:
        return "做 2 件装 / 3 件装，提高 AOV，再用低价引流"
    if price <= 99:
        return "做 hero bundle + 单品对照，首页突出节省金额"
    if vertical in {"recovery", "fitness", "tech"}:
        return "做价值堆栈 offer，强调结果、质保、赠品和分期"
    if variants >= 4:
        return "突出多规格/多颜色选择，做 best variant 默认推荐"
    return "先上单品测试，再补 bundle 和加购链"


def suggest_creative_angles(product: dict) -> list[str]:
    title = normalize_query(product.get("title", ""))
    trend_query = normalize_query(product.get("trend_query", "")) or title
    vertical = product.get("vertical", "default")
    if vertical == "recovery":
        return [
            f"痛点开头: 用'{trend_query}'解决酸痛/疲劳场景",
            "演示开头: 10 秒展示使用前后状态变化",
            "口碑开头: 评论截屏 + 使用反馈 + 明确 CTA",
        ]
    if vertical in {"beauty", "fashion"}:
        return [
            f"结果开头: '{trend_query}' 前后效果对比",
            "UGC 开头: 真人上脸/上身/开箱体验",
            "种草开头: 限时折扣 + 热评 + 快速转化 CTA",
        ]
    return [
        f"问题开头: '{trend_query}' 当前替代方案不够好",
        "演示开头: 功能点拆成 3 个镜头快速讲清",
        "对比开头: 老方案 vs 新方案，突出更省时/更方便",
    ]


def compute_action_score(product: dict) -> tuple[int, list[str]]:
    score = 0.0
    reasons = []

    trend_status = product.get("trend_status", "unverified")
    if trend_status == "hot":
        score += 32
        reasons.append("趋势热度强")
    elif trend_status == "watch":
        score += 24
        reasons.append("趋势稳定可测")
    elif trend_status == "emerging":
        score += 16
        reasons.append("趋势新兴")
    elif trend_status == "weak":
        score += 6

    price = safe_float(product.get("price_min"), 0.0)
    if 25 <= price <= 120:
        score += 18
        reasons.append("价格带适合冷启动")
    elif 121 <= price <= 220:
        score += 10
        reasons.append("可做中高客单验证")
    elif 10 <= price < 25:
        score += 8
    else:
        score -= 6

    discount = discount_pct(product)
    if discount >= 20:
        score += 15
        reasons.append("有强折扣锚点")
    elif discount >= 10:
        score += 10
        reasons.append("有价格锚点")

    if int(product.get("images_count") or 0) >= 4:
        score += 8
        reasons.append("素材位充足")
    if int(product.get("variants_count") or 0) >= 3:
        score += 7
        reasons.append("可做变体/套装扩展")
    if product.get("source") == "sitemap":
        score += 10
        reasons.append("站点近期新增信号")
    if product.get("available"):
        score += 6
    else:
        score -= 12
    if product.get("trend_verified"):
        score += 10
    if product.get("body_text"):
        score += 4
    if product.get("vendor"):
        score += 3

    return max(0, min(100, int(round(score)))), reasons[:5]


def action_bucket(score: int) -> str:
    if score >= 78:
        return "scale_now"
    if score >= 62:
        return "test_now"
    if score >= 45:
        return "watch_close"
    return "low_priority"


def action_label(bucket: str) -> str:
    labels = {
        "scale_now": "立即抄作业",
        "test_now": "本周立刻测试",
        "watch_close": "持续观察",
        "low_priority": "低优先级",
    }
    return labels.get(bucket, bucket)


def queue_priority(bucket: str) -> str:
    mapping = {
        "scale_now": "P1",
        "test_now": "P1",
        "watch_close": "P2",
        "low_priority": "P3",
    }
    return mapping.get(bucket, "P3")


def product_short_reason(product: dict) -> str:
    reasons = product.get("action_reasons", []) or []
    if reasons:
        return " / ".join(reasons[:3])
    return "监控触发了可行动信号"


def enrich_product_actionability(product: dict) -> dict:
    enriched = dict(product)
    enriched["vertical"] = detect_vertical(enriched)
    score, reasons = compute_action_score(enriched)
    bucket = action_bucket(score)
    enriched["discount_pct"] = discount_pct(enriched)
    enriched["action_score"] = score
    enriched["action_bucket"] = bucket
    enriched["action_label"] = action_label(bucket)
    enriched["action_reasons"] = reasons
    enriched["primary_channel"] = CHANNEL_HINTS.get(enriched["vertical"], CHANNEL_HINTS["default"])
    enriched["offer_strategy"] = suggest_offer_strategy(enriched)
    enriched["creative_angles"] = suggest_creative_angles(enriched)
    enriched["next_step"] = (
        "立刻查广告 -> 拆页面 -> 做 offer -> 开 3 组创意测试"
        if bucket in {"scale_now", "test_now"}
        else "进入观察，等待更多趋势或价格变化确认"
    )
    enriched.update(make_meta_urls(enriched))
    return enriched


def score_change_event(event: dict) -> dict:
    counts = event.get("counts", {})
    score = (
        counts.get("price", 0) * 5
        + counts.get("image", 0) * 4
        + counts.get("copy", 0) * 4
        + counts.get("variant", 0) * 3
        + counts.get("publish", 0) * 6
        + counts.get("availability", 0) * 4
        + counts.get("new", 0) * 8
    )
    scored = dict(event)
    scored["competitive_heat_score"] = min(100, score)
    return scored


def build_opportunity_board(new_products: list[dict], change_events: list[dict], limit: int = 20) -> dict:
    ranked_products = sorted(
        new_products,
        key=lambda x: (
            x.get("action_score", 0),
            1 if x.get("trend_verified") else 0,
            x.get("trend_interest_avg", 0),
        ),
        reverse=True,
    )
    hot_domains = sorted(
        [score_change_event(event) for event in change_events],
        key=lambda x: x.get("competitive_heat_score", 0),
        reverse=True,
    )[:limit]
    return {
        "top_products": ranked_products[:limit],
        "hot_domains": hot_domains,
    }


def make_queue_item(product: dict, queue_type: str, next_step: str, why_now: str) -> dict:
    timebox = {
        "ad_research_queue": 20,
        "creative_queue": 45,
        "landing_page_queue": 40,
        "offer_test_queue": 25,
        "research_queue": 20,
        "watch_queue": 10,
    }.get(queue_type, 20)
    impact = "high" if product.get("action_score", 0) >= 78 else ("medium" if product.get("action_score", 0) >= 55 else "low")
    return {
        "queue_type": queue_type,
        "priority": queue_priority(product.get("action_bucket", "low_priority")),
        "domain": product.get("domain", ""),
        "product_id": product.get("product_id", ""),
        "title": product.get("title", ""),
        "handle": product.get("handle", ""),
        "action_score": product.get("action_score", 0),
        "action_label": product.get("action_label", ""),
        "primary_channel": product.get("primary_channel", ""),
        "offer_strategy": product.get("offer_strategy", ""),
        "trend_status": product.get("trend_status", ""),
        "trend_verified": product.get("trend_verified", False),
        "next_step": next_step,
        "why_now": why_now,
        "meta_brand_url": product.get("meta_brand_url", ""),
        "meta_domain_url": product.get("meta_domain_url", ""),
        "google_ads_transparency_url": product.get("google_ads_transparency_url", ""),
        "tiktok_creative_center_url": product.get("tiktok_creative_center_url", ""),
        "tiktok_query": product.get("tiktok_query", ""),
        "timebox_min": timebox,
        "expected_impact": impact,
        "generated_at": datetime.now().isoformat(),
        "status": "todo",
    }


def build_task_queues(new_products: list[dict], change_events: list[dict], hero_products: list[dict] | None = None, limit_per_queue: int = 30) -> dict:
    queues = {
        "research_queue": [],
        "ad_research_queue": [],
        "creative_queue": [],
        "landing_page_queue": [],
        "offer_test_queue": [],
        "watch_queue": [],
        "hero_queue": [],
    }

    ranked_products = sorted(new_products, key=lambda x: x.get("action_score", 0), reverse=True)
    for product in ranked_products:
        bucket = product.get("action_bucket", "low_priority")
        why_now = product_short_reason(product)

        if bucket in {"scale_now", "test_now"}:
            queues["ad_research_queue"].append(
                make_queue_item(product, "ad_research_queue", "先看 Meta/TikTok 广告素材，再拆 hook / CTA / 评论角度", why_now)
            )
            queues["creative_queue"].append(
                make_queue_item(product, "creative_queue", "按系统给的 3 个角度先产出 6-9 条短视频脚本", why_now)
            )
            queues["landing_page_queue"].append(
                make_queue_item(product, "landing_page_queue", "重做首屏、评价、FAQ、推荐链和支付理由", why_now)
            )
            queues["offer_test_queue"].append(
                make_queue_item(product, "offer_test_queue", f"执行 offer: {product.get('offer_strategy','')}", why_now)
            )
        elif bucket == "watch_close":
            queues["research_queue"].append(
                make_queue_item(product, "research_queue", "补查市场、价格带、广告强度，再决定是否升级测试", why_now)
            )
            queues["watch_queue"].append(
                make_queue_item(product, "watch_queue", "继续观察 3-7 天，等趋势/改价/广告动作二次确认", why_now)
            )
        else:
            queues["watch_queue"].append(
                make_queue_item(product, "watch_queue", "低优先级观察，不进入本周实战队列", why_now)
            )

    hot_domains = sorted(
        [score_change_event(event) for event in change_events],
        key=lambda x: x.get("competitive_heat_score", 0),
        reverse=True,
    )
    for event in hot_domains[:limit_per_queue]:
        if event.get("competitive_heat_score", 0) >= 40:
            queues["research_queue"].append({
                "queue_type": "research_queue",
                "priority": "P1" if event.get("competitive_heat_score", 0) >= 70 else "P2",
                "domain": event.get("domain", ""),
                "product_id": "",
                "title": "",
                "handle": "",
                "action_score": event.get("competitive_heat_score", 0),
                "action_label": "站点层高变化",
                "primary_channel": "",
                "offer_strategy": "",
                "trend_status": "",
                "trend_verified": False,
                "next_step": "优先人工复盘该站最近改价/图片/文案/上架动作，找出主推产品",
                "why_now": f"变化热度高: {event.get('counts', {})}",
                "meta_brand_url": "",
                "meta_domain_url": f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={requests.utils.quote(event.get('domain',''))}&search_type=keyword_unordered",
                "google_ads_transparency_url": f"https://adstransparency.google.com/?region=US&domain={requests.utils.quote(normalize_domain(event.get('domain','')))}",
                "tiktok_creative_center_url": "https://ads.tiktok.com/business/creativecenter/inspiration/topads/pc/en",
                "tiktok_query": normalize_domain(event.get("domain", "")).split(".")[0],
                "generated_at": datetime.now().isoformat(),
                "status": "todo",
            })

    for hero in sorted(hero_products or [], key=lambda x: x.get("hero_score", 0), reverse=True)[:limit_per_queue]:
        queues["hero_queue"].append({
            "queue_type": "hero_queue",
            "priority": "P1" if hero.get("hero_bucket") == "evergreen_hero" else "P2",
            "domain": hero.get("domain", ""),
            "product_id": hero.get("product_id", ""),
            "title": hero.get("title", ""),
            "handle": hero.get("handle", ""),
            "action_score": hero.get("hero_score", 0),
            "action_label": hero.get("hero_bucket", ""),
            "primary_channel": " / ".join(hero.get("likely_traffic_sources", [])),
            "offer_strategy": hero.get("offer_strategy", ""),
            "trend_status": hero.get("evergreen_status", ""),
            "trend_verified": hero.get("evergreen_avg", 0) > 0,
            "next_step": hero.get("hero_next_step", ""),
            "why_now": " / ".join(hero.get("hero_reasons", [])),
            "meta_brand_url": hero.get("meta_brand_url", ""),
            "meta_domain_url": hero.get("meta_domain_url", ""),
            "google_ads_transparency_url": hero.get("google_ads_transparency_url", ""),
            "tiktok_creative_center_url": hero.get("tiktok_creative_center_url", ""),
            "tiktok_query": hero.get("tiktok_query", ""),
            "timebox_min": 35,
            "expected_impact": "high" if hero.get("hero_bucket") == "evergreen_hero" else "medium",
            "generated_at": datetime.now().isoformat(),
            "status": "todo",
        })

    for key in queues:
        queues[key] = sorted(queues[key], key=lambda x: (x.get("priority", "P3"), -int(x.get("action_score", 0))))[:limit_per_queue]
    return queues


def write_task_queues(queues: dict) -> tuple[Path, Path]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = QUEUE_DIR / f"task_queues_{ts}.json"
    md_path = QUEUE_DIR / f"task_queues_{ts}.md"
    json_path.write_text(json.dumps(queues, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [f"# Task Queues {ts}", ""]
    for queue_name, items in queues.items():
        lines.append(f"## {queue_name}")
        if not items:
            lines.append("- (empty)")
            lines.append("")
            continue
        for item in items:
            title = item.get("title") or item.get("domain") or "(unknown)"
            lines.append(f"- [{item.get('priority','P3')}] {title} | {item.get('action_label','')} | Score {item.get('action_score',0)}")
            lines.append(f"  Next: {item.get('next_step','')}")
            lines.append(f"  Why: {item.get('why_now','')}")
        lines.append("")
    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path


def build_daily_checklist(board: dict, task_queues: dict, hero_products: list[dict] | None = None) -> dict:
    def first_item(queue_name):
        items = task_queues.get(queue_name, [])
        return items[0] if items else None

    morning = []
    afternoon = []
    evening = []

    for queue_name in ("ad_research_queue", "research_queue", "offer_test_queue", "creative_queue", "landing_page_queue"):
        item = first_item(queue_name)
        if not item:
            continue
        task = {
            "queue": queue_name,
            "title": item.get("title") or item.get("domain") or "(unknown)",
            "priority": item.get("priority", "P3"),
            "next_step": item.get("next_step", ""),
            "why_now": item.get("why_now", ""),
            "timebox_min": item.get("timebox_min", 20),
            "expected_impact": item.get("expected_impact", "medium"),
        }
        if queue_name in ("ad_research_queue", "research_queue"):
            morning.append(task)
        elif queue_name in ("offer_test_queue", "creative_queue"):
            afternoon.append(task)
        else:
            evening.append(task)

    top_products = board.get("top_products", [])[:10]
    if not top_products and hero_products:
        top_products = hero_products[:10]
    top_domains = board.get("hot_domains", [])[:10]
    return {
        "generated_at": datetime.now().isoformat(),
        "today_focus": [p.get("title") or p.get("domain") for p in top_products[:5]],
        "morning_tasks": morning,
        "afternoon_tasks": afternoon,
        "evening_tasks": evening,
        "top_products": top_products,
        "hot_domains": top_domains,
        "hero_products": (hero_products or [])[:10],
    }


def write_daily_checklist(checklist: dict) -> tuple[Path, Path]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = DAILY_DIR / f"daily_checklist_{ts}.json"
    md_path = DAILY_DIR / f"daily_checklist_{ts}.md"
    json_path.write_text(json.dumps(checklist, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        f"# Daily Checklist {ts}",
        "",
        "## Today Focus",
    ]
    for idx, title in enumerate(checklist.get("today_focus", []), start=1):
        lines.append(f"{idx}. {title}")

    for section_key, title in (
        ("morning_tasks", "Morning Tasks"),
        ("afternoon_tasks", "Afternoon Tasks"),
        ("evening_tasks", "Evening Tasks"),
    ):
        lines.extend(["", f"## {title}"])
        tasks = checklist.get(section_key, [])
        if not tasks:
            lines.append("- (empty)")
            continue
        for task in tasks:
            lines.append(
                f"- [{task.get('priority','P3')}] {task.get('title','')} | {task.get('timebox_min',20)}min | {task.get('expected_impact','medium')}"
            )
            lines.append(f"  Next: {task.get('next_step','')}")
            lines.append(f"  Why: {task.get('why_now','')}")

    lines.extend(["", "## Top Products"])
    for idx, product in enumerate(checklist.get("top_products", []), start=1):
        lines.append(
            f"{idx}. {product.get('title','')} | {product.get('domain','')} | {product.get('action_label','')} | Score {product.get('action_score',0)}"
        )

    lines.extend(["", "## Hot Domains"])
    for idx, event in enumerate(checklist.get("hot_domains", []), start=1):
        lines.append(f"{idx}. {event.get('domain','')} | Heat {event.get('competitive_heat_score',0)} | {event.get('counts',{})}")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path


def write_playbook(board: dict) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = PLAYBOOK_DIR / f"top_opportunities_{ts}.md"
    lines = [
        f"# Shopify 竞品机会榜 {ts}",
        "",
        "## Top Products",
    ]
    for idx, product in enumerate(board.get("top_products", []), start=1):
        lines.extend([
            f"{idx}. {product.get('title','')} | {product.get('domain','')} | {product.get('action_label','')}",
            f"   - Score: {product.get('action_score',0)} | Trend: {product.get('trend_status','')} | Channel: {product.get('primary_channel','')}",
            f"   - Offer: {product.get('offer_strategy','')}",
            f"   - Reasons: {', '.join(product.get('action_reasons', []))}",
            f"   - Meta Brand: {product.get('meta_brand_url','')}",
            f"   - Meta Domain: {product.get('meta_domain_url','')}",
            f"   - Google Ads: {product.get('google_ads_transparency_url','')}",
            f"   - TikTok Query: {product.get('tiktok_query','')} | {product.get('tiktok_creative_center_url','')}",
            "   - Creative Angles:",
        ])
        for angle in product.get("creative_angles", []):
            lines.append(f"     - {angle}")
        lines.append("")
    lines.append("## Hot Domains")
    for idx, domain_event in enumerate(board.get("hot_domains", []), start=1):
        lines.append(
            f"{idx}. {domain_event.get('domain','')} | Heat {domain_event.get('competitive_heat_score',0)} | {domain_event.get('counts',{})}"
        )
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def product_summary_from_payload(product: dict, domain: str, *, source: str = "products_json", sitemap_lastmod: str = "") -> dict:
    variants = product.get("variants", []) or []
    images = product.get("images", []) or []
    image_urls = extract_image_urls(images)
    prices = [safe_float(v.get("price"), 0.0) for v in variants if v.get("price") not in (None, "", 0)]
    body_text = strip_html(product.get("body_html", ""))
    available = any(bool(v.get("available")) for v in variants) if variants else False
    status = "published" if product.get("published_at") else "draft"
    image_url = image_urls[0] if image_urls else str(product.get("featured_image", "") or "")
    compare_prices = [safe_float(v.get("compare_at_price"), 0.0) for v in variants if v.get("compare_at_price")]
    has_subscription = bool(product.get("selling_plan_groups")) or any(bool(v.get("selling_plan_allocations")) for v in variants)
    tags = product.get("tags", [])
    if isinstance(tags, str):
        tags = [t.strip() for t in tags.split(",") if t.strip()]

    return {
        "domain": domain,
        "product_id": str(product.get("id", "")),
        "title": product.get("title", "") or "",
        "handle": product.get("handle", "") or "",
        "product_type": product.get("product_type", "") or "",
        "vendor": product.get("vendor", "") or "",
        "price_min": min(prices) if prices else 0.0,
        "price_max": max(prices) if prices else 0.0,
        "compare_at_max": max(compare_prices) if compare_prices else 0.0,
        "variants_count": len(variants),
        "images_count": len(image_urls),
        "image_url": image_url,
        "url": f"https://{domain}/products/{product.get('handle', '')}",
        "created_at": product.get("created_at", "") or "",
        "published_at": product.get("published_at", "") or "",
        "updated_at": product.get("updated_at", "") or "",
        "status": status,
        "available": available,
        "body_text": body_text[:300],
        "body_hash": hashlib.md5(body_text.encode("utf-8", errors="ignore")).hexdigest() if body_text else "",
        "tags": tags,
        "has_subscription": has_subscription,
        "source": source,
        "sitemap_lastmod": sitemap_lastmod,
        "detected_at": datetime.now().isoformat(),
    }


def extract_review_signals(html: str) -> dict:
    count = 0
    rating = 0.0
    patterns = [
        r'"rating_count"\s*:\s*(\d+)',
        r'"reviewCount"\s*:\s*"?(\d+)"?',
        r'"MetafieldLooxCount"\s*=\s*(\d+)',
        r'LooxCount\s*=\s*(\d+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.I)
        if match:
            count = max(count, int(match.group(1)))
    for pattern in [r'"ratingValue"\s*:\s*"?([0-9.]+)"?', r'"value"\s*:\s*"?([0-9.]+)"?']:
        match = re.search(pattern, html, re.I)
        if match:
            rating = max(rating, safe_float(match.group(1), 0.0))
    return {"review_count": count, "rating_value": rating}


def extract_page_conversion_signals(html: str, product: dict) -> dict:
    lower = html.lower()
    description_blob = (product.get("body_text", "") or "").lower()
    combined = f"{lower}\n{description_blob}"
    rec_count = len(re.findall(r'pr_rec_pid=', html))
    flags = {
        "has_google_ads": "aw-" in lower or "googletagmanager.com/gtag/js?id=aw-" in lower,
        "has_ga4": "googletagmanager.com/gtag/js?id=g-" in lower,
        "has_klaviyo": "klaviyo" in lower,
        "has_loox": "loox" in lower,
        "has_ab_test": "intelligems" in lower,
        "has_referral": "socialsnowball" in lower,
        "has_cart_upsell": "upcart" in lower or "free-gift" in lower or "upsell" in lower,
        "has_subscription_ui": "subscribe" in combined or "selling_plan" in lower or "subscription" in lower or product.get("has_subscription", False),
        "has_free_gift": "free " in combined and ("gift" in combined or "socks" in combined or "belt" in combined),
        "has_training_bonus": "training guide" in combined or "video training" in combined or "workout series" in combined,
        "has_gift_card_bonus": "gift card" in combined,
        "has_refund_policy": "refund policy" in lower or "full refund" in combined or "money back" in combined,
        "has_shipping_info": "shipping info" in lower or "free shipping" in combined,
        "has_shop_pay": "shop pay" in lower,
    }
    return {
        "recommendation_count": rec_count,
        **extract_review_signals(html),
        **flags,
    }


def fetch_product_page_html(domain: str, handle: str, session) -> str:
    if not handle:
        return ""
    try:
        r = session.get(f"{get_base_url(domain)}/products/{handle}", timeout=TIMEOUT)
        if r.status_code != 200:
            return ""
        return r.text
    except Exception:
        return ""


def fetch_product_recommendations(domain: str, product_id: str, session, limit: int = 8) -> list[dict]:
    if not product_id:
        return []
    try:
        r = session.get(
            f"{get_base_url(domain)}/recommendations/products.json?product_id={product_id}&limit={limit}",
            timeout=TIMEOUT,
        )
        if r.status_code != 200:
            return []
        data = r.json()
        return data.get("products", []) if isinstance(data, dict) else []
    except Exception:
        return []


def local_hero_candidate_score(product: dict) -> tuple[int, list[str]]:
    score = 0
    reasons = []
    age_days = days_since(product.get("created_at", ""))
    maintenance_days = days_between(product.get("created_at", ""), product.get("updated_at", ""))
    recent_update_days = days_since(product.get("updated_at", ""))
    if age_days is not None and age_days >= 120:
        score += 18
        reasons.append("老品仍在售卖")
    if maintenance_days is not None and maintenance_days >= 30:
        score += 14
        reasons.append("长期持续维护")
    if recent_update_days is not None and recent_update_days <= 120:
        score += 12
        reasons.append("近期仍在更新")
    if discount_pct(product) >= 15:
        score += 10
        reasons.append("折扣锚点强")
    if int(product.get("images_count") or 0) >= 6:
        score += 8
        reasons.append("素材丰富")
    if int(product.get("variants_count") or 0) >= 3:
        score += 6
        reasons.append("变体策略完整")
    if product.get("available"):
        score += 5
    if product.get("has_subscription"):
        score += 6
        reasons.append("有订阅/续费结构")
    return score, reasons[:5]


def pick_domain_hero_candidates(current_products: dict, limit: int = HERO_CANDIDATES_PER_DOMAIN) -> list[dict]:
    ranked = []
    for product in current_products.values():
        score, reasons = local_hero_candidate_score(product)
        if score <= 0:
            continue
        candidate = dict(product)
        candidate["hero_candidate_score"] = score
        candidate["hero_candidate_reasons"] = reasons
        ranked.append(candidate)
    ranked.sort(key=lambda x: x.get("hero_candidate_score", 0), reverse=True)
    return ranked[:limit]


def classify_hero_bucket(hero_score: int) -> str:
    if hero_score >= 72:
        return "evergreen_hero"
    if hero_score >= 52:
        return "maintained_winner"
    if hero_score >= 36:
        return "hero_candidate"
    return "regular"


def identify_evergreen_hero(product: dict, geo: str, cache: dict, session) -> dict:
    enriched = dict(product)
    page_html = fetch_product_page_html(product.get("domain", ""), product.get("handle", ""), session)
    page_signals = extract_page_conversion_signals(page_html, product) if page_html else {}
    evergreen = verify_evergreen_demand(product, geo, cache)
    recommendations = fetch_product_recommendations(product.get("domain", ""), product.get("product_id", ""), session, limit=8)

    score = 0
    reasons = []

    if evergreen.get("evergreen_avg", 0) >= 35:
        score += 20
        reasons.append("长期需求稳定")
    elif evergreen.get("evergreen_avg", 0) >= 20:
        score += 12
        reasons.append("长期需求存在")

    if evergreen.get("evergreen_latest", 0) >= max(15, evergreen.get("evergreen_avg", 0) * 0.75):
        score += 10
        reasons.append("近期需求没有掉")

    if page_signals.get("review_count", 0) >= 100:
        score += 18
        reasons.append("社证足够强")
    elif page_signals.get("review_count", 0) >= 30:
        score += 10
        reasons.append("已有明显社证")

    if page_signals.get("has_cart_upsell"):
        score += 8
        reasons.append("有加购/购物车加售")
    if page_signals.get("has_free_gift"):
        score += 8
        reasons.append("有赠品堆栈")
    if page_signals.get("has_training_bonus"):
        score += 7
        reasons.append("有内容型增值")
    if page_signals.get("has_subscription_ui"):
        score += 8
        reasons.append("有订阅/复购结构")
    if page_signals.get("has_ab_test"):
        score += 6
        reasons.append("页面在持续 A/B 测试")
    if page_signals.get("has_klaviyo"):
        score += 5
        reasons.append("有邮件承接")
    if page_signals.get("has_referral"):
        score += 5
        reasons.append("有 referral/联盟")
    if page_signals.get("has_google_ads"):
        score += 8
        reasons.append("Google 投流信号")

    recommendation_count = len(recommendations)
    if recommendation_count >= 3:
        score += 10
        reasons.append("推荐链完整")

    maintenance_days = days_between(product.get("created_at", ""), product.get("updated_at", ""))
    if maintenance_days is not None and maintenance_days >= 60:
        score += 10
        reasons.append("不是一次性上架品")

    hero_bucket = classify_hero_bucket(score)
    likely_sources = []
    if page_signals.get("has_google_ads"):
        likely_sources.append("Google Ads / Shopping")
    if page_signals.get("has_klaviyo"):
        likely_sources.append("Email retention")
    if page_signals.get("has_referral"):
        likely_sources.append("Referral / creator")
    if not likely_sources:
        likely_sources.append("Organic + retargeting mix")

    enriched.update(page_signals)
    enriched.update(evergreen)
    enriched["hero_score"] = min(100, score)
    enriched["hero_bucket"] = hero_bucket
    enriched["hero_reasons"] = reasons[:6]
    enriched["recommendation_count"] = recommendation_count
    enriched["recommended_handles"] = [p.get("handle", "") for p in recommendations if p.get("handle")][:8]
    enriched["likely_traffic_sources"] = likely_sources
    enriched["hero_next_step"] = (
        "拆它的广告、offer、推荐链和页面结构，优先复制成你自己的 hero page"
        if hero_bucket in {"evergreen_hero", "maintained_winner"}
        else "继续观察是否进入长期赢家区间"
    )
    return enriched


def parse_sitemap_locs(xml_text: str) -> list[dict]:
    if not xml_text:
        return []
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []

    items = []
    for node in root.findall(".//{*}sitemap") + root.findall(".//{*}url"):
        loc = node.findtext("{*}loc", default="").strip()
        lastmod = node.findtext("{*}lastmod", default="").strip()
        if loc:
            items.append({"loc": loc, "lastmod": lastmod})
    return items


def fetch_product_by_handle(domain: str, handle: str, session) -> dict | None:
    if not handle:
        return None
    url = f"{get_base_url(domain)}/products/{handle}.js"
    try:
        r = session.get(url, timeout=TIMEOUT)
        if r.status_code != 200:
            return None
        data = r.json()
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def fetch_recent_sitemap_products(domain: str, session, existing_handles: set[str], limit: int = MAX_SITEMAP_CANDIDATES) -> list[dict]:
    base = get_base_url(domain)
    sitemap_urls = []
    discovered = []

    try:
        r = session.get(f"{base}/sitemap.xml", timeout=TIMEOUT)
        if r.status_code == 200:
            sitemap_urls = [
                item["loc"]
                for item in parse_sitemap_locs(r.text)
                if "sitemap_products" in item["loc"] or "/products/" in item["loc"]
            ]
    except Exception:
        sitemap_urls = []

    if not sitemap_urls:
        sitemap_urls = [f"{base}/sitemap_products_1.xml"]

    product_entries = []
    for sitemap_url in sitemap_urls[:6]:
        try:
            r = session.get(sitemap_url, timeout=TIMEOUT)
            if r.status_code != 200:
                continue
            for item in parse_sitemap_locs(r.text):
                loc = item["loc"]
                if "/products/" not in loc:
                    continue
                handle = loc.rstrip("/").split("/products/")[-1].split("?")[0].strip()
                if not handle:
                    continue
                product_entries.append({"handle": handle, "loc": loc, "lastmod": item.get("lastmod", "")})
        except Exception:
            continue

    product_entries.sort(key=lambda x: x.get("lastmod", ""), reverse=True)

    seen = set()
    for entry in product_entries:
        handle = entry["handle"].lower()
        if handle in existing_handles or handle in seen:
            continue
        seen.add(handle)
        product = fetch_product_by_handle(domain, entry["handle"], session)
        if product and product.get("id"):
            discovered.append(product_summary_from_payload(product, domain, source="sitemap", sitemap_lastmod=entry.get("lastmod", "")))
        if len(discovered) >= limit:
            break
    return discovered


def detect_catalog_changes(old_products: dict, current_products: dict) -> dict:
    old_ids = set(old_products.keys())
    current_ids = set(current_products.keys())

    new_ids = current_ids - old_ids
    removed_ids = old_ids - current_ids
    common_ids = current_ids & old_ids

    price_changes = []
    image_changes = []
    copy_changes = []
    variant_changes = []
    publish_changes = []
    availability_changes = []

    for pid in sorted(common_ids):
        old = old_products[pid]
        new = current_products[pid]

        if (old.get("price_min"), old.get("price_max"), old.get("compare_at_max")) != (
            new.get("price_min"), new.get("price_max"), new.get("compare_at_max")
        ):
            price_changes.append({
                "domain": new.get("domain"),
                "product_id": pid,
                "title": new.get("title"),
                "handle": new.get("handle"),
                "old_price_min": old.get("price_min", 0),
                "new_price_min": new.get("price_min", 0),
                "old_price_max": old.get("price_max", 0),
                "new_price_max": new.get("price_max", 0),
                "old_compare_at_max": old.get("compare_at_max", 0),
                "new_compare_at_max": new.get("compare_at_max", 0),
                "url": new.get("url", ""),
            })

        if old.get("image_url") != new.get("image_url"):
            image_changes.append({
                "domain": new.get("domain"),
                "product_id": pid,
                "title": new.get("title"),
                "handle": new.get("handle"),
                "old_image_url": old.get("image_url", ""),
                "new_image_url": new.get("image_url", ""),
                "url": new.get("url", ""),
            })

        if old.get("body_hash") != new.get("body_hash") or old.get("title") != new.get("title"):
            copy_changes.append({
                "domain": new.get("domain"),
                "product_id": pid,
                "title": new.get("title"),
                "handle": new.get("handle"),
                "old_title": old.get("title", ""),
                "new_title": new.get("title", ""),
                "url": new.get("url", ""),
            })

        if old.get("variants_count") != new.get("variants_count"):
            variant_changes.append({
                "domain": new.get("domain"),
                "product_id": pid,
                "title": new.get("title"),
                "handle": new.get("handle"),
                "old_variants_count": old.get("variants_count", 0),
                "new_variants_count": new.get("variants_count", 0),
                "url": new.get("url", ""),
            })

        if old.get("status") != new.get("status"):
            publish_changes.append({
                "domain": new.get("domain"),
                "product_id": pid,
                "title": new.get("title"),
                "handle": new.get("handle"),
                "old_status": old.get("status", ""),
                "new_status": new.get("status", ""),
                "url": new.get("url", ""),
            })

        if bool(old.get("available")) != bool(new.get("available")):
            availability_changes.append({
                "domain": new.get("domain"),
                "product_id": pid,
                "title": new.get("title"),
                "handle": new.get("handle"),
                "old_available": bool(old.get("available")),
                "new_available": bool(new.get("available")),
                "url": new.get("url", ""),
            })

    return {
        "new_products": [current_products[pid] for pid in sorted(new_ids)],
        "removed_products": [old_products[pid] for pid in sorted(removed_ids)],
        "price_changes": price_changes,
        "image_changes": image_changes,
        "copy_changes": copy_changes,
        "variant_changes": variant_changes,
        "publish_changes": publish_changes,
        "availability_changes": availability_changes,
        "counts": {
            "new": len(new_ids),
            "removed": len(removed_ids),
            "price": len(price_changes),
            "image": len(image_changes),
            "copy": len(copy_changes),
            "variant": len(variant_changes),
            "publish": len(publish_changes),
            "availability": len(availability_changes),
        },
    }


def load_snapshot(domain):
    """加载域名快照 (产品ID集合)"""
    path = SNAPSHOT_DIR / f"{normalize_domain(domain).replace('.', '_')}.json"
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                products = data.get("products", {})
                legacy_only = False
                if not products and data.get("product_ids"):
                    products = {str(pid): {"product_id": str(pid)} for pid in data.get("product_ids", [])}
                    legacy_only = True
                product_ids = data.get("product_ids") or list(products.keys())
                if products and not legacy_only:
                    rich_sample = next(iter(products.values()), {})
                    legacy_only = not any(
                        key in rich_sample for key in ("price_min", "title", "body_hash", "variants_count")
                    )
                return {
                    "schema_version": data.get("schema_version", 2),
                    "product_ids": [str(pid) for pid in product_ids],
                    "products": products,
                    "last_check": data.get("last_check"),
                    "count": data.get("count", len(product_ids)),
                    "legacy_only": legacy_only,
                }
        except Exception:
            pass
    return {"schema_version": 2, "product_ids": [], "products": {}, "last_check": None, "count": 0, "legacy_only": False}


def save_snapshot(domain, products_map):
    """保存快照"""
    path = SNAPSHOT_DIR / f"{normalize_domain(domain).replace('.', '_')}.json"
    data = {
        "schema_version": 2,
        "product_ids": sorted(products_map.keys()),
        "products": products_map,
        "last_check": datetime.now().isoformat(),
        "count": len(products_map)
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_products_fast(domain, session):
    """快速抓取产品列表 (只取ID+基本信息，不做深度验证)"""
    base = get_base_url(domain)
    products = []
    page = 1
    page_cap_hit = False

    while page <= MAX_PRODUCT_PAGES:  # 最多10页 = 2500产品
        url = f"{base}/products.json?limit=250&page={page}"
        try:
            r = session.get(url, timeout=TIMEOUT)
            if r.status_code == 401 or r.status_code == 403:
                return None  # 被封或需要密码
            if r.status_code != 200:
                break
            data = r.json()
        except Exception:
            break

        batch = data.get("products", [])
        if not batch:
            break

        products.extend(batch)
        if len(batch) < 250:
            break
        if page == MAX_PRODUCT_PAGES:
            page_cap_hit = True

        page += 1
        time.sleep(DELAY)

    return {"products": products, "page_cap_hit": page_cap_hit}


def check_domain(domain, session):
    """检查单个域名的新品"""
    domain = normalize_domain(domain)
    old_snap = load_snapshot(domain)
    old_products = {str(k): v for k, v in (old_snap.get("products", {}) or {}).items()}

    fetch_result = fetch_products_fast(domain, session)
    if fetch_result is None:
        return {"domain": domain, "status": "blocked", "new_products": []}
    products = fetch_result.get("products", [])
    if not products:
        return {"domain": domain, "status": "empty", "new_products": []}

    current_products = {}
    existing_handles = set()
    for p in products:
        summary = product_summary_from_payload(p, domain, source="products_json")
        pid = summary["product_id"]
        if pid:
            current_products[pid] = summary
        if summary.get("handle"):
            existing_handles.add(summary["handle"].lower())

    sitemap_candidates = fetch_recent_sitemap_products(
        domain,
        session,
        existing_handles,
        limit=MAX_SITEMAP_CANDIDATES if fetch_result.get("page_cap_hit") else max(10, MAX_SITEMAP_CANDIDATES // 2),
    )
    for summary in sitemap_candidates:
        pid = summary["product_id"]
        if pid and pid not in current_products:
            current_products[pid] = summary

    changes = detect_catalog_changes(old_products, current_products)

    # 保存新快照
    save_snapshot(domain, current_products)

    # 如果是首次扫描（无旧快照）或从旧版 product_id-only 快照升级，不算新品/变化
    if not old_products or old_snap.get("legacy_only"):
        return {
            "domain": domain,
            "status": "first_scan",
            "total": len(current_products),
            "new_products": [],
            "baseline_upgraded": bool(old_snap.get("legacy_only")),
        }

    return {
        "domain": domain,
        "status": "ok",
        "total": len(current_products),
        "new_count": len(changes["new_products"]),
        "new_products": changes["new_products"],
        "changes": changes,
        "page_cap_hit": fetch_result.get("page_cap_hit", False),
        "sitemap_discovered": len(sitemap_candidates),
    }


def run_monitor(limit=0, max_workers=MAX_WORKERS, trend_workers=MAX_TREND_WORKERS, trend_geo="US", enable_trend=True):
    """主监控循环"""
    # 加载域名列表
    if not WATCHLIST_FILE.exists():
        print("❌ 无监控列表: competitors_watchlist.json")
        return

    watchlist = json.loads(WATCHLIST_FILE.read_text(encoding="utf-8"))
    domains = [item["domain"] for item in watchlist if isinstance(item, dict)]

    if limit > 0:
        domains = domains[:limit]

    print(f"🔍 监控 {len(domains)} 个独立站 (并发={max_workers})")
    print(f"   快照目录: {SNAPSHOT_DIR}")
    print(f"   输出目录: {OUTPUT_DIR}")
    print(f"   开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    all_new_products = []
    change_events = []
    stats = {"ok": 0, "first_scan": 0, "empty": 0, "blocked": 0, "error": 0}
    start_time = time.time()

    def process_domain(domain):
        session = get_session()
        try:
            return check_domain(domain, session)
        except Exception as e:
            return {"domain": domain, "status": "error", "error": str(e), "new_products": []}

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_domain, d): d for d in domains}
        done_count = 0

        for future in as_completed(futures):
            done_count += 1
            result = future.result()
            status = result.get("status", "error")
            stats[status] = stats.get(status, 0) + 1

            # 收集新品
            new_prods = result.get("new_products", [])
            if new_prods:
                all_new_products.extend(new_prods)
                print(f"  🆕 [{done_count}/{len(domains)}] {result['domain']}: {len(new_prods)} 新品!")

            changes = result.get("changes") or {}
            counts = changes.get("counts") or {}
            if any(counts.get(key, 0) for key in ("price", "image", "copy", "variant", "publish", "availability")):
                change_events.append({
                    "domain": result["domain"],
                    "counts": counts,
                    "page_cap_hit": result.get("page_cap_hit", False),
                    "sitemap_discovered": result.get("sitemap_discovered", 0),
                })
                print(
                    f"  🔄 [{done_count}/{len(domains)}] {result['domain']}: "
                    f"改价{counts.get('price',0)} 图变{counts.get('image',0)} 文案{counts.get('copy',0)} "
                    f"变体{counts.get('variant',0)} 上架{counts.get('publish',0)} 库存{counts.get('availability',0)}"
                )

            # 进度 (每50个打印)
            if done_count % 50 == 0:
                elapsed = time.time() - start_time
                rate = done_count / elapsed * 60
                eta = (len(domains) - done_count) / (done_count / elapsed)
                print(f"  ⏱️ 进度 {done_count}/{len(domains)} | "
                      f"{rate:.0f}站/分 | ETA {eta/60:.1f}分 | "
                      f"新品累计 {len(all_new_products)}")

    elapsed = time.time() - start_time

    if enable_trend and all_new_products:
        print(f"\n📈 自动趋势验证 {len(all_new_products)} 个新品 (geo={trend_geo}, workers={trend_workers})")
        cache = load_trend_cache()

        def process_trend(product):
            enriched = dict(product)
            enriched.update(verify_product_trend(product, trend_geo, cache))
            return enriched

        enriched_products = []
        with ThreadPoolExecutor(max_workers=max(1, trend_workers)) as executor:
            futures = [executor.submit(process_trend, product) for product in all_new_products]
            for idx, future in enumerate(as_completed(futures), start=1):
                enriched_products.append(future.result())
                if idx % 20 == 0 or idx == len(all_new_products):
                    print(f"  📈 趋势验证进度 {idx}/{len(all_new_products)}")

        save_trend_cache(cache)
        all_new_products = sorted(
            enriched_products,
            key=lambda x: (
                1 if x.get("trend_verified") else 0,
                x.get("trend_interest_avg", 0),
                x.get("trend_momentum_pct", 0),
            ),
            reverse=True,
        )

    if all_new_products:
        all_new_products = [enrich_product_actionability(product) for product in all_new_products]
        all_new_products = sorted(
            all_new_products,
            key=lambda x: (
                x.get("action_score", 0),
                1 if x.get("trend_verified") else 0,
                x.get("trend_interest_avg", 0),
            ),
            reverse=True,
        )

    # 输出统计
    print(f"\n{'='*60}")
    print(f"✅ 完成! 耗时 {elapsed/60:.1f} 分钟")
    print(f"   成功: {stats['ok']} | 首次扫描: {stats['first_scan']} | "
          f"空站: {stats['empty']} | 被封: {stats['blocked']} | 错误: {stats['error']}")
    print(f"   🆕 新品总数: {len(all_new_products)}")
    if change_events:
        print(f"   🔄 有关键变化站点: {len(change_events)}")

    trend_stats = trend_summary(all_new_products) if all_new_products else {"verified_count": 0, "status_breakdown": {}, "top_trending": []}
    if enable_trend and all_new_products:
        print(f"   📈 趋势已验证: {trend_stats['verified_count']}")
        print(f"   热门: {trend_stats['status_breakdown'].get('hot', 0)} | 观察: {trend_stats['status_breakdown'].get('watch', 0)} | "
              f"新兴: {trend_stats['status_breakdown'].get('emerging', 0)}")
        print("   🎯 今日优先抄作业:")
        for product in all_new_products[:5]:
            print(
                f"      - {product.get('title','')[:42]} | {product.get('domain','')} | "
                f"{product.get('action_label','')} | Score {product.get('action_score',0)} | "
                f"{product.get('primary_channel','')}"
            )

    # 写 Excel
    if all_new_products:
        write_excel(all_new_products)
    else:
        print("   无新品（首次运行需要先建立基线快照）")

    board = build_opportunity_board(all_new_products, change_events, limit=20)
    playbook_path = write_playbook(board)
    task_queues = build_task_queues(all_new_products, change_events, limit_per_queue=30)
    queue_json_path, queue_md_path = write_task_queues(task_queues)
    daily_checklist = build_daily_checklist(board, task_queues)
    daily_json_path, daily_md_path = write_daily_checklist(daily_checklist)

    report_path = REPORT_DIR / f"new_products_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    report_payload = {
        "generated_at": datetime.now().isoformat(),
        "watch_domains": len(domains),
        "stats": stats,
        "changed_domains": len(change_events),
        "change_events": change_events,
        "trend_geo": trend_geo,
        "trend_enabled": enable_trend,
        "trend_summary": trend_stats,
        "opportunity_board": board,
        "task_queues": task_queues,
        "daily_checklist": daily_checklist,
        "new_products": all_new_products,
    }
    report_path.write_text(json.dumps(report_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"   🧾 报告: {report_path}")
    print(f"   📌 作战清单: {playbook_path}")
    print(f"   📥 队列 JSON: {queue_json_path}")
    print(f"   📋 队列 Markdown: {queue_md_path}")
    print(f"   🗓️ Daily JSON: {daily_json_path}")
    print(f"   🗓️ Daily Markdown: {daily_md_path}")

    print("   🚀 今天先做:")
    for queue_name in ("ad_research_queue", "creative_queue", "landing_page_queue", "offer_test_queue", "research_queue"):
        items = task_queues.get(queue_name, [])
        if items:
            top = items[0]
            label = top.get("title") or top.get("domain") or "(unknown)"
            print(f"      - {queue_name}: {label} | {top.get('next_step','')}")

    return all_new_products


def write_excel(new_products):
    """新品写入 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("❌ 需要 openpyxl: pip3 install openpyxl")
        # fallback: 写 JSON
        out_path = OUTPUT_DIR / f"new_products_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(new_products, f, ensure_ascii=False, indent=2)
        print(f"   JSON 备份: {out_path}")
        return

    today = datetime.now().strftime("%Y%m%d")
    out_path = OUTPUT_DIR / f"new_products_{today}.xlsx"

    # 如果今天的文件已存在，追加
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        existing_rows = ws.max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "新品监控"
        existing_rows = 0

        # 写表头
        headers = [
            "域名", "产品名", "类型", "品牌", "最低价", "最高价",
            "变体数", "图片数", "产品链接", "主图", "创建时间", "发现时间",
            "趋势状态", "趋势验证", "趋势关键词", "趋势方向", "趋势均值", "趋势最新值", "趋势动量%", "相关搜索词",
            "动作评分", "动作标签", "主渠道", "Offer打法", "Meta品牌词", "Meta域名词", "Google广告库", "TikTok查询词"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 列宽
        widths = [25, 40, 15, 20, 10, 10, 8, 8, 50, 50, 20, 20, 12, 10, 28, 10, 10, 10, 12, 40, 10, 14, 22, 34, 40, 40, 40, 24]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

        existing_rows = 1

    # 写数据
    for p in new_products:
        row = existing_rows + 1
        existing_rows += 1
        ws.cell(row=row, column=1, value=p.get("domain", ""))
        ws.cell(row=row, column=2, value=p.get("title", ""))
        ws.cell(row=row, column=3, value=p.get("product_type", ""))
        ws.cell(row=row, column=4, value=p.get("vendor", ""))
        ws.cell(row=row, column=5, value=p.get("price_min", 0))
        ws.cell(row=row, column=6, value=p.get("price_max", 0))
        ws.cell(row=row, column=7, value=p.get("variants_count", 0))
        ws.cell(row=row, column=8, value=p.get("images_count", 0))
        ws.cell(row=row, column=9, value=p.get("url", ""))
        ws.cell(row=row, column=10, value=p.get("image_url", ""))
        ws.cell(row=row, column=11, value=p.get("created_at", "")[:19])
        ws.cell(row=row, column=12, value=p.get("detected_at", "")[:19])
        ws.cell(row=row, column=13, value=p.get("trend_status", ""))
        ws.cell(row=row, column=14, value="是" if p.get("trend_verified") else "否")
        ws.cell(row=row, column=15, value=p.get("trend_query", ""))
        ws.cell(row=row, column=16, value=p.get("trend_direction", ""))
        ws.cell(row=row, column=17, value=p.get("trend_interest_avg", 0))
        ws.cell(row=row, column=18, value=p.get("trend_interest_latest", 0))
        ws.cell(row=row, column=19, value=p.get("trend_momentum_pct", 0))
        ws.cell(row=row, column=20, value=" | ".join(p.get("trend_related_queries", [])))
        ws.cell(row=row, column=21, value=p.get("action_score", 0))
        ws.cell(row=row, column=22, value=p.get("action_label", ""))
        ws.cell(row=row, column=23, value=p.get("primary_channel", ""))
        ws.cell(row=row, column=24, value=p.get("offer_strategy", ""))
        ws.cell(row=row, column=25, value=p.get("meta_brand_url", ""))
        ws.cell(row=row, column=26, value=p.get("meta_domain_url", ""))
        ws.cell(row=row, column=27, value=p.get("google_ads_transparency_url", ""))
        ws.cell(row=row, column=28, value=p.get("tiktok_query", ""))

    wb.save(out_path)
    print(f"   📊 Excel: {out_path} ({existing_rows - 1} 行)")


# ─── CLI ───
if __name__ == "__main__":
    limit = 0
    workers = MAX_WORKERS
    trend_workers = MAX_TREND_WORKERS
    trend_geo = "US"
    enable_trend = True

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == "--limit" and i + 1 < len(sys.argv):
            limit = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == "--workers" and i + 1 < len(sys.argv):
            workers = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == "--trend-workers" and i + 1 < len(sys.argv):
            trend_workers = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == "--geo" and i + 1 < len(sys.argv):
            trend_geo = sys.argv[i + 1].upper()
            i += 2
        elif sys.argv[i] == "--no-trend":
            enable_trend = False
            i += 1
        else:
            i += 1

    run_monitor(limit=limit, max_workers=workers, trend_workers=trend_workers, trend_geo=trend_geo, enable_trend=enable_trend)
