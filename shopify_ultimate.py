#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 shopify_ultimate.py — Shopify 终极工具套件 v5.0 + OpenMythos
 合并自: shopify_competitor_monitor + shopify_discovery + shopify_manager
 功能: 竞品店铺监控 + 选品发现(Selofy) + 店铺管理(CLI+GraphQL) + 商业智能
 v4.0 新增: 数据质量验证 (title/price/url/image_url 必须非空且有效)
 v4.1 新增: OpenMythos RDT循环推理 + ACT自适应停止 (2026-05-01)
 v5.0 新增: 商业智能内核 — 定价策略引擎/竞品对比矩阵/利润计算器/市场分析器 (2026-05-13)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

CLI:
  === 竞品监控 (monitor) ===
  python3 shopify_ultimate.py scrape competitor1.com
  python3 shopify_ultimate.py check competitor1.com
  python3 shopify_ultimate.py auto competitor1.com
  python3 shopify_ultimate.py export competitor1.com

  === 选品发现 (discovery) ===
  python3 shopify_ultimate.py search "wireless charger"
  python3 shopify_ultimate.py search "wireless charger" --max 50 --export
  python3 shopify_ultimate.py batch "kw1,kw2" --export
  python3 shopify_ultimate.py inspire
  python3 shopify_ultimate.py verify "keyword"
  python3 shopify_ultimate.py verify-batch input.xlsx --top 20

  === 店铺管理 (manager) ===
  python3 shopify_ultimate.py auth <store>
  python3 shopify_ultimate.py info
  python3 shopify_ultimate.py products --limit 5
  python3 shopify_ultimate.py create-product --json data.json
  python3 shopify_ultimate.py create-product --from-xlsx report.xlsx --row 2
  python3 shopify_ultimate.py update-product <id> --json data.json
  python3 shopify_ultimate.py delete-product <id>
  python3 shopify_ultimate.py themes
  python3 shopify_ultimate.py theme-dev / theme-push / theme-pull
  python3 shopify_ultimate.py collections
  python3 shopify_ultimate.py create-collection --json data.json
  python3 shopify_ultimate.py bulk-import report.xlsx
"""

import argparse
import html as html_lib
import hashlib
import json
import os
import random
import re
import subprocess
import sys
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urljoin, quote
import xml.etree.ElementTree as ET

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# OpenMythos v3 Core
try:
    from openmythos_core import RecurrentDepthTransformer, AdaptiveExecutor
    HAS_MYTHOS = True
except ImportError:
    HAS_MYTHOS = False

# Mythos Guard v3 + Trajectory
try:
    _guard_path = str(Path.home() / '.hermes/skills/mythos-agent-orchestration')
    if _guard_path not in sys.path:
        sys.path.insert(0, _guard_path)
    from mythos_guard import api_guard, guarded_call
    from trajectory_logger import get_logger as get_trajectory, reset_logger
    HAS_GUARD = True
except ImportError:
    HAS_GUARD = False
    def api_guard(name, **kw):
        def d(fn): return fn
        return d
    def guarded_call(name, fn, *a, **kw): return fn(*a, **kw)
    def get_trajectory(tid, **kw): return None
    def reset_logger(tid, **kw): return None

# ─── 配置 ───
_HOME = Path(os.environ.get("HOME", "")).expanduser()
if not _HOME.exists() or not (_HOME / "Desktop" / "amazon").exists():
    _HOME = Path("/Users/doi")
BASE_DIR = _HOME / "Desktop/amazon"
DATA_DIR = BASE_DIR / "output/shopify_monitor"
DISCOVERY_OUTPUT_DIR = str(BASE_DIR / "output/shopify_discovery")
BLACKLIST_FILE = DATA_DIR / "non_shopify_blacklist.json"  # 非独立站黑名单
DELAY = 0.5
TIMEOUT = 10  # 基础超时，自适应会动态调整
MAX_DOMAIN_TIMEOUT = 30  # 单域名最大总耗时
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36"
BEST_EFFORT_PAGE_LIMIT = 12
BEST_EFFORT_SITEMAP_LIMIT = 6
BEST_EFFORT_PRODUCT_LIMIT = 80

# ─── v6.0 风控模块 ───
class RiskControl:
    """请求风控：UA 轮换、请求头随机化、速率控制、域名冷却"""

    UA_POOL = [
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.4 Safari/605.1.15",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:138.0) Gecko/20100101 Firefox/138.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:138.0) Gecko/20100101 Firefox/138.0",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36 Edg/134.0.0.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64; rv:138.0) Gecko/20100101 Firefox/138.0",
    ]

    ACCEPT_LANGS = [
        "en-US,en;q=0.9",
        "en-US,en;q=0.9,zh-CN;q=0.8",
        "en-GB,en;q=0.9",
        "en-US,en;q=0.8,fr;q=0.6",
        "en;q=0.9",
    ]

    def __init__(self):
        self._domain_last_req: dict[str, float] = {}
        self._global_last_req: float = 0
        self._domain_cooldown = 1.0  # 同域名最小间隔 1s
        self._global_rps = 10  # 每秒最多 10 请求
        self._lock = threading.Lock()

    def get_ua(self) -> str:
        return random.choice(self.UA_POOL)

    def get_headers(self) -> dict:
        return {
            "User-Agent": self.get_ua(),
            "Accept": "application/json,text/html,*/*",
            "Accept-Language": random.choice(self.ACCEPT_LANGS),
            "Connection": "keep-alive",
            "Cache-Control": "no-cache",
        }

    def wait_domain(self, domain: str):
        with self._lock:
            now = time.time()
            last = self._domain_last_req.get(domain, 0)
            wait = self._domain_cooldown - (now - last)
            if wait > 0:
                time.sleep(wait)
            self._domain_last_req[domain] = time.time()

    def wait_global(self):
        with self._lock:
            now = time.time()
            min_interval = 1.0 / self._global_rps
            wait = min_interval - (now - self._global_last_req)
            if wait > 0:
                time.sleep(wait)
            self._global_last_req = time.time()

    def before_request(self, domain: str):
        self.wait_global()
        self.wait_domain(domain)


_risk = RiskControl()


# ─── v6.0 域名健康评分 ───
_HEALTH_FILE = DATA_DIR / "domain_health.json"
_HEALTH_LOCK = threading.Lock()
_HEALTH_CACHE: dict = {}
_HEALTH_CACHE_TS: float = 0
_HEALTH_CACHE_TTL = 300  # 5 分钟缓存


def _load_health() -> dict:
    global _HEALTH_CACHE, _HEALTH_CACHE_TS
    now = time.time()
    if _HEALTH_CACHE and (now - _HEALTH_CACHE_TS) < _HEALTH_CACHE_TTL:
        return _HEALTH_CACHE
    if _HEALTH_FILE.exists():
        try:
            _HEALTH_CACHE = json.loads(_HEALTH_FILE.read_text(encoding="utf-8"))
        except:
            _HEALTH_CACHE = {}
    else:
        _HEALTH_CACHE = {}
    _HEALTH_CACHE_TS = now
    return _HEALTH_CACHE


def _save_health(health: dict):
    _HEALTH_FILE.parent.mkdir(parents=True, exist_ok=True)
    _HEALTH_FILE.write_text(json.dumps(health, ensure_ascii=False), encoding="utf-8")


def record_domain_result(domain: str, success: bool, response_ms: float = 0, error_type: str = ""):
    """记录域名请求结果，更新健康评分"""
    with _HEALTH_LOCK:
        health = _load_health()
        h = health.get(domain, {
            "success": 0, "fail": 0, "avg_ms": 0,
            "last_ok": None, "last_fail": None, "consec_fails": 0, "status": "active",
        })
        now = datetime.now().isoformat()
        if success:
            h["success"] = h.get("success", 0) + 1
            h["consec_fails"] = 0
            h["last_ok"] = now
            h["status"] = "active"
        else:
            h["fail"] = h.get("fail", 0) + 1
            h["consec_fails"] = h.get("consec_fails", 0) + 1
            h["last_fail"] = now
            if h["consec_fails"] >= 5:
                h["status"] = "dead"
            elif h["consec_fails"] >= 3:
                h["status"] = "protected"
        # 滚动平均响应时间
        if response_ms > 0:
            old_avg = h.get("avg_ms", 0)
            total = h.get("success", 0) + h.get("fail", 0)
            h["avg_ms"] = round(old_avg * 0.8 + response_ms * 0.2, 1) if old_avg > 0 else round(response_ms, 1)
        if error_type:
            h["last_error"] = error_type
        health[domain] = h
        _HEALTH_CACHE = health
        _save_health(health)


def get_domain_health(domain: str) -> dict:
    health = _load_health()
    return health.get(domain, {"status": "active", "consec_fails": 0, "avg_ms": 0})


def get_adaptive_timeout(domain: str) -> float:
    """根据域名历史响应时间返回自适应超时"""
    h = get_domain_health(domain)
    avg = h.get("avg_ms", 0)
    if avg > 8000:
        return 25.0
    if avg > 5000:
        return 20.0
    if avg > 2000:
        return 15.0
    return TIMEOUT


# ─── v6.0 黑名单缓存 ───
_BL_CACHE: dict = {}
_BL_CACHE_TS: float = 0
_BL_CACHE_TTL = 300
_BL_PENDING: list = []  # 待批量写入
_BL_BATCH_SIZE = 50
_BLACKLIST_LOCK = threading.Lock()


def load_blacklist():
    """加载非独立站黑名单（带内存缓存）"""
    global _BL_CACHE, _BL_CACHE_TS
    now = time.time()
    if _BL_CACHE and (now - _BL_CACHE_TS) < _BL_CACHE_TTL:
        return _BL_CACHE
    if not BLACKLIST_FILE.exists():
        _BL_CACHE = {}
    else:
        try:
            _BL_CACHE = json.loads(BLACKLIST_FILE.read_text(encoding="utf-8"))
        except:
            _BL_CACHE = {}
    _BL_CACHE_TS = now
    return _BL_CACHE


def save_to_blacklist(domain, platform, reason=""):
    """将非 Shopify 域名加入黑名单（批量缓冲写入）"""
    global _BL_CACHE, _BL_PENDING
    with _BLACKLIST_LOCK:
        if domain in _BL_CACHE:
            return
        _BL_CACHE[domain] = {
            "platform": platform,
            "reason": reason,
            "detected_at": datetime.now().isoformat()
        }
        _BL_PENDING.append(domain)
        if len(_BL_PENDING) >= _BL_BATCH_SIZE:
            _flush_blacklist()
    print(f"  🚫 已加入黑名单: {domain} ({platform})")


def _flush_blacklist():
    """将内存黑名单写入磁盘"""
    global _BL_PENDING
    if not _BL_PENDING:
        return
    BLACKLIST_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(BLACKLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(_BL_CACHE, f, ensure_ascii=False)
    _BL_PENDING = []


def flush_blacklist_on_complete():
    """刷新结束时调用，确保所有待写入数据落盘"""
    with _BLACKLIST_LOCK:
        _flush_blacklist()


def cleanup_stale_blacklist(max_age_days=90):
    """清理超过 max_age_days 的黑名单条目"""
    with _BLACKLIST_LOCK:
        bl = load_blacklist()
        cutoff = (datetime.now() - timedelta(days=max_age_days)).isoformat()
        stale = [d for d, v in bl.items() if v.get("detected_at", "") < cutoff]
        for d in stale:
            del bl[d]
        if stale:
            _BL_CACHE = bl
            _flush_blacklist()
            print(f"  🧹 清理 {len(stale)} 条过期黑名单（>{max_age_days}天）")
    return len(stale)


def is_blacklisted(domain):
    """检查域名是否在黑名单中（纯内存查询）"""
    bl = load_blacklist()
    return domain in bl


# 品类黑名单 — 自动过滤的产品类型（只保留独立站卖实体产品）
CATEGORY_BLACKLIST = [
    # 招聘/移民服务
    "recruitment", "hiring", "job", "career", "immigration", "visa", "work permit",
    "employment", "staffing", "hr service", "talent acquisition",
    
    # 商业服务（可持续环保）
    "sustainability consulting", "carbon offset", "esg", "green business",
    "environmental consulting", "carbon credit", "climate solution",
    "renewable energy consulting", "eco certification",
    
    # 汽车配件/改装（科技配件）
    "car accessories", "auto parts", "vehicle modification", "car tech",
    "automotive electronics", "car gadget", "dash cam", "car mount",
    "car charger", "obd", "tire pressure", "car diagnostic",
    
    # 烘焙
    "baking", "bakery", "pastry", "cake decorating", "bread making",
    "baking supplies", "baking tools", "cake pan", "cookie cutter",
    
    # 水果/美食/健身
    "fruit", "fruits", "fresh fruit", "organic fruit", "fruit box",
    "food", "gourmet", "meal", "cuisine", "recipe", "cooking",
    "fitness", "gym", "workout", "exercise", "training", "bodybuilding",
    "protein", "supplement", "nutrition", "diet", "weight loss",
    
    # 在线教育
    "online course", "e-learning", "online education", "online training",
    "webinar", "online class", "educational platform", "learning management",
    "tutoring service", "certification course",
    
    # 票务/演出
    "ticket", "ticketing", "event ticket", "concert ticket", "show ticket",
    "box office", "event booking", "live performance", "theater ticket",
    
    # 物流
    "logistics", "shipping service", "freight", "courier", "delivery service",
    "warehousing", "fulfillment service", "cargo", "express delivery",
    
    # 服务类（非实体产品）
    "consulting", "service", "subscription", "membership", "saas", "software",
    "app download", "digital service", "platform", "marketplace"
]

# Selofy API config (discovery)
SELOFY_BASE_URL = "https://www.selofy.com/api/shopifier"
SELOFY_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    "Referer": "https://www.selofy.com/zh/tools/shopifier",
    "Accept": "application/json",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}

# Shopify CLI config (manager)
SHOPIFY_CLI = os.path.expanduser("~/.npm-global/bin/shopify")
CONFIG_FILE = str(BASE_DIR / ".shopify_config.json")

# v5.0 竞品监控列表 (从 competitors_watchlist.json 动态加载)
_WATCHLIST_FILE = DATA_DIR / "competitors_watchlist.json"
def _load_watchlist():
    if _WATCHLIST_FILE.exists():
        try:
            import json as _j
            data = _j.loads(_WATCHLIST_FILE.read_text(encoding="utf-8"))
            return [item["domain"] for item in data if isinstance(item, dict)]
        except:
            pass
    return []

DEFAULT_COMPETITORS = _load_watchlist()


# ================================================================
#  Section 1: Competitor Monitor (from shopify_competitor_monitor)
# ================================================================

def is_category_blacklisted(product):
    """检查产品是否属于黑名单品类"""
    # 提取产品文本信息
    title = product.get("title", "").lower()
    product_type = product.get("product_type", "").lower()
    vendor = product.get("vendor", "").lower()
    tags = " ".join(product.get("tags", [])).lower()
    
    # 合并所有文本
    text = f"{title} {product_type} {vendor} {tags}"
    
    # 检查是否命中黑名单关键词
    for keyword in CATEGORY_BLACKLIST:
        if keyword.lower() in text:
            return True, keyword
    
    return False, None

_PLATFORM_CACHE: dict = {}
_PLATFORM_CACHE_LOCK = threading.Lock()


def detect_platform(domain, session=None, fast=True):
    """检测电商平台类型: shopify / lightfunnels / magento / woocommerce / unknown
    fast=True 时跳过 HTML 检测（默认），只试 products.json
    """
    # 缓存命中
    with _PLATFORM_CACHE_LOCK:
        if domain in _PLATFORM_CACHE:
            return _PLATFORM_CACHE[domain]

    import urllib.request, ssl
    base = get_base_url(domain)

    # 快速路径：直接试 products.json（3s，跳过 SSL 验证）
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        # 不发送 Accept-Encoding，避免 gzip 解压问题
        headers = _risk.get_headers()
        headers.pop("Accept-Encoding", None)
        req = urllib.request.Request(f"{base}/products.json?limit=1", headers=headers)
        with urllib.request.urlopen(req, timeout=3, context=ctx) as rj:
            if rj.status == 200:
                data = json.loads(rj.read())
                if isinstance(data, dict) and "products" in data:
                    with _PLATFORM_CACHE_LOCK:
                        _PLATFORM_CACHE[domain] = "shopify"
                    return "shopify"
    except:
        pass

    # fast 模式跳过 HTML 检测
    if fast:
        with _PLATFORM_CACHE_LOCK:
            _PLATFORM_CACHE[domain] = "unknown"
        return "unknown"

    # 慢速路径：首页 HTML 检测（仅 fast=False 时）
    if session is None:
        session = get_session()
    try:
        r = session.get(base, timeout=8, allow_redirects=True)
        html = r.text.lower()
        headers_str = str(r.headers).lower()

        if "lightfunnels" in html or "lightfunnels" in headers_str:
            result = "lightfunnels"
        elif "woocommerce" in html or "wp-content/plugins/woocommerce" in html:
            result = "woocommerce"
        elif any(x in html for x in ["magento", "mage-", "catalogsearch", "checkout/cart"]):
            result = "magento"
        elif "x-magento" in headers_str or "mage-cache-storage" in html:
            result = "magento"
        elif "cdn.shopify.com" in html or "shopify" in headers_str:
            result = "shopify"
        else:
            result = "unknown"
    except Exception as e:
        print(f"  ⚠️ 平台检测失败: {e}")
        result = "unknown"

    with _PLATFORM_CACHE_LOCK:
        _PLATFORM_CACHE[domain] = result
    return result


def fetch_products_bb_browser(domain):
    """用 bb-browser（真实Chrome）抓取非 Shopify 店铺的产品"""
    import subprocess, json as _json

    base = get_base_url(domain)
    products = []

    # 打开店铺页面
    import time as _time
    try:
        subprocess.run(["bb-browser", "open", base], capture_output=True, timeout=20)
        _time.sleep(3)  # 等页面加载完成
    except Exception as e:
        print(f"  ❌ bb-browser 打开失败: {e}")
        return products

    # 多种 JS 提取策略，适配不同平台
    js_extract = """
    (() => {
        const products = [];

        // 策略1: JSON-LD 结构化数据
        document.querySelectorAll('script[type="application/ld+json"]').forEach(s => {
            try {
                const d = JSON.parse(s.textContent);
                const items = d['@type'] === 'ItemList' ? d.itemListElement : (Array.isArray(d) ? d : [d]);
                items.forEach(item => {
                    const p = item.item || item;
                    if (p['@type'] === 'Product' && p.name) {
                        const offers = p.offers || {};
                        const offer = Array.isArray(offers) ? offers[0] : offers;
                        products.push({
                            title: p.name,
                            handle: (p.url || '').split('/').pop() || '',
                            product_type: p.category || '',
                            vendor: p.brand?.name || p.brand || '',
                            tags: [],
                            body_html: (p.description || '').substring(0, 500),
                            variants: [{title: 'Default', price: String(offer.price || offer.lowPrice || '0'), sku: '', compare_at_price: '', grams: 0, requires_shipping: true, taxable: true, available: true}],
                            images: (Array.isArray(p.image) ? p.image : [p.image || '']).filter(Boolean).map(src => ({src: typeof src === 'string' ? src : src.url || ''})),
                            published_at: new Date().toISOString(),
                            created_at: new Date().toISOString(),
                            updated_at: new Date().toISOString(),
                        });
                    }
                });
            } catch(e) {}
        });

        if (products.length > 0) return JSON.stringify(products);

        // 策略2: 通用产品卡片抓取
        const cards = document.querySelectorAll('.product-card, .product-item, .product, [class*="product"], .card, .item-product, .grid-item, .collection-item');
        const seen = new Set();
        cards.forEach(card => {
            const link = card.querySelector('a[href*="/products/"], a[href*="/product/"]');
            const img = card.querySelector('img');
            const title = card.querySelector('.product-title, .product-name, h2, h3, h4, .title, [class*="title"], [class*="name"]');
            const price = card.querySelector('.price, .product-price, [class*="price"], .money, .amount');

            if (!link && !title) return;

            const titleText = (title?.textContent || link?.textContent || '').trim();
            if (!titleText || seen.has(titleText)) return;
            seen.add(titleText);

            const href = link?.href || '';
            const imgSrc = img?.src || img?.dataset?.src || '';
            let priceVal = '0';
            if (price) {
                const m = price.textContent.match(/[\\d,.]+/);
                if (m) priceVal = m[0].replace(',', '');
            }

            products.push({
                title: titleText.substring(0, 200),
                handle: href.split('/').pop()?.split('?')[0] || '',
                product_type: '',
                vendor: '',
                tags: [],
                body_html: '',
                variants: [{title: 'Default', price: priceVal, sku: '', compare_at_price: '', grams: 0, requires_shipping: true, taxable: true, available: true}],
                images: imgSrc ? [{src: imgSrc}] : [],
                published_at: new Date().toISOString(),
                created_at: new Date().toISOString(),
                updated_at: new Date().toISOString(),
            });
        });

        if (products.length > 0) return JSON.stringify(products);

        // 策略3: 所有产品链接
        const links = document.querySelectorAll('a[href*="/products/"], a[href*="/product/"]');
        const seenLinks = new Set();
        links.forEach(a => {
            const href = a.href;
            if (seenLinks.has(href)) return;
            seenLinks.add(href);
            const title = a.textContent?.trim() || a.querySelector('img')?.alt || '';
            if (!title || title.length < 3) return;
            const img = a.querySelector('img');
            products.push({
                title: title.substring(0, 200),
                handle: href.split('/').pop()?.split('?')[0] || '',
                product_type: '', vendor: '', tags: [], body_html: '',
                variants: [{title: 'Default', price: '0', sku: '', compare_at_price: '', grams: 0, requires_shipping: true, taxable: true, available: true}],
                images: img?.src ? [{src: img.src}] : [],
                published_at: new Date().toISOString(), created_at: new Date().toISOString(), updated_at: new Date().toISOString(),
            });
        });

        // 策略4: 单产品落地页（Lightfunnels / 单品商城）
        if (products.length === 0) {
            // 标题: document.title 或 h1
            let title = document.title || '';
            // 只按 | 分割（不按 - 分割，避免截断 "3-in-1" 等产品名）
            title = title.split('|')[0].trim()
                .replace(/[\\u{1F300}-\\u{1F9FF}]/gu, '')
                .replace(/BEST PRICE|FREE GIFTS?|TODAY ONLY|FLASH SALE|UP TO \\d+% OFF/gi, '')
                .trim();

            // 如果 title 太短或太营销化，尝试 h1
            if (title.length < 5) {
                const h1 = document.querySelector('h1');
                if (h1) title = h1.textContent.trim();
            }

            // 图片: 包含 lazy-loaded（不检查 naturalWidth）
            const productImgs = [];
            const seenSrcs = new Set();
            document.querySelectorAll('img').forEach(img => {
                const src = img.src || img.dataset?.src || img.getAttribute('data-lazy-src') || '';
                if (!src || src.length < 20) return;
                // 排除追踪/像素/logo 图
                if (/pixel|tracking|analytics|spacer|blank|logo|icon|favicon/i.test(src)) return;
                // 排除极小图（通过 URL 判断尺寸）
                const sizeMatch = src.match(/width=(\\d+)/);
                if (sizeMatch && parseInt(sizeMatch[1]) < 100) return;
                if (!seenSrcs.has(src)) {
                    seenSrcs.add(src);
                    productImgs.push(src);
                }
            });

            // 价格: 从全文匹配
            let price = '0';
            const bodyText = document.body?.innerText || '';
            const priceMatches = bodyText.match(/\\$\\d+\\.?\\d*/g) || [];
            if (priceMatches.length > 0) {
                // 取最常见价格或最低价格
                const prices = priceMatches.map(p => parseFloat(p.replace('$',''))).filter(p => p > 5 && p < 5000);
                if (prices.length > 0) price = String(Math.min(...prices));
            }

            // 也从结构化数据取价格
            document.querySelectorAll('script[type="application/ld+json"]').forEach(s => {
                try {
                    const d = JSON.parse(s.textContent);
                    const offer = d.offers || (Array.isArray(d['@graph']) ? d['@graph'].find(i=>i.offers)?.offers : null);
                    if (offer?.price && parseFloat(offer.price) > 0) price = String(offer.price);
                    if (!title && d.name) title = d.name;
                } catch(e){}
            });

            // 描述
            const desc = document.querySelector('[class*="description"], .product-description, [class*="detail"], .wys')?.textContent?.trim() || '';

            if (title && title.length > 3) {
                products.push({
                    title: title.substring(0, 200),
                    handle: window.location.pathname.split('/').pop()?.replace(/^\\//,'') || 'main',
                    product_type: '',
                    vendor: window.location.hostname,
                    tags: [],
                    body_html: desc.substring(0, 500),
                    variants: [{title: 'Default', price: price, sku: '', compare_at_price: '', grams: 0, requires_shipping: true, taxable: true, available: true}],
                    images: productImgs.slice(0, 8).map(src => ({src})),
                    published_at: new Date().toISOString(),
                    created_at: new Date().toISOString(),
                    updated_at: new Date().toISOString(),
                });
            }
        }

        return JSON.stringify(products);
    })()
    """

    try:
        result = subprocess.run(
            ["bb-browser", "eval", js_extract],
            capture_output=True, text=True, timeout=30
        )
        raw = result.stdout.strip()
        if raw and raw.startswith('['):
            products = _json.loads(raw)
            print(f"  🌐 bb-browser: {len(products)} 个产品")
    except Exception as e:
        print(f"  ❌ bb-browser 提取失败: {e}")

    # 滚动加载更多
    if products:
        for _ in range(3):
            try:
                subprocess.run(
                    ["bb-browser", "eval", "window.scrollTo(0, document.body.scrollHeight); 'scrolled'"],
                    capture_output=True, timeout=10
                )
                _time.sleep(2)
                result = subprocess.run(
                    ["bb-browser", "eval", js_extract],
                    capture_output=True, text=True, timeout=30
                )
                raw = result.stdout.strip()
                if raw and raw.startswith('['):
                    new_products = _json.loads(raw)
                    existing_titles = {p['title'] for p in products}
                    for p in new_products:
                        if p['title'] not in existing_titles:
                            products.append(p)
                            existing_titles.add(p['title'])
            except:
                break

    return products


def get_session():
    """带重试的 requests session"""
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "application/json",
    })
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def normalize_domain(domain):
    """标准化域名"""
    domain = domain.strip().rstrip("/")
    domain = re.sub(r"^https?://", "", domain)
    domain = re.sub(r"^www\.", "", domain)
    return domain


def get_base_url(domain):
    return f"https://{domain}"


# ─── v4.0 数据质量验证 ───
def validate_product_quality(product, domain, session=None):
    """
    v4.0 数据质量门：验证产品字段完整性和有效性
    返回: (is_valid: bool, errors: list)
    """
    errors = []
    
    # 1. title 必须非空
    title = product.get("title", "").strip()
    if not title:
        errors.append("title为空")
    
    # 2. price > 0
    variants = product.get("variants", [])
    if not variants:
        errors.append("无variants")
    else:
        prices = [float(v.get("price", 0)) for v in variants if v.get("price")]
        if not prices or all(p <= 0 for p in prices):
            errors.append("price<=0")
    
    # 3. url 格式完整
    handle = product.get("handle", "").strip()
    if not handle:
        errors.append("handle为空(无法生成url)")
    else:
        url = f"{get_base_url(domain)}/products/{handle}"
        if not url.startswith("https://"):
            errors.append("url格式无效")
    
    # 4. image_url 可访问
    images = product.get("images", [])
    if not images:
        errors.append("无images")
    else:
        image_url = images[0].get("src", "").strip()
        if not image_url:
            errors.append("image_url为空")
        elif not image_url.startswith(("http://", "https://")):
            errors.append("image_url格式无效")
        # else:
        #     # 快速HEAD检查图片可访问性 (DISABLED: 太慢，导致监控超时)
        #     try:
        #         if session is None:
        #             session = get_session()
        #         r = session.head(image_url, timeout=5, allow_redirects=True)
        #         if r.status_code >= 400:
        #             errors.append(f"image_url不可访问(HTTP {r.status_code})")
        #     except Exception as e:
        #         errors.append(f"image_url不可访问({str(e)[:50]})")
    
    return len(errors) == 0, errors


def save_rejected_products(domain, rejected_products):
    """保存被拒绝的产品到 rejected_products.json"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    rejected_file = DATA_DIR / f"{normalize_domain(domain)}_rejected_products.json"
    
    with open(rejected_file, "w", encoding="utf-8") as f:
        json.dump(rejected_products, f, ensure_ascii=False, indent=2)
    
    return rejected_file


# ─── v6.0 核心抓取（自适应超时 + 智能重试 + 风控）───

def _classify_error(exc: Exception) -> str:
    """将异常分类为错误类型"""
    s = str(exc).lower()
    if isinstance(exc, requests.exceptions.SSLError) or "ssl" in s:
        return "ssl"
    if isinstance(exc, requests.exceptions.Timeout) or "timeout" in s or "timed out" in s:
        return "timeout"
    if isinstance(exc, requests.exceptions.ConnectionError) or "connection" in s:
        return "network"
    if isinstance(exc, json.JSONDecodeError) or "expecting value" in s or "json" in s:
        return "json_error"
    if hasattr(exc, 'response') and exc.response is not None:
        code = exc.response.status_code
        if code == 403:
            return "403"
        if code == 404:
            return "404"
        if code == 429:
            return "429"
        if code >= 500:
            return "5xx"
    if "403" in s:
        return "403"
    if "404" in s:
        return "404"
    if "429" in s:
        return "429"
    return "unknown"


def _fetch_shopify_products(domain, session):
    """Shopify JSON API 分页抓取（v6.0: 自适应超时 + 智能重试 + 风控）"""
    base = get_base_url(domain)
    products = []
    page = 1
    domain_timeout = get_adaptive_timeout(domain)
    domain_start = time.time()

    while True:
        # 单域名总耗时保护
        elapsed = (time.time() - domain_start) * 1000
        if elapsed > MAX_DOMAIN_TIMEOUT * 1000:
            print(f"  ⏱️ 域名总耗时 {elapsed/1000:.1f}s 超限，停止")
            break

        url = f"{base}/products.json?limit=250&page={page}"

        # 智能重试：指数退避
        max_retries = 3
        last_error = None
        last_error_type = None
        for attempt in range(max_retries):
            try:
                _risk.before_request(domain)
                headers = _risk.get_headers()
                r = session.get(url, timeout=domain_timeout, headers=headers)
                r.raise_for_status()
                data = r.json()
                last_error = None
                break
            except requests.exceptions.SSLError as e:
                # SSL 错误快速失败，不重试
                last_error = e
                last_error_type = "ssl"
                break
            except requests.exceptions.HTTPError as e:
                if hasattr(e, 'response') and e.response is not None:
                    code = e.response.status_code
                    if code == 403:
                        # 403 快速失败，不重试
                        last_error = e
                        last_error_type = "403"
                        break
                    if code == 404:
                        last_error = e
                        last_error_type = "404"
                        break
                    if code == 429:
                        # 429 指数退避
                        wait = 2 ** attempt
                        time.sleep(wait)
                        last_error = e
                        last_error_type = "429"
                        continue
                last_error = e
                last_error_type = _classify_error(e)
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
            except requests.exceptions.Timeout as e:
                last_error = e
                last_error_type = "timeout"
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
            except Exception as e:
                last_error = e
                last_error_type = _classify_error(e)
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)

        if last_error:
            print(f"  ❌ 第{page}页失败 [{last_error_type}]: {last_error}")
            record_domain_result(domain, False, error_type=last_error_type)
            break

        batch = data.get("products", [])
        if not batch:
            break

        products.extend(batch)
        if page == 1:
            print(f"  📦 第{page}页: {len(batch)} 个产品")

        if len(batch) < 250:
            break

        page += 1
        time.sleep(DELAY)

    return products


def _safe_float(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        if isinstance(value, str):
            cleaned = re.sub(r"[^0-9.\-]", "", value.replace(",", ""))
            if cleaned in ("", ".", "-", "-."):
                return default
            return float(cleaned)
        return float(value)
    except (TypeError, ValueError):
        return default


def _clean_text(value, limit=240):
    text = html_lib.unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def _absolute_url(base_url, url):
    if not url:
        return ""
    return urljoin(base_url.rstrip("/") + "/", str(url).strip())


def _slug_from_url(url):
    path = re.sub(r"^https?://[^/]+", "", str(url or "")).split("?")[0].strip("/")
    slug = path.rstrip("/").split("/")[-1] if path else "main"
    slug = re.sub(r"[^a-zA-Z0-9._-]+", "-", slug).strip("-").lower()
    return slug or "main"


def _stable_product_id(domain, url, title=""):
    raw = f"{normalize_domain(domain)}|{url}|{title}"
    return "best_effort_" + hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()[:16]


def _meta_content(html_text, *names):
    if not html_text:
        return ""
    for name in names:
        pattern = (
            r'<meta\s+[^>]*(?:property|name)=["\']'
            + re.escape(name)
            + r'["\'][^>]*content=["\']([^"\']+)["\'][^>]*>'
        )
        match = re.search(pattern, html_text, re.I)
        if match:
            return html_lib.unescape(match.group(1)).strip()
        pattern = (
            r'<meta\s+[^>]*content=["\']([^"\']+)["\'][^>]*(?:property|name)=["\']'
            + re.escape(name)
            + r'["\'][^>]*>'
        )
        match = re.search(pattern, html_text, re.I)
        if match:
            return html_lib.unescape(match.group(1)).strip()
    return ""


def _title_from_html(html_text):
    for pattern in [r"<h1[^>]*>(.*?)</h1>", r"<title[^>]*>(.*?)</title>"]:
        match = re.search(pattern, html_text or "", re.I | re.S)
        if match:
            title = _clean_text(match.group(1), limit=180)
            if title:
                return re.split(r"\s+[|]\s+", title)[0].strip()
    return ""


def _first_price_from_text(text):
    if not text:
        return 0.0
    prices = []
    for match in re.findall(r"(?:US\$|\$|USD\s*)\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)", text, re.I):
        value = _safe_float(match)
        if 1 <= value <= 10000:
            prices.append(value)
    return min(prices) if prices else 0.0


def _jsonld_objects(value):
    if isinstance(value, list):
        for item in value:
            yield from _jsonld_objects(item)
    elif isinstance(value, dict):
        yield value
        for key in ("@graph", "itemListElement"):
            child = value.get(key)
            if child:
                yield from _jsonld_objects(child)
        item = value.get("item")
        if isinstance(item, dict):
            yield from _jsonld_objects(item)


def _jsonld_type_matches(node, wanted="Product"):
    raw_type = node.get("@type") if isinstance(node, dict) else None
    if isinstance(raw_type, list):
        return any(str(t).lower() == wanted.lower() for t in raw_type)
    return str(raw_type or "").lower() == wanted.lower()


def _iter_jsonld_products(html_text):
    for raw in re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', html_text or "", re.I | re.S):
        try:
            data = json.loads(html_lib.unescape(raw.strip()))
        except Exception:
            continue
        for node in _jsonld_objects(data):
            if isinstance(node, dict) and _jsonld_type_matches(node) and node.get("name"):
                yield node


def _offer_price(offers):
    if isinstance(offers, list):
        for offer in offers:
            price = _offer_price(offer)
            if price:
                return price
        return 0.0
    if isinstance(offers, dict):
        for key in ("price", "lowPrice", "highPrice"):
            price = _safe_float(offers.get(key))
            if price:
                return price
    return 0.0


def _image_list_from_value(value, base_url=""):
    if not value:
        return []
    values = value if isinstance(value, list) else [value]
    images = []
    seen = set()
    for item in values:
        if isinstance(item, dict):
            src = item.get("url") or item.get("src") or item.get("contentUrl")
        else:
            src = str(item)
        src = _absolute_url(base_url, src)
        if src and src not in seen:
            images.append({"src": src})
            seen.add(src)
    return images


def _product_from_page(domain, page_url, html_text, source):
    base_url = page_url if page_url.startswith(("http://", "https://")) else get_base_url(domain)
    for node in _iter_jsonld_products(html_text):
        title = _clean_text(node.get("name"), limit=200)
        if not title:
            continue
        url = _absolute_url(base_url, node.get("url") or page_url)
        price = _offer_price(node.get("offers"))
        images = _image_list_from_value(node.get("image"), base_url=base_url)
        brand = node.get("brand")
        if isinstance(brand, dict):
            vendor = brand.get("name", "")
        else:
            vendor = str(brand or "")
        return {
            "id": _stable_product_id(domain, url, title),
            "title": title,
            "handle": _slug_from_url(url),
            "product_type": _clean_text(node.get("category"), limit=120),
            "vendor": _clean_text(vendor, limit=120) or normalize_domain(domain),
            "tags": ["best_effort", source],
            "body_html": _clean_text(node.get("description"), limit=500),
            "variants": [{
                "title": "Default",
                "price": f"{price:.2f}" if price else "0",
                "sku": "",
                "compare_at_price": "",
                "grams": 0,
                "requires_shipping": True,
                "taxable": True,
                "available": True,
            }],
            "images": images,
            "published_at": datetime.now().isoformat(),
            "created_at": "",
            "updated_at": datetime.now().isoformat(),
            "url": url,
            "source": source,
            "platform": "best_effort",
            "_best_effort": True,
        }

    title = (
        _meta_content(html_text, "og:title", "twitter:title")
        or _title_from_html(html_text)
    )
    title = _clean_text(re.sub(r"\s+[|-]\s+.*$", "", title or ""), limit=200)
    if not title or len(title) < 3:
        return None
    url = _absolute_url(base_url, _meta_content(html_text, "og:url") or page_url)
    image = _absolute_url(base_url, _meta_content(html_text, "og:image", "twitter:image"))
    price = _safe_float(
        _meta_content(html_text, "product:price:amount", "og:price:amount", "twitter:data1")
    )
    if not price:
        price = _first_price_from_text(html_text[:120000])
    description = _meta_content(html_text, "og:description", "description", "twitter:description")
    images = [{"src": image}] if image else []
    return {
        "id": _stable_product_id(domain, url, title),
        "title": title,
        "handle": _slug_from_url(url),
        "product_type": "",
        "vendor": normalize_domain(domain),
        "tags": ["best_effort", source],
        "body_html": _clean_text(description, limit=500),
        "variants": [{
            "title": "Default",
            "price": f"{price:.2f}" if price else "0",
            "sku": "",
            "compare_at_price": "",
            "grams": 0,
            "requires_shipping": True,
            "taxable": True,
            "available": True,
        }],
        "images": images,
        "published_at": datetime.now().isoformat(),
        "created_at": "",
        "updated_at": datetime.now().isoformat(),
        "url": url,
        "source": source,
        "platform": "best_effort",
        "_best_effort": True,
    }


def _parse_sitemap_locs(xml_text):
    if not xml_text:
        return []
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []
    items = []
    for node in root.findall(".//{*}sitemap") + root.findall(".//{*}url"):
        loc = node.findtext("{*}loc", default="").strip()
        lastmod = node.findtext("{*}lastmod", default="").strip()
        if loc:
            items.append({"loc": loc, "lastmod": lastmod})
    return items


def _looks_like_product_url(url):
    path = re.sub(r"^https?://[^/]+", "", str(url or "").lower()).split("?")[0]
    if not path or any(x in path for x in ["/cart", "/checkout", "/account", "/blog", "/pages/", "/policy", "/privacy"]):
        return False
    product_markers = ["/products/", "/product/", "/shop/", "/item/", "/p/"]
    if any(marker in path for marker in product_markers):
        return True
    parts = [p for p in path.strip("/").split("/") if p]
    return len(parts) == 1 and bool(re.search(r"[a-z].*-[a-z]", parts[0]))


def _get_text(session, url, domain, timeout=TIMEOUT):
    try:
        _risk.before_request(domain)
        headers = _risk.get_headers()
        headers["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
        r = session.get(url, timeout=timeout, headers=headers, allow_redirects=True)
        if r.status_code != 200:
            return ""
        return r.text or ""
    except Exception:
        return ""


def _discover_product_urls(domain, session):
    base = get_base_url(domain)
    sitemap_urls = [
        f"{base}/sitemap.xml",
        f"{base}/sitemap_products_1.xml",
        f"{base}/product-sitemap.xml",
        f"{base}/product-sitemap1.xml",
    ]
    product_urls = []
    seen_sitemaps = set()
    seen_urls = set()

    for sitemap_url in sitemap_urls:
        if sitemap_url in seen_sitemaps or len(seen_sitemaps) >= BEST_EFFORT_SITEMAP_LIMIT:
            continue
        seen_sitemaps.add(sitemap_url)
        xml_text = _get_text(session, sitemap_url, domain)
        for item in _parse_sitemap_locs(xml_text):
            loc = item["loc"]
            if loc in seen_sitemaps:
                continue
            if "sitemap" in loc.lower() and any(word in loc.lower() for word in ["product", "shop", "item"]):
                seen_sitemaps.add(loc)
                child_xml = _get_text(session, loc, domain)
                for child in _parse_sitemap_locs(child_xml):
                    child_loc = child["loc"]
                    if _looks_like_product_url(child_loc) and child_loc not in seen_urls:
                        product_urls.append(child_loc)
                        seen_urls.add(child_loc)
            elif _looks_like_product_url(loc) and loc not in seen_urls:
                product_urls.append(loc)
                seen_urls.add(loc)
            if len(product_urls) >= BEST_EFFORT_PRODUCT_LIMIT:
                return product_urls

    return product_urls[:BEST_EFFORT_PRODUCT_LIMIT]


def _discover_home_product_links(domain, session):
    base = get_base_url(domain)
    html_text = _get_text(session, base, domain)
    urls = []
    seen = set()
    for href in re.findall(r'<a\s+[^>]*href=["\']([^"\']+)["\']', html_text or "", re.I):
        url = _absolute_url(base, href)
        if _looks_like_product_url(url) and url not in seen:
            urls.append(url)
            seen.add(url)
        if len(urls) >= BEST_EFFORT_PAGE_LIMIT:
            break
    if not urls and html_text:
        product = _product_from_page(domain, base, html_text, "homepage")
        return [], [product] if product else []
    return urls, []


def _dedupe_products(products):
    deduped = []
    seen = set()
    for product in products:
        if not product or not product.get("title"):
            continue
        key = str(product.get("id") or product.get("url") or product.get("handle") or product.get("title")).lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(product)
    return deduped


def fetch_best_effort_products(domain, session=None, platform="unknown", with_browser=False):
    """Best-effort scanner for any ecommerce-like site, including non-Shopify stores."""
    if session is None:
        session = get_session()
    domain = normalize_domain(domain)
    products = []

    urls = _discover_product_urls(domain, session)
    if not urls:
        home_urls, home_products = _discover_home_product_links(domain, session)
        urls = home_urls
        products.extend([p for p in home_products if p])

    for url in urls[:BEST_EFFORT_PAGE_LIMIT]:
        html_text = _get_text(session, url, domain)
        if not html_text:
            continue
        product = _product_from_page(domain, url, html_text, "best_effort_sitemap")
        if product:
            product["platform"] = platform or "best_effort"
            products.append(product)
        if len(products) >= BEST_EFFORT_PRODUCT_LIMIT:
            break

    products = _dedupe_products(products)
    if products:
        print(f"  🧭 best-effort: {len(products)} 个产品线索")
        return products

    if with_browser:
        browser_products = fetch_products_bb_browser(domain)
        for product in browser_products or []:
            product["_best_effort"] = True
            product["source"] = product.get("source") or "bb_browser"
            product["platform"] = platform or "best_effort"
            product["id"] = product.get("id") or _stable_product_id(domain, product.get("url", ""), product.get("title", ""))
            product["url"] = product.get("url") or f"{get_base_url(domain)}/products/{product.get('handle', '')}"
        return _dedupe_products(browser_products or [])

    return []


def validate_best_effort_product_quality(product):
    errors = []
    if not str(product.get("title", "")).strip():
        errors.append("title为空")
    if not str(product.get("handle", "")).strip():
        errors.append("handle为空")
    variants = product.get("variants", []) or []
    images = product.get("images", []) or []
    price = max([_safe_float(v.get("price")) for v in variants] or [0])
    if price <= 0 and not images:
        errors.append("缺少价格和图片")
    return len(errors) == 0, errors


def fetch_all_products(domain, session=None, with_browser=False, best_effort=False):
    """分页抓取店铺所有产品（自动检测平台）+ v4.0 数据质量验证
    v6.0: with_browser=False 时跳过 bb-browser（默认禁用，太慢）
    """
    if session is None:
        session = get_session()

    domain = normalize_domain(domain)

    # 黑名单检查：普通模式保守跳过；全站尽力扫描模式仍继续尝试。
    if is_blacklisted(domain) and not best_effort:
        return []

    # 健康评分检查：普通模式跳过 dead；全站尽力扫描模式给一次机会。
    health = get_domain_health(domain)
    if health.get("status") == "dead" and not best_effort:
        return []

    # 平台检测
    platform = detect_platform(domain, session, fast=True)
    print(f"  🔍 平台: {platform}")

    if platform == "shopify":
        products = _fetch_shopify_products(domain, session)
        if best_effort and not products:
            products = fetch_best_effort_products(domain, session, platform=platform, with_browser=with_browser)
    elif with_browser and platform in ("lightfunnels", "magento", "woocommerce"):
        print(f"  🌐 非 Shopify 平台 ({platform})，启动 bb-browser 抓取...")
        products = fetch_products_bb_browser(domain)
        if best_effort and not products:
            products = fetch_best_effort_products(domain, session, platform=platform, with_browser=False)
        if not products and not best_effort:
            save_to_blacklist(domain, platform, reason="bb-browser 无法抓取产品")
            return []
    elif platform in ("lightfunnels", "magento", "woocommerce"):
        if best_effort:
            print(f"  🧭 非 Shopify 平台 ({platform})，改走 best-effort 扫描...")
            products = fetch_best_effort_products(domain, session, platform=platform, with_browser=with_browser)
        else:
            # v6.0: 明确的非 Shopify 平台直接加入黑名单
            save_to_blacklist(domain, platform, reason=f"非 Shopify 平台 ({platform})")
            return []
    else:
        # unknown: 可能是 Shopify 但 products.json 被禁用，尝试 HTML 检测
        html_platform = detect_platform(domain, session, fast=False)
        if html_platform == "shopify":
            print(f"  🔍 HTML 检测确认为 Shopify（products.json 被禁用）")
            products = fetch_best_effort_products(domain, session, platform=html_platform, with_browser=with_browser) if best_effort else []
        elif html_platform == "unknown":
            if best_effort:
                products = fetch_best_effort_products(domain, session, platform="unknown", with_browser=with_browser)
            else:
                save_to_blacklist(domain, "unknown", reason="products.json + HTML 均无法识别")
                return []
        else:
            if best_effort:
                products = fetch_best_effort_products(domain, session, platform=html_platform, with_browser=with_browser)
            else:
                save_to_blacklist(domain, html_platform, reason=f"HTML 检测为 {html_platform}")
                return []

    if products is None:
        products = []

    # ─── v4.0 数据质量验证 ───
    print(f"  🔬 v4.0 数据质量验证中...")
    valid_products = []
    rejected_products = []
    validation_stats = {
        "total": len(products),
        "valid": 0,
        "rejected": 0,
        "error_breakdown": {}
    }

    if validation_stats["total"] == 0:
        print(f"\n  📊 v4.0 验证统计:")
        print(f"     总产品数: 0")
        print(f"     ✅ 通过: 0 (0.0%)")
        print(f"     ❌ 拒绝: 0 (0.0%)")
        return []

    for i, product in enumerate(products, 1):
        # 品类黑名单检查
        is_blacklisted_cat, keyword = is_category_blacklisted(product)
        if is_blacklisted_cat:
            rejected_products.append({
                "product": product,
                "reason": f"品类黑名单: {keyword}"
            })
            validation_stats["rejected"] += 1
            validation_stats["error_breakdown"]["category_blacklist"] = \
                validation_stats["error_breakdown"].get("category_blacklist", 0) + 1
            continue
        
        if best_effort and product.get("_best_effort"):
            is_valid, errors = validate_best_effort_product_quality(product)
        else:
            is_valid, errors = validate_product_quality(product, domain, session)
        
        if is_valid:
            valid_products.append(product)
            validation_stats["valid"] += 1
        else:
            rejected_products.append({
                "product_id": product.get("id"),
                "title": product.get("title", ""),
                "handle": product.get("handle", ""),
                "errors": errors
            })
            validation_stats["rejected"] += 1
            
            # 统计错误类型
            for error in errors:
                validation_stats["error_breakdown"][error] = \
                    validation_stats["error_breakdown"].get(error, 0) + 1
        
        # 每100个产品显示进度
        if i % 100 == 0:
            print(f"  ✓ 已验证 {i}/{len(products)} 个产品...")

    # 保存被拒绝的产品
    if rejected_products:
        rejected_file = save_rejected_products(domain, rejected_products)
        print(f"  📝 已保存 {len(rejected_products)} 个被拒绝产品到: {rejected_file}")

    # 输出验证统计
    print(f"\n  📊 v4.0 验证统计:")
    print(f"     总产品数: {validation_stats['total']}")
    print(f"     ✅ 通过: {validation_stats['valid']} ({validation_stats['valid']/validation_stats['total']*100:.1f}%)")
    print(f"     ❌ 拒绝: {validation_stats['rejected']} ({validation_stats['rejected']/validation_stats['total']*100:.1f}%)")
    
    if validation_stats["error_breakdown"]:
        print(f"     错误分布:")
        for error_type, count in sorted(validation_stats["error_breakdown"].items(), 
                                        key=lambda x: x[1], reverse=True):
            print(f"       • {error_type}: {count}")

    return valid_products


@api_guard("shopify_api", cooldown=1, use_cache=True) if HAS_GUARD else lambda f: f
def fetch_collections(domain, session=None):
    """抓取所有集合"""
    if session is None:
        session = get_session()

    base = get_base_url(domain)
    collections = []
    page = 1

    while True:
        url = f"{base}/collections.json?limit=250&page={page}"
        try:
            r = session.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"  ❌ 集合第{page}页失败: {e}")
            break

        batch = data.get("collections", [])
        if not batch:
            break

        collections.extend(batch)
        if len(batch) < 250:
            break

        page += 1
        time.sleep(DELAY)

    return collections


@api_guard("shopify_api", cooldown=1, use_cache=True) if HAS_GUARD else lambda f: f
def fetch_shop_info(domain, session=None):
    """抓取店铺基本信息"""
    if session is None:
        session = get_session()

    base = get_base_url(domain)
    try:
        r = session.get(f"{base}/shop.json", timeout=TIMEOUT)
        if r.status_code == 200:
            return r.json().get("shop", {})
    except:
        pass
    return {}


# ─── 数据处理 ───
def parse_product(p):
    """解析单个产品为扁平字典"""
    variants = p.get("variants", []) or []
    images = p.get("images", []) or []

    prices = [_safe_float(v.get("price", 0)) for v in variants if v.get("price")]
    prices = [price for price in prices if price >= 0]
    min_price = min(prices) if prices else 0
    max_price = max(prices) if prices else 0
    tags = p.get("tags", [])
    if isinstance(tags, str):
        tags_text = tags
    else:
        tags_text = ", ".join(str(t) for t in tags)
    handle = p.get("handle", "")
    product_url = p.get("url") or (f"{get_base_url(p.get('domain', ''))}/products/{handle}" if handle else "")

    return {
        "id": p.get("id"),
        "handle": handle,
        "标题": p.get("title", ""),
        "产品类型": p.get("product_type", ""),
        "供应商": p.get("vendor", ""),
        "标签": tags_text,
        "最低价": min_price,
        "最高价": max_price,
        "变体数": len(variants),
        "图片数": len(images),
        "主图": images[0].get("src", "") if images else "",
        "状态": "发布" if p.get("published_at") else "草稿",
        "发布时间": p.get("published_at", ""),
        "创建时间": p.get("created_at", ""),
        "更新时间": p.get("updated_at", ""),
        "链接": product_url,
        "source": p.get("source", ""),
        "platform": p.get("platform", ""),
        "描述摘要": (p.get("body_html", "") or "")[:200].replace("<", "").replace(">", ""),
    }


def product_fingerprint(product):
    """产品指纹（用于检测变化）"""
    key = f"{product.get('id')}_{product.get('updated_at')}_{product.get('title')}"
    return hashlib.md5(key.encode()).hexdigest()


# ─── 快照管理 ───
def get_snapshot_path(domain):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    return DATA_DIR / f"{normalize_domain(domain)}_snapshot.json"


def load_snapshot(domain):
    path = get_snapshot_path(domain)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {"products": {}, "last_check": None}


def save_snapshot(domain, products_parsed):
    domain = normalize_domain(domain)
    path = get_snapshot_path(domain)
    snapshot = {
        "domain": domain,
        "last_check": datetime.now().isoformat(),
        "product_count": len(products_parsed),
        "products": {},
    }
    for p in products_parsed:
        tags = p.get("标签", "")
        tag_list = [t.strip() for t in tags.split(",") if t.strip()] if tags else []
        handle = p.get("handle", "")
        product_url = p.get("链接", "") or (f"https://{domain}/products/{handle}" if handle else f"https://{domain}")
        if product_url.startswith("https:///") or product_url.startswith("https://products/"):
            product_url = f"https://{domain}/products/{handle}" if handle else f"https://{domain}"
        snapshot["products"][str(p["id"])] = {
            "title": p["标题"],
            "handle": handle,
            "price": p["最低价"],
            "updated_at": p["更新时间"],
            "created_at": p.get("创建时间", ""),
            "published_at": p.get("发布时间", ""),
            "product_type": p.get("产品类型", ""),
            "vendor": p.get("供应商", ""),
            "tags": tag_list,
            "url": product_url,
            "source": p.get("source", ""),
            "platform": p.get("platform", ""),
            "fingerprint": product_fingerprint({
                "id": p["id"],
                "updated_at": p["更新时间"],
                "title": p["标题"],
            }),
        }
    path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


# ─── 新品检测 ───
def detect_changes(domain, current_products):
    """对比快照，检测新品和变化"""
    snapshot = load_snapshot(domain)
    old_products = snapshot.get("products", {})
    last_check = snapshot.get("last_check")

    current_map = {}
    for p in current_products:
        parsed = parse_product(p)
        current_map[str(p["id"])] = parsed

    new_products = []
    updated_products = []
    removed_ids = set(old_products.keys()) - set(current_map.keys())

    for pid, p in current_map.items():
        if pid not in old_products:
            new_products.append(p)
        else:
            old = old_products[pid]
            if old.get("updated_at") != p["更新时间"]:
                updated_products.append({
                    "title": p["标题"],
                    "old_price": old.get("price"),
                    "new_price": p["最低价"],
                    "handle": p["handle"],
                })

    return {
        "domain": normalize_domain(domain),
        "last_check": last_check,
        "current_count": len(current_map),
        "new_products": new_products,
        "updated_products": updated_products,
        "removed_count": len(removed_ids),
    }


# ─── Excel 导出（完整参数，域名命名，追加更新）───
def get_excel_path(domain):
    """Excel 文件按域名命名，固定一个文件"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    return DATA_DIR / f"{normalize_domain(domain)}.xlsx"


def monitor_monitor_export_excel(domain, products, collections=None, shop_info=None):
    """导出完整产品参数 Excel，同域名始终更新同一个文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("❌ 需要 openpyxl: pip install openpyxl")
        return None

    excel_path = get_excel_path(domain)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    domain_clean = normalize_domain(domain)

    # ── Sheet 1: 产品完整参数 ──
    headers = [
        "抓取时间", "产品ID", "Handle", "标题", "描述HTML", "产品类型", "供应商",
        "标签", "状态", "发布时间", "创建时间", "更新时间",
        "选项1名称", "选项1值", "选项2名称", "选项2值", "选项3名称", "选项3值",
        "变体ID", "变体标题", "SKU", "价格", "原价(Compare)", "重量(g)",
        "需要运输", "需纳税", "库存可用", "变体图片",
        "图片数", "主图URL", "图2", "图3", "图4", "图5",
        "产品链接", "店铺域名",
    ]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 追加模式：如果文件已存在，追加到已有 Sheet
    if excel_path.exists():
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb["产品列表"] if "产品列表" in wb.sheetnames else wb.active
        # 不写表头，直接追加
        write_header = False
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "产品列表"
        write_header = True

    if write_header:
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        # 冻结首行
        ws.freeze_panes = "A2"

    # 写入产品数据
    row_count = 0
    for p_raw in products:
        pid = p_raw.get("id", "")
        handle = p_raw.get("handle", "")
        title = p_raw.get("title", "")
        body_html = (p_raw.get("body_html", "") or "").replace("<br>", "\n").replace("<br/>", "\n")
        # 去除 HTML 标签保留文本
        body_text = re.sub(r'<[^>]+>', '', body_html).strip()[:500]
        product_type = p_raw.get("product_type", "")
        vendor = p_raw.get("vendor", "")
        tags = ", ".join(p_raw.get("tags", []))
        status = "发布" if p_raw.get("published_at") else "草稿"
        published_at = p_raw.get("published_at", "")
        created_at = p_raw.get("created_at", "")
        updated_at = p_raw.get("updated_at", "")

        # 选项
        options = p_raw.get("options", [])
        opt1_name = options[0]["name"] if len(options) > 0 else ""
        opt2_name = options[1]["name"] if len(options) > 1 else ""
        opt3_name = options[2]["name"] if len(options) > 2 else ""

        # 图片
        images = p_raw.get("images", [])
        img_count = len(images)
        img_urls = [img.get("src", "") for img in images[:5]]
        while len(img_urls) < 5:
            img_urls.append("")

        # 每个变体一行
        variants = p_raw.get("variants", [])
        if not variants:
            # 没变体也写一行
            variants = [{}]

        for v in variants:
            opt1_val = v.get("option1", "")
            opt2_val = v.get("option2", "")
            opt3_val = v.get("option3", "")

            row = [
                now_str, pid, handle, title, body_text, product_type, vendor,
                tags, status, published_at, created_at, updated_at,
                opt1_name, opt1_val, opt2_name, opt2_val, opt3_name, opt3_val,
                v.get("id", ""), v.get("title", ""), v.get("sku", ""),
                v.get("price", ""), v.get("compare_at_price", ""),
                v.get("grams", 0),
                "是" if v.get("requires_shipping") else "否",
                "是" if v.get("taxable") else "否",
                "是" if v.get("available") else "否",
                (v.get("featured_image", {}).get("src", "") if isinstance(v.get("featured_image"), dict) else str(v.get("featured_image", ""))),
                img_count, img_urls[0], img_urls[1], img_urls[2], img_urls[3], img_urls[4],
                f"https://{domain_clean}/products/{handle}",
                domain_clean,
            ]
            ws.append(row)
            row_count += 1

    # 自动列宽（仅首次）
    if write_header:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col[:50])  # 只看前50行
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # ── Sheet 2: 集合 ──
    if collections:
        if "集合" in wb.sheetnames:
            del wb["集合"]
        ws2 = wb.create_sheet("集合")
        ws2.append(["集合名称", "Handle", "描述", "产品数量"])
        for c in collections:
            ws2.append([
                c.get("title", ""),
                c.get("handle", ""),
                re.sub(r'<[^>]+>', '', c.get("body_html", "") or "")[:200],
                c.get("products_count", 0),
            ])

    # ── Sheet 3: 店铺信息 ──
    if shop_info:
        if "店铺信息" in wb.sheetnames:
            del wb["店铺信息"]
        ws3 = wb.create_sheet("店铺信息")
        ws3.append(["字段", "值", "抓取时间"])
        for k, v in shop_info.items():
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)[:500]
            ws3.append([k, str(v), now_str])

    # 保存
    wb.save(str(excel_path))
    print(f"  💾 Excel: {excel_path}")
    print(f"  📊 新增 {row_count} 行产品数据")
    return excel_path


# ─── 广告追踪检测 ───
def detect_ad_tracking(domain, session=None):
    """检测 Shopify 店铺的广告追踪代码"""
    if session is None:
        session = get_session()

    base = get_base_url(domain)
    result = {
        "facebook": {"active": False, "pixel_id": "", "events": []},
        "google_ads": {"active": False, "aw_id": "", "merchant_id": "", "events": []},
        "google_analytics": {"active": False, "ga_id": ""},
        "tiktok": {"active": False},
        "pinterest": {"active": False},
        "email_marketing": {"active": False, "platform": ""},
        "other": [],
    }

    try:
        r = session.get(base, timeout=15)
        html = r.text

        # Facebook Pixel（多种格式 + Shopify pixels manager 三重转义）
        fb_pixel = re.findall(r'\\?"pixel_id\\?"\s*:\\?\s*\\?"(\d{10,})\\?"', html)
        if not fb_pixel:
            fb_pixel = re.findall(r'pixel_id.*?(\d{15,})', html)
        if not fb_pixel:
            fb_pixel = re.findall(r'fbq\s*\(\s*["\']init["\']\s*,\s*["\'](\d{10,})["\']', html)
        if fb_pixel:
            result["facebook"]["active"] = True
            result["facebook"]["pixel_id"] = fb_pixel[0]

        # GA4（从 Shopify config 提取 GT- ID 或 G- ID）
        ga4_ids = re.findall(r'GT-[A-Z0-9]+', html)
        if not ga4_ids:
            ga4_ids = re.findall(r'G-[A-Z0-9]{8,}', html)
        ua_ids = re.findall(r'UA-\d{4,}-\d+', html)
        if ga4_ids or ua_ids:
            result["google_analytics"]["active"] = True
            result["google_analytics"]["ga_id"] = (ga4_ids[0] if ga4_ids else ua_ids[0])

        # Google Ads AW ID
        aw_ids = re.findall(r'AW-(\d{8,})', html)
        if aw_ids:
            result["google_ads"]["active"] = True
            result["google_ads"]["aw_id"] = f"AW-{aw_ids[0]}"

        # Google Merchant Center
        mc_ids = re.findall(r'MC-([A-Z0-9]+)', html)
        if mc_ids:
            result["google_ads"]["merchant_id"] = f"MC-{mc_ids[0]}"

        # GA4 / UA
        ga_ids = re.findall(r'(G-[A-Z0-9]{8,}|UA-\d{4,}-\d+)', html)
        if ga_ids:
            result["google_analytics"]["active"] = True
            result["google_analytics"]["ga_id"] = ga_ids[0]

        # Google Ads 事件类型
        gtag_events = re.findall(r'"type"\s*:\s*"(\w+)"', html)
        known_events = ["page_view", "view_item", "add_to_cart", "begin_checkout", "purchase", "search", "add_payment_info"]
        result["google_ads"]["events"] = [e for e in known_events if e in gtag_events]

        # TikTok
        if re.search(r'tiktok|ttq\.', html, re.IGNORECASE):
            result["tiktok"]["active"] = True

        # Pinterest
        if re.search(r'pintrk|pinterest.*tag', html, re.IGNORECASE):
            result["pinterest"]["active"] = True

        # Email marketing
        email_platforms = {
            "klaviyo": r'klaviyo',
            "omnisend": r'omnisend',
            "mailchimp": r'mailchimp',
            "privy": r'privy',
            "justuno": r'justuno',
            "drip": r'getdrip',
        }
        for platform, pattern in email_platforms.items():
            if re.search(pattern, html, re.IGNORECASE):
                result["email_marketing"]["active"] = True
                result["email_marketing"]["platform"] = platform
                break

        # Other tools
        other_tools = {
            "judge.me": r'judge\.me|judgeme',
            "loox": r'loox\.io',
            "stamped": r'stamped\.io',
            "yotpo": r'yotpo',
            "hotjar": r'hotjar',
            "luckyorange": r'luckyorange',
            "recart": r'recart',
            "carthook": r'carthook',
            "reconvert": r'reconvert',
        }
        for tool, pattern in other_tools.items():
            if re.search(pattern, html, re.IGNORECASE):
                result["other"].append(tool)

    except Exception as e:
        result["error"] = str(e)

    return result


# ─── v6.0 着陆页情报采集 ───
def fetch_landing_page_intel(domain, session=None):
    """采集竞品着陆页的转化策略信号
    提取：hero 产品、弹窗优惠、urgency 元素、社会证明、订阅策略、bundle 策略
    """
    if session is None:
        session = get_session()

    base = get_base_url(domain)
    intel = {
        "domain": normalize_domain(domain),
        "captured_at": datetime.now().isoformat(),
        "hero": {},
        "popup": {},
        "urgency": [],
        "social_proof": {},
        "subscription": {},
        "bundles": [],
        "nav_categories": [],
        "tech_stack": {},
    }

    try:
        headers = _risk.get_headers()
        _risk.before_request(domain)
        r = session.get(base, timeout=15, headers=headers, allow_redirects=True)
        html = r.text.lower()
        raw_html = r.text

        # ── Hero 产品/系列 ──
        hero_titles = re.findall(r'<h[12][^>]*>([^<]{10,80})</h[12]>', raw_html)
        if hero_titles:
            intel["hero"]["main_heading"] = hero_titles[0].strip()

        cta_texts = re.findall(r'(shop\s+now|add\s+to\s+cart|buy\s+now|get\s+yours|order\s+now|learn\s+more)', html)
        intel["hero"]["cta_count"] = len(cta_texts)
        intel["hero"]["cta_text"] = list(set(cta_texts))[:5]

        # ── 弹窗/注册优惠 ──
        popup_indicators = {
            "email_popup": bool(re.search(r'email.*popup|popup.*email|newsletter.*modal|klaviyo.*form', html)),
            "discount_offer": bool(re.search(r'\d+%\s*off|save\s+\$?\d+|free\s+shipping|first\s+order', html)),
            "discount_code": re.findall(r'([A-Z0-9]{4,12})\s*(?:to\s+get|for\s+\d+%|at\s+checkout)', raw_html),
        }
        discount_match = re.findall(r'(\d+)%\s*off', html)
        if discount_match:
            intel["popup"]["discount_pct"] = max(int(x) for x in discount_match)
        intel["popup"]["has_popup"] = popup_indicators["email_popup"]
        intel["popup"]["has_discount"] = popup_indicators["discount_offer"]
        intel["popup"]["discount_codes"] = popup_indicators["discount_code"][:3]

        # ── Urgency 元素 ──
        urgency_signals = []
        if re.search(r'countdown|timer|limited\s+time|ends\s+in|flash\s+sale', html):
            urgency_signals.append("countdown_timer")
        if re.search(r'only\s+\d+\s+left|low\s+stock|selling\s+fast|almost\s+gone', html):
            urgency_signals.append("scarcity")
        if re.search(r'\d+\s+people\s+(?:are\s+)?(?:viewing|looking|watching)', html):
            urgency_signals.append("live_visitors")
        if re.search(r'back\s+in\s+stock|just\s+restocked', html):
            urgency_signals.append("back_in_stock")
        if re.search(r'free\s+shipping.*(?:today|limited|ends)', html):
            urgency_signals.append("free_shipping_urgency")
        intel["urgency"] = urgency_signals

        # ── 社会证明 ──
        review_systems = {
            "judge_me": bool(re.search(r'judge\.me|judgeme', html)),
            "loox": bool(re.search(r'loox\.io', html)),
            "yotpo": bool(re.search(r'yotpo', html)),
            "stamped": bool(re.search(r'stamped\.io', html)),
            "okendo": bool(re.search(r'okendo', html)),
            "trustpilot": bool(re.search(r'trustpilot', html)),
        }
        intel["social_proof"]["review_system"] = [k for k, v in review_systems.items() if v]

        review_count = re.findall(r'(\d[\d,]*)\s*(?:reviews?|ratings?|reviews?\s*\)', html)
        if review_count:
            intel["social_proof"]["review_count"] = review_count[0]

        media_mentions = re.findall(r'(?:as\s+seen\s+on|featured\s+in|press)', html)
        intel["social_proof"]["media_mentions"] = len(media_mentions) > 0

        user_count = re.findall(r'([\d,]+)\s*(?:happy\s+customers|customers|users|community\s+members)', html)
        if user_count:
            intel["social_proof"]["customer_count"] = user_count[0]

        # ── 订阅策略 ──
        has_subscription = bool(re.search(r'subscribe|subscription|auto-?replenish|recurring', html))
        intel["subscription"]["available"] = has_subscription
        if has_subscription:
            sub_discount = re.findall(r'subscribe.*?(\d+)%\s*(?:off|save|discount)', html)
            if sub_discount:
                intel["subscription"]["discount_pct"] = int(sub_discount[0])

        # ── Bundle 策略 ──
        bundle_signals = re.findall(r'(?:bundle|pack|set|kit)\s*(?:&|and)?\s*(?:save|discount)?\s*(\d+)?%?\s*(?:off)?', html)
        intel["bundles"] = list(set([b for b in bundle_signals if b]))[:5]

        # ── 导航品类 ──
        nav_items = re.findall(r'<a[^>]*href="(?:/collections/[^"]+)"[^>]*>([^<]+)</a>', raw_html)
        intel["nav_categories"] = [n.strip() for n in nav_items if n.strip()][:15]

        # ── 技术栈 ──
        tech = detect_ad_tracking(domain, session)
        intel["tech_stack"] = {
            "fb_pixel": tech.get("facebook", {}).get("active", False),
            "ga4": tech.get("google_analytics", {}).get("active", False),
            "tiktok": tech.get("tiktok", {}).get("active", False),
            "email_platform": tech.get("email_marketing", {}).get("platform", ""),
            "review_system": intel["social_proof"]["review_system"][:1],
        }

    except Exception as e:
        intel["error"] = str(e)[:200]

    # 保存
    intel_path = DATA_DIR / f"{normalize_domain(domain)}_landing_intel.json"
    intel_path.parent.mkdir(parents=True, exist_ok=True)
    intel_path.write_text(json.dumps(intel, ensure_ascii=False, indent=2), encoding='utf-8')

    return intel


def detect_lp_changes(domain):
    """对比两次着陆页情报，检测策略变化"""
    intel_path = DATA_DIR / f"{normalize_domain(domain)}_landing_intel.json"
    if not intel_path.exists():
        return None

    current = json.loads(intel_path.read_text(encoding='utf-8'))

    history_path = DATA_DIR / f"{normalize_domain(domain)}_lp_history.json"
    if not history_path.exists():
        history_path.write_text(json.dumps([current], ensure_ascii=False), encoding='utf-8')
        return None

    history = json.loads(history_path.read_text(encoding='utf-8'))
    if not history:
        return None

    previous = history[-1]
    changes = []

    if current.get("hero", {}).get("main_heading") != previous.get("hero", {}).get("main_heading"):
        changes.append({
            "type": "hero_change",
            "old": previous.get("hero", {}).get("main_heading", ""),
            "new": current.get("hero", {}).get("main_heading", ""),
        })

    old_disc = previous.get("popup", {}).get("discount_pct", 0)
    new_disc = current.get("popup", {}).get("discount_pct", 0)
    if old_disc != new_disc and (old_disc > 0 or new_disc > 0):
        changes.append({
            "type": "discount_change",
            "old": f"{old_disc}%",
            "new": f"{new_disc}%",
        })

    old_urgency = set(previous.get("urgency", []))
    new_urgency = set(current.get("urgency", []))
    added = new_urgency - old_urgency
    removed = old_urgency - new_urgency
    if added:
        changes.append({"type": "urgency_added", "signals": list(added)})
    if removed:
        changes.append({"type": "urgency_removed", "signals": list(removed)})

    old_cats = set(previous.get("nav_categories", []))
    new_cats = set(current.get("nav_categories", []))
    added_cats = new_cats - old_cats
    if added_cats:
        changes.append({"type": "new_categories", "categories": list(added_cats)})

    history.append(current)
    history = history[-30:]
    history_path.write_text(json.dumps(history, ensure_ascii=False), encoding='utf-8')

    return changes if changes else None


# ─── v6.0 FB Ad Library 监控 ───
def fetch_fb_ad_library(domain, session=None):
    """从 Facebook Ad Library 获取竞品的活跃广告数据"""
    if session is None:
        session = get_session()

    domain_clean = normalize_domain(domain)
    brand_name = domain_clean.split('.')[0]

    ad_library_url = f"https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={brand_name}&search_type=keyword_unordered"

    result = {
        "domain": domain_clean,
        "brand_name": brand_name,
        "search_url": ad_library_url,
        "captured_at": datetime.now().isoformat(),
        "active_ads": 0,
        "ad_details": [],
        "error": None,
    }

    try:
        headers = _risk.get_headers()
        _risk.before_request(domain)
        r = session.get(ad_library_url, timeout=15, headers=headers, allow_redirects=True)
        html = r.text

        count_match = re.findall(r'(\d[\d,]*)\s+results?\s+(?:for|matching)', html, re.IGNORECASE)
        if count_match:
            result["active_ads"] = int(count_match[0].replace(',', ''))

        advertiser_match = re.findall(r'"advertiser_name"\s*:\s*"([^"]+)"', html)
        if advertiser_match:
            result["advertiser_name"] = advertiser_match[0]

        creative_types = []
        if re.search(r'video|mp4|\.mp4', html, re.IGNORECASE):
            creative_types.append("video")
        if re.search(r'image|photo|\.jpg|\.png', html, re.IGNORECASE):
            creative_types.append("image")
        if re.search(r'carousel|swipe|multiple', html, re.IGNORECASE):
            creative_types.append("carousel")
        result["creative_types"] = creative_types

        ad_texts = re.findall(r'"text"\s*:\s*"([^"]{20,200})"', html)
        result["ad_texts"] = list(set(ad_texts))[:5]

        platforms = []
        if re.search(r'facebook|fb', html, re.IGNORECASE):
            platforms.append("facebook")
        if re.search(r'instagram|ig', html, re.IGNORECASE):
            platforms.append("instagram")
        if re.search(r'audience.?network', html, re.IGNORECASE):
            platforms.append("audience_network")
        if re.search(r'messenger', html, re.IGNORECASE):
            platforms.append("messenger")
        result["platforms"] = platforms

    except Exception as e:
        result["error"] = str(e)[:200]

    ad_path = DATA_DIR / f"{domain_clean}_fb_ads.json"
    ad_path.parent.mkdir(parents=True, exist_ok=True)
    ad_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')

    return result


def analyze_ad_velocity(domain):
    """分析竞品广告投放速度变化"""
    ad_path = DATA_DIR / f"{normalize_domain(domain)}_fb_ads.json"
    if not ad_path.exists():
        return None

    current = json.loads(ad_path.read_text(encoding='utf-8'))

    history_path = DATA_DIR / f"{normalize_domain(domain)}_fb_ad_history.json"
    if not history_path.exists():
        history_path.write_text(json.dumps([current], ensure_ascii=False), encoding='utf-8')
        return None

    history = json.loads(history_path.read_text(encoding='utf-8'))
    if not history:
        return None

    previous = history[-1]
    changes = []

    old_count = previous.get("active_ads", 0)
    new_count = current.get("active_ads", 0)
    if old_count > 0 and new_count > 0:
        pct_change = ((new_count - old_count) / old_count) * 100
        if abs(pct_change) > 50:
            changes.append({
                "type": "ad_volume_change",
                "old": old_count,
                "new": new_count,
                "change_pct": round(pct_change, 1),
                "alert": "surge" if pct_change > 100 else "drop",
            })

    history.append(current)
    history = history[-30:]
    history_path.write_text(json.dumps(history, ensure_ascii=False), encoding='utf-8')

    return changes if changes else None


# ─── HTML 导出（专业版，带产品图+链接+参数）───
def get_html_path(domain):
    """HTML 文件按域名命名"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    return DATA_DIR / f"{normalize_domain(domain)}.html"


def monitor_monitor_export_html(domain, products, collections=None, shop_info=None, changes=None):
    """导出专业 HTML 报告，带产品图、链接、完整参数"""
    domain_clean = normalize_domain(domain)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    html_path = get_html_path(domain)
    now_dt = datetime.now()

    # 解析产品
    parsed = []
    for i, p in enumerate(products):
        variants = p.get("variants", [])
        images = p.get("images", [])
        prices = [float(v.get("price", 0)) for v in variants if v.get("price")]
        body_html = (p.get("body_html", "") or "")
        body_text = re.sub(r'<[^>]+>', '', body_html).strip()[:300]
        tags = p.get("tags", [])

        # 上架时间
        published_at = p.get("published_at", "")
        created_at = p.get("created_at", "")

        # 判断新品（30天内）
        is_new = False
        if created_at:
            try:
                created_dt = datetime.fromisoformat(created_at.replace("Z", "+00:00").rsplit("+", 1)[0] if "+" not in created_at and "Z" not in created_at else created_at.replace("Z", "+00:00"))
                is_new = (now_dt - created_dt.replace(tzinfo=None)).days <= 30
            except:
                pass

        # TOP 判断（多维度：图片多 + 变体多 + 有标签 + 有描述）
        score = 0
        if len(images) >= 3: score += 2
        if len(images) >= 5: score += 2
        if len(variants) >= 2: score += 1
        if len(variants) >= 4: score += 2
        if len(tags) >= 3: score += 1
        if len(body_text) > 100: score += 1
        if prices and min(prices) >= 15: score += 1
        is_top = score >= 5

        # BEST SELLING 推断（标签含 best/seller/popular 或 价格适中 + 图片多 + 变体多）
        best_selling_signals = ["best", "seller", "popular", "hot", "top", "favorite", "trending"]
        is_best = any(s in " ".join(tags).lower() for s in best_selling_signals)
        if not is_best and len(variants) >= 3 and len(images) >= 3 and prices and 10 <= min(prices) <= 50:
            is_best = True

        parsed.append({
            "id": p.get("id"),
            "handle": p.get("handle", ""),
            "title": p.get("title", ""),
            "type": p.get("product_type", ""),
            "vendor": p.get("vendor", ""),
            "tags": tags,
            "status": "发布" if published_at else "草稿",
            "price_min": min(prices) if prices else 0,
            "price_max": max(prices) if prices else 0,
            "variant_count": len(variants),
            "image_count": len(images),
            "main_image": images[0].get("src", "") if images else "",
            "images": [img.get("src", "") for img in images[:6]],
            "variants": [{
                "title": v.get("title", ""),
                "sku": v.get("sku", ""),
                "price": v.get("price", ""),
                "compare_at": v.get("compare_at_price", ""),
                "available": v.get("available", False),
            } for v in variants[:10]],
            "body": body_text,
            "published_at": published_at,
            "created_at": created_at,
            "updated_at": p.get("updated_at", ""),
            "link": f"https://{domain_clean}/products/{p.get('handle', '')}",
            "is_new": is_new,
            "is_top": is_top,
            "is_best": is_best,
            "position": i + 1,  # 在原始列表中的位置（可视为推荐顺序）
        })

    # 广告追踪检测
    ad_tracking = detect_ad_tracking(domain)
    fb = ad_tracking.get("facebook", {})
    ga = ad_tracking.get("google_ads", {})
    g4 = ad_tracking.get("google_analytics", {})
    tiktok = ad_tracking.get("tiktok", {})
    pinterest = ad_tracking.get("pinterest", {})
    email = ad_tracking.get("email_marketing", {})
    other_tools = ad_tracking.get("other", [])

    # 按产品类型分组
    type_groups = {}
    for p in parsed:
        t = p["type"] or "未分类"
        type_groups.setdefault(t, []).append(p)

    # 统计
    total_products = len(parsed)
    total_variants = sum(p["variant_count"] for p in parsed)
    price_avg = sum(p["price_min"] for p in parsed) / total_products if total_products else 0
    types_count = len(type_groups)
    new_count = sum(1 for p in parsed if p["is_new"])
    top_count = sum(1 for p in parsed if p["is_top"])
    best_count = sum(1 for p in parsed if p["is_best"])

    # 新品标记（来自 changes 检测 + 30天内创建）
    new_ids = set()
    if changes and changes.get("new_products"):
        new_ids = {p["id"] for p in changes["new_products"]}
    for p in parsed:
        if p["is_new"]:
            new_ids.add(p["id"])

    # 生成 HTML
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>竞品分析报告 — {domain_clean}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f0f2f5; color: #1a1a2e; line-height: 1.6; }}
.container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}

/* Header */
.header {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%); color: white; padding: 40px; border-radius: 16px; margin-bottom: 24px; }}
.header h1 {{ font-size: 28px; margin-bottom: 8px; }}
.header .domain {{ color: #e94560; font-size: 18px; font-weight: 600; }}
.header .meta {{ display: flex; gap: 24px; margin-top: 16px; font-size: 14px; opacity: 0.85; }}

/* Stats */
.stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px; }}
.stat-card {{ background: white; border-radius: 12px; padding: 20px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.stat-card .number {{ font-size: 32px; font-weight: 700; color: #0f3460; }}
.stat-card .label {{ font-size: 13px; color: #666; margin-top: 4px; }}

/* Section */
.section {{ background: white; border-radius: 12px; padding: 24px; margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.section-title {{ font-size: 20px; font-weight: 700; margin-bottom: 16px; padding-bottom: 12px; border-bottom: 2px solid #f0f2f5; display: flex; align-items: center; gap: 8px; }}

/* Filter */
.filter-bar {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 16px; }}
.filter-btn {{ padding: 6px 16px; border: 2px solid #e0e0e0; border-radius: 20px; background: white; cursor: pointer; font-size: 13px; transition: all 0.2s; }}
.filter-btn:hover, .filter-btn.active {{ border-color: #0f3460; background: #0f3460; color: white; }}
.filter-btn .count {{ background: #f0f2f5; padding: 1px 6px; border-radius: 10px; font-size: 11px; margin-left: 4px; }}
.filter-btn:hover .count, .filter-btn.active .count {{ background: rgba(255,255,255,0.2); }}

/* Product Grid */
.product-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap: 20px; }}
.product-card {{ border: 1px solid #e8e8e8; border-radius: 12px; overflow: hidden; transition: all 0.3s; position: relative; }}
.product-card:hover {{ box-shadow: 0 8px 24px rgba(0,0,0,0.12); transform: translateY(-2px); }}
.product-card.new-product {{ border-color: #e94560; }}
.product-card.new-product::before {{ content: '🆕 NEW'; position: absolute; top: 12px; left: 12px; background: #e94560; color: white; padding: 2px 10px; border-radius: 4px; font-size: 11px; font-weight: 700; z-index: 2; }}

.product-img {{ width: 100%; height: 280px; object-fit: cover; background: #f8f8f8; }}
.product-img-placeholder {{ width: 100%; height: 280px; background: linear-gradient(135deg, #f0f0f0, #e0e0e0); display: flex; align-items: center; justify-content: center; color: #999; font-size: 14px; }}

.product-info {{ padding: 16px; }}
.product-title {{ font-size: 15px; font-weight: 600; margin-bottom: 8px; line-height: 1.4; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; }}
.product-title a {{ color: #1a1a2e; text-decoration: none; }}
.product-title a:hover {{ color: #e94560; }}

.product-meta {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; }}
.product-price {{ font-size: 18px; font-weight: 700; color: #e94560; }}
.product-price .compare {{ font-size: 13px; color: #999; text-decoration: line-through; margin-left: 6px; font-weight: 400; }}
.product-type {{ font-size: 11px; background: #f0f2f5; padding: 2px 8px; border-radius: 4px; color: #666; }}

.product-tags {{ display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 8px; }}
.product-tag {{ font-size: 10px; background: #e8f4fd; color: #0f3460; padding: 2px 6px; border-radius: 3px; }}

.product-detail {{ font-size: 12px; color: #888; margin-bottom: 4px; }}
.product-desc {{ font-size: 12px; color: #666; line-height: 1.5; display: -webkit-box; -webkit-line-clamp: 3; -webkit-box-orient: vertical; overflow: hidden; margin-top: 8px; }}

.product-footer {{ padding: 12px 16px; border-top: 1px solid #f0f2f5; display: flex; justify-content: space-between; align-items: center; font-size: 12px; color: #999; }}
.product-link {{ color: #0f3460; text-decoration: none; font-weight: 500; }}
.product-link:hover {{ color: #e94560; }}

/* Images gallery */
.img-gallery {{ display: flex; gap: 4px; margin-top: 8px; }}
.img-gallery img {{ width: 36px; height: 36px; object-fit: cover; border-radius: 4px; border: 1px solid #e0e0e0; cursor: pointer; }}

/* Variant table */
.variant-table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 11px; }}
.variant-table th {{ background: #f8f8f8; padding: 6px 8px; text-align: left; font-weight: 600; border-bottom: 1px solid #e0e0e0; }}
.variant-table td {{ padding: 6px 8px; border-bottom: 1px solid #f0f0f0; }}

/* Type filter content */
.type-section {{ display: none; }}
.type-section.active {{ display: block; }}

/* Responsive */
@media (max-width: 768px) {{
    .product-grid {{ grid-template-columns: 1fr; }}
    .header {{ padding: 24px; }}
    .stats {{ grid-template-columns: repeat(2, 1fr); }}
}}

/* Table view */
.table-view {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
.table-view th {{ background: #1a1a2e; color: white; padding: 10px 12px; text-align: left; font-weight: 600; position: sticky; top: 0; }}
.table-view td {{ padding: 10px 12px; border-bottom: 1px solid #f0f0f0; }}
.table-view tr:hover {{ background: #f8f9fa; }}
.table-view img {{ width: 50px; height: 50px; object-fit: cover; border-radius: 6px; }}
.table-view .row-new {{ background: #fff5f5; }}

/* View toggle */
.view-toggle {{ display: flex; gap: 4px; }}
.view-btn {{ padding: 6px 12px; border: 1px solid #e0e0e0; background: white; cursor: pointer; font-size: 13px; }}
.view-btn.active {{ background: #1a1a2e; color: white; border-color: #1a1a2e; }}

/* Badges */
.badge {{ display: inline-block; padding: 2px 10px; border-radius: 4px; font-size: 11px; font-weight: 700; margin-bottom: 6px; margin-right: 4px; }}
.badge-best {{ background: #eafaf1; color: #27ae60; border: 1px solid #27ae60; }}
.badge-top {{ background: #fef9e7; color: #f5a623; border: 1px solid #f5a623; }}
.badge-new {{ background: #fdedec; color: #e94560; border: 1px solid #e94560; }}
</style>
</head>
<body>
<div class="container">
  <!-- Header -->
  <div class="header">
    <h1>🏪 竞品产品分析报告</h1>
    <div class="domain">{domain_clean}</div>
    <div class="meta">
      <span>📅 抓取时间: {now_str}</span>
      <span>🔗 数据来源: Shopify JSON API</span>
      <span>📊 产品总数: {total_products}</span>
    </div>
  </div>

  <!-- Stats -->
  <div class="stats">
    <div class="stat-card">
      <div class="number">{total_products}</div>
      <div class="label">产品总数</div>
    </div>
    <div class="stat-card">
      <div class="number">{total_variants}</div>
      <div class="label">SKU 变体数</div>
    </div>
    <div class="stat-card">
      <div class="number">{types_count}</div>
      <div class="label">产品分类</div>
    </div>
    <div class="stat-card">
      <div class="number">${price_avg:.2f}</div>
      <div class="label">平均最低价</div>
    </div>
    <div class="stat-card">
      <div class="number" style="color:#e94560">{new_count}</div>
      <div class="label">🆕 新品 (30天内)</div>
    </div>
    <div class="stat-card">
      <div class="number" style="color:#f5a623">{top_count}</div>
      <div class="label">🏆 TOP 产品</div>
    </div>
    <div class="stat-card">
      <div class="number" style="color:#27ae60">{best_count}</div>
      <div class="label">🔥 Best Selling</div>
    </div>
  </div>

  <!-- Ad Intelligence -->
  <div class="section">
    <div class="section-title">📢 广告情报</div>
    <div class="stats">
      <div class="stat-card" style="border-left: 4px solid {'#1877f2' if fb.get('active') else '#ccc'}">
        <div class="number" style="font-size:24px">{'✅' if fb.get('active') else '❌'}</div>
        <div class="label"><strong>Facebook Ads</strong></div>
        <div style="font-size:11px; margin-top:4px">
          {('Pixel: ' + fb.get('pixel_id', '')) if fb.get('active') else '未检测到'}
        </div>
        <div style="margin-top:6px">
          <a href="https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={domain_clean.split('.')[0]}&search_type=keyword_unordered" target="_blank" style="font-size:11px;color:#1877f2;text-decoration:none;font-weight:600">🔍 查看在投广告 →</a>
        </div>
        <div style="margin-top:3px">
          <a href="https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={domain_clean}&search_type=keyword_unordered" target="_blank" style="font-size:10px;color:#999;text-decoration:none">按域名搜索</a>
        </div>
      </div>
      <div class="stat-card" style="border-left: 4px solid {'#4285f4' if ga.get('active') else '#ccc'}">
        <div class="number" style="font-size:24px">{'✅' if ga.get('active') else '❌'}</div>
        <div class="label"><strong>Google Ads</strong></div>
        <div style="font-size:11px; margin-top:4px">
          {(ga.get('aw_id', '') + '<br>MC: ' + ga.get('merchant_id', '')) if ga.get('active') else '未检测到'}
        </div>
        <div style="margin-top:6px">
          <a href="https://adstransparency.google.com/?region=US&domain={domain_clean}" target="_blank" style="font-size:11px;color:#4285f4;text-decoration:none;font-weight:600">🔍 查看广告透明度 →</a>
        </div>
      </div>
      <div class="stat-card" style="border-left: 4px solid {'#ff0050' if tiktok.get('active') else '#ccc'}">
        <div class="number" style="font-size:24px">{'✅' if tiktok.get('active') else '❌'}</div>
        <div class="label"><strong>TikTok Ads</strong></div>
        <div style="font-size:11px; margin-top:4px">{'在投' if tiktok.get('active') else '未检测到'}</div>
      </div>
      <div class="stat-card" style="border-left: 4px solid {'#e60023' if pinterest.get('active') else '#ccc'}">
        <div class="number" style="font-size:24px">{'✅' if pinterest.get('active') else '❌'}</div>
        <div class="label"><strong>Pinterest Ads</strong></div>
        <div style="font-size:11px; margin-top:4px">{'在投' if pinterest.get('active') else '未检测到'}</div>
      </div>
      <div class="stat-card" style="border-left: 4px solid {'#00c4cc' if email.get('active') else '#ccc'}">
        <div class="number" style="font-size:24px">{'✅' if email.get('active') else '❌'}</div>
        <div class="label"><strong>邮件营销</strong></div>
        <div style="font-size:11px; margin-top:4px">{email.get('platform', '未检测到').upper() if email.get('active') else '未检测到'}</div>
      </div>
      <div class="stat-card" style="border-left: 4px solid {'#7c3aed' if g4.get('active') else '#ccc'}">
        <div class="number" style="font-size:24px">{'✅' if g4.get('active') else '❌'}</div>
        <div class="label"><strong>Google Analytics</strong></div>
        <div style="font-size:11px; margin-top:4px">{g4.get('ga_id', '') if g4.get('active') else '未检测到'}</div>
      </div>
    </div>

    {"" if not ga.get('events') else '<div style="margin-top:12px;padding:12px;background:#f0f4ff;border-radius:8px;font-size:13px"><strong>🎯 Google Ads 追踪事件:</strong> ' + ' | '.join(ga['events']) + '</div>'}

    {"" if not other_tools else '<div style="margin-top:12px;padding:12px;background:#f5f0ff;border-radius:8px;font-size:13px"><strong>🧰 其他工具:</strong> ' + ' | '.join(other_tools) + '</div>'}

    <div style="margin-top:12px;padding:12px;background:#fff8e1;border-radius:8px;font-size:13px">
      <strong>🔗 广告情报快捷入口：</strong><br>
      <a href="https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={domain_clean.split('.')[0]}&search_type=keyword_unordered" target="_blank" style="color:#1877f2">📘 Facebook Ad Library（品牌名搜）</a> ｜
      <a href="https://www.facebook.com/ads/library/?active_status=active&ad_type=all&country=US&q={domain_clean}&search_type=keyword_unordered" target="_blank" style="color:#1877f2">📘 Facebook Ad Library（域名搜）</a> ｜
      <a href="https://adstransparency.google.com/?region=US&domain={domain_clean}" target="_blank" style="color:#4285f4">📊 Google Ads Transparency</a> ｜
      <a href="https://www.similarweb.com/website/{domain_clean}/" target="_blank" style="color:#0070f3">📈 SimilarWeb 流量</a>
    </div>
  </div>

  <!-- Quick Filter -->
  <div class="section">
    <div class="section-title">⚡ 快速筛选</div>
    <div class="filter-bar">
      <button class="filter-btn active" onclick="filterTag('all')">📦 全部 <span class="count">{total_products}</span></button>
      <button class="filter-btn" onclick="filterTag('best')" style="border-color:#27ae60">🔥 Best Selling <span class="count">{best_count}</span></button>
      <button class="filter-btn" onclick="filterTag('top')" style="border-color:#f5a623">🏆 TOP <span class="count">{top_count}</span></button>
      <button class="filter-btn" onclick="filterTag('new')" style="border-color:#e94560">🆕 新品 <span class="count">{new_count}</span></button>
    </div>
  </div>

  <!-- Sort Bar -->
  <div class="section">
    <div class="section-title">🔃 排序方式</div>
    <div class="filter-bar">
      <button class="filter-btn active" onclick="sortProducts('default')">默认顺序</button>
      <button class="filter-btn" onclick="sortProducts('price-asc')">价格 ↑ 低到高</button>
      <button class="filter-btn" onclick="sortProducts('price-desc')">价格 ↓ 高到低</button>
      <button class="filter-btn" onclick="sortProducts('newest')">🆕 最新上架</button>
      <button class="filter-btn" onclick="sortProducts('name')">名称 A-Z</button>
    </div>
  </div>

  <!-- Category Filter -->
  <div class="section">
    <div class="section-title">📂 按产品类型筛选</div>
    <div class="filter-bar">
      <button class="filter-btn active" onclick="filterType('all')">全部 <span class="count">{total_products}</span></button>
"""

    # 分类按钮
    for type_name, items in sorted(type_groups.items(), key=lambda x: -len(x[1])):
        html += f'      <button class="filter-btn" onclick="filterType(\'{type_name}\')">{type_name} <span class="count">{len(items)}</span></button>\n'

    html += """    </div>
  </div>

  <!-- Product Grid -->
  <div class="section">
    <div class="section-title">📦 产品列表</div>
    <div class="product-grid" id="productGrid">
"""

    # 产品卡片
    for p in parsed:
        new_class = "new-product" if p["id"] in new_ids else ""
        img_tag = f'<img class="product-img" src="{p["main_image"]}" alt="{p["title"]}" loading="lazy">' if p["main_image"] else '<div class="product-img-placeholder">暂无图片</div>'

        # 价格显示
        if p["price_min"] == p["price_max"]:
            price_html = f'<span class="product-price">${p["price_min"]:.2f}</span>'
        else:
            price_html = f'<span class="product-price">${p["price_min"]:.2f} - ${p["price_max"]:.2f}</span>'

        # Compare at price
        compare_prices = [v["compare_at"] for v in p["variants"] if v.get("compare_at")]
        if compare_prices:
            price_html += f'<span class="compare">${compare_prices[0]}</span>'

        # 标签
        tags_html = ""
        for t in p["tags"][:5]:
            tags_html += f'<span class="product-tag">{t}</span>'

        # 分类标签（BEST / TOP / NEW）
        badge_html = ""
        if p["is_best"]:
            badge_html += '<span class="badge badge-best">🔥 BEST SELLING</span>'
        if p["is_top"]:
            badge_html += '<span class="badge badge-top">🏆 TOP</span>'
        if p["is_new"]:
            badge_html += '<span class="badge badge-new">🆕 NEW</span>'

        # 图片缩略图
        gallery_html = ""
        if len(p["images"]) > 1:
            gallery_html = '<div class="img-gallery">'
            for img_url in p["images"][:6]:
                gallery_html += f'<img src="{img_url}" alt="">'
            gallery_html += '</div>'

        # 变体表格
        variant_html = ""
        if len(p["variants"]) > 1:
            variant_html = '<table class="variant-table"><tr><th>变体</th><th>SKU</th><th>价格</th><th>状态</th></tr>'
            for v in p["variants"][:8]:
                status_icon = "✅" if v["available"] else "❌"
                compare_cell = f' <s>${v["compare_at"]}</s>' if v.get("compare_at") else ""
                variant_html += f'<tr><td>{v["title"]}</td><td>{v["sku"]}</td><td>${v["price"]}{compare_cell}</td><td>{status_icon}</td></tr>'
            variant_html += '</table>'

        # 时间格式化
        published = p["published_at"][:10] if p["published_at"] else ""
        created = p["created_at"][:10] if p["created_at"] else ""
        updated = p["updated_at"][:10] if p["updated_at"] else ""

        html += f"""
      <div class="product-card {new_class}" data-type="{p['type'] or '未分类'}" data-price="{p['price_min']}" data-created="{p['created_at']}" data-name="{p['title']}" data-best="{1 if p['is_best'] else 0}" data-top="{1 if p['is_top'] else 0}" data-new="{1 if p['is_new'] else 0}">
        {img_tag}
        <div class="product-info">
          {badge_html}
          <div class="product-title"><a href="{p['link']}" target="_blank">{p['title']}</a></div>
          <div class="product-meta">
            {price_html}
            <span class="product-type">{p['type']}</span>
          </div>
          <div class="product-tags">{tags_html}</div>
          <div class="product-detail">📦 {p['variant_count']} 变体 | 🖼 {p['image_count']} 图片 | 🏷 {p['vendor']}</div>
          <div class="product-detail" style="color:#e94560;font-weight:600">📅 上架: {published}</div>
          <div class="product-detail">创建: {created} | 更新: {updated}</div>
          {gallery_html}
          <div class="product-desc">{p['body']}</div>
          {variant_html}
        </div>
        <div class="product-footer">
          <span>SKU: {p['variants'][0]['sku'] if p['variants'] else 'N/A'}</span>
          <a class="product-link" href="{p['link']}" target="_blank">查看商品 →</a>
        </div>
      </div>
"""

    html += """    </div>
  </div>

  <!-- Table View -->
  <div class="section">
    <div class="section-title">📊 数据表格</div>
    <div style="overflow-x:auto;">
    <table class="table-view">
      <tr>
        <th>图片</th>
        <th>产品名称</th>
        <th>分类</th>
        <th>价格</th>
        <th>变体</th>
        <th>SKU</th>
        <th>状态</th>
        <th>创建时间</th>
        <th>链接</th>
      </tr>
"""

    for p in parsed:
        row_class = "row-new" if p["id"] in new_ids else ""
        img_cell = f'<img src="{p["main_image"]}">' if p["main_image"] else "无图"
        price_cell = f"${p['price_min']:.2f}"
        if p["price_min"] != p["price_max"]:
            price_cell = f"${p['price_min']:.2f}-${p['price_max']:.2f}"
        sku = p["variants"][0]["sku"] if p["variants"] else ""
        created = p["created_at"][:10] if p["created_at"] else ""

        html += f"""      <tr class="{row_class}">
        <td>{img_cell}</td>
        <td><a href="{p['link']}" target="_blank">{p['title']}</a></td>
        <td>{p['type']}</td>
        <td>{price_cell}</td>
        <td>{p['variant_count']}</td>
        <td>{sku}</td>
        <td>{p['status']}</td>
        <td>{created}</td>
        <td><a href="{p['link']}" target="_blank">🔗</a></td>
      </tr>
"""

    html += f"""    </table>
    </div>
  </div>

  <!-- Footer -->
  <div style="text-align:center; padding:20px; color:#999; font-size:12px;">
    Hermes Agent 竞品监控系统 | 抓取时间: {now_str} | 数据来源: {domain_clean}/products.json
  </div>
</div>

<script>
// 当前筛选状态
let currentType = 'all';
let currentTag = 'all';

function filterType(type) {{
  currentType = type;
  applyFilters();
  // 激活按钮
  const section = event.target.closest('.section');
  section.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
}}

function filterTag(tag) {{
  currentTag = tag;
  applyFilters();
  const section = event.target.closest('.section');
  section.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
}}

function applyFilters() {{
  const cards = document.querySelectorAll('.product-card');
  cards.forEach(card => {{
    let show = true;
    // 类型筛选
    if (currentType !== 'all' && card.dataset.type !== currentType) show = false;
    // 标签筛选
    if (currentTag === 'best' && card.dataset.best !== '1') show = false;
    if (currentTag === 'top' && card.dataset.top !== '1') show = false;
    if (currentTag === 'new' && card.dataset.new !== '1') show = false;
    card.style.display = show ? '' : 'none';
  }});
}}

function sortProducts(mode) {{
  const grid = document.getElementById('productGrid');
  const cards = Array.from(grid.querySelectorAll('.product-card'));

  // 激活按钮
  const section = event.target.closest('.section');
  section.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');

  cards.sort((a, b) => {{
    switch(mode) {{
      case 'price-asc':
        return parseFloat(a.dataset.price) - parseFloat(b.dataset.price);
      case 'price-desc':
        return parseFloat(b.dataset.price) - parseFloat(a.dataset.price);
      case 'newest':
        return (b.dataset.created || '').localeCompare(a.dataset.created || '');
      case 'name':
        return (a.dataset.name || '').localeCompare(b.dataset.name || '');
      default:
        return 0; // 保持原始顺序
    }}
  }});

  cards.forEach(card => grid.appendChild(card));
}}
</script>
</body>
</html>
"""

    html_path.write_text(html, encoding="utf-8")
    print(f"  💾 HTML: {html_path}")
    print(f"  📊 {total_products} 产品 | {types_count} 分类 | {len(new_ids)} 新品")
    return html_path


# ─── 通知报告 ───
def format_change_report(changes_list):
    """格式化新品检测报告"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = [f"🏪 Shopify 竞品监控报告 — {now}\n"]

    total_new = 0
    total_updated = 0

    for ch in changes_list:
        domain = ch["domain"]
        new = ch["new_products"]
        updated = ch["updated_products"]
        total_new += len(new)
        total_updated += len(updated)

        lines.append(f"\n{'='*50}")
        lines.append(f"📍 {domain}")
        lines.append(f"   上次检查: {ch['last_check'] or '首次'}")
        lines.append(f"   当前产品: {ch['current_count']} 个")

        if new:
            lines.append(f"\n   🆕 新品（{len(new)}个）:")
            for p in new:
                lines.append(f"      • {p['标题'][:50]} — ${p['最低价']:.2f}")
                lines.append(f"        {p['链接']}")

        if updated:
            lines.append(f"\n   🔄 更新（{len(updated)}个）:")
            for p in updated:
                price_change = ""
                if p["old_price"] != p["new_price"]:
                    price_change = f" (${p['old_price']:.2f} → ${p['new_price']:.2f})"
                lines.append(f"      • {p['title'][:50]}{price_change}")

        if ch["removed_count"]:
            lines.append(f"\n   🗑️ 下架: {ch['removed_count']} 个")

        if not new and not updated and not ch["removed_count"]:
            lines.append(f"   ✅ 无变化")

    lines.append(f"\n{'='*50}")
    lines.append(f"汇总: 🆕 {total_new} 新品 | 🔄 {total_updated} 更新")

    return "\n".join(lines)




# ================================================================
#  Section 5: Business Intelligence Kernel v5.0
# ================================================================

import difflib as _difflib
import statistics as _statistics
import math as _math
from collections import Counter as _Counter, defaultdict as _defaultdict


class PricingStrategyEngine:
    """定价策略引擎 — 分析竞品价格分布，建议最优定价"""

    @staticmethod
    def analyze_pricing(products):
        """分析产品价格分布 → 四分位/中位数/众数带/价格缺口"""
        prices = []
        for p in products:
            variants = p.get("variants", [])
            for v in variants:
                try:
                    pr = float(v.get("price", 0))
                    if pr > 0:
                        prices.append(pr)
                except (ValueError, TypeError):
                    continue

        if not prices:
            return {"error": "无有效价格数据", "count": 0}

        prices.sort()
        n = len(prices)
        q1 = prices[n // 4] if n >= 4 else prices[0]
        q2 = _statistics.median(prices)
        q3 = prices[3 * n // 4] if n >= 4 else prices[-1]
        mean = _statistics.mean(prices)
        stdev = _statistics.stdev(prices) if n > 1 else 0

        # 众数带 (10美元区间)
        bands = _Counter(int(p // 10) * 10 for p in prices)
        mode_band = bands.most_common(1)[0] if bands else (0, 0)

        # 价格缺口 (相邻价格差 > 50美元的区间)
        gaps = []
        sorted_unique = sorted(set(int(p) for p in prices))
        for i in range(len(sorted_unique) - 1):
            diff = sorted_unique[i + 1] - sorted_unique[i]
            if diff > 50:
                gaps.append({"from": sorted_unique[i], "to": sorted_unique[i + 1], "gap": diff})

        return {
            "count": n,
            "min": prices[0],
            "max": prices[-1],
            "mean": round(mean, 2),
            "median": round(q2, 2),
            "q1": round(q1, 2),
            "q3": round(q3, 2),
            "stdev": round(stdev, 2),
            "mode_band": f"${mode_band[0]}-${mode_band[0]+10}",
            "mode_band_count": mode_band[1],
            "price_gaps": gaps[:5],
            "premium_threshold": round(q3 + 1.5 * (q3 - q1), 2),
            "budget_threshold": round(max(0, q1 - 1.5 * (q3 - q1)), 2),
        }

    @staticmethod
    def suggest_price(cost, target_margin=0.67, competitor_prices=None):
        """建议零售价 — 基于成本+目标毛利+竞品参考"""
        floor_price = cost / (1 - target_margin)  # 最低可接受价
        ideal_price = cost * 3  # 3x markup

        result = {
            "cost": cost,
            "target_margin": f"{target_margin*100:.0f}%",
            "floor_price": round(floor_price, 2),
            "ideal_price": round(ideal_price, 2),
            "roas3x_min": round(cost + cost * 0.3 * 3, 2),  # cost + 3x ad spend
        }

        if competitor_prices:
            comp = [p for p in competitor_prices if p > 0]
            if comp:
                comp_median = _statistics.median(comp)
                result["competitor_median"] = round(comp_median, 2)
                result["competitor_range"] = f"${min(comp):.0f}-${max(comp):.0f}"
                # 建议: 竞品中位数 +5%（轻微溢价定位）
                result["suggested_price"] = round(comp_median * 1.05, 2)
                result["undercut_price"] = round(comp_median * 0.9, 2)
                result["premium_price"] = round(comp_median * 1.2, 2)
            else:
                result["suggested_price"] = round(ideal_price, 2)
        else:
            result["suggested_price"] = round(ideal_price, 2)

        return result

    @staticmethod
    def detect_price_anomalies(products):
        """检测价格异常值 (同品类内 >2σ)"""
        by_type = _defaultdict(list)
        for p in products:
            ptype = p.get("product_type", "unknown") or "unknown"
            variants = p.get("variants", [])
            for v in variants:
                try:
                    pr = float(v.get("price", 0))
                    if pr > 0:
                        by_type[ptype].append({"product": p.get("title", ""), "price": pr})
                except (ValueError, TypeError):
                    continue

        anomalies = []
        for ptype, items in by_type.items():
            if len(items) < 3:
                continue
            item_prices = [x["price"] for x in items]
            mean = _statistics.mean(item_prices)
            stdev = _statistics.stdev(item_prices)
            if stdev == 0:
                continue
            for item in items:
                z = abs(item["price"] - mean) / stdev
                if z > 2:
                    anomalies.append({
                        "title": item["product"],
                        "price": item["price"],
                        "category": ptype,
                        "z_score": round(z, 2),
                        "category_mean": round(mean, 2),
                        "deviation": f"{'高' if item['price'] > mean else '低'}于均值{abs(item['price']-mean):.0f}美元",
                    })

        return sorted(anomalies, key=lambda x: x["z_score"], reverse=True)


class CompetitorMatrix:
    """竞品对比矩阵 — 多店铺横向对比"""

    @staticmethod
    def _title_similarity(a, b):
        """标题相似度 (0-1)"""
        return _difflib.SequenceMatcher(None, a.lower(), b.lower()).ratio()

    def build_matrix(self, stores):
        """构建竞品矩阵 {domain: [products]} → 对比表"""
        matrix = {}
        for domain, products in stores.items():
            prices = []
            categories = set()
            for p in products:
                for v in p.get("variants", []):
                    try:
                        pr = float(v.get("price", 0))
                        if pr > 0:
                            prices.append(pr)
                    except:
                        pass
                cat = p.get("product_type", "")
                if cat:
                    categories.add(cat)

            matrix[domain] = {
                "product_count": len(products),
                "category_count": len(categories),
                "categories": sorted(categories),
                "avg_price": round(_statistics.mean(prices), 2) if prices else 0,
                "median_price": round(_statistics.median(prices), 2) if prices else 0,
                "price_range": f"${min(prices):.0f}-${max(prices):.0f}" if prices else "$0",
                "total_variants": sum(len(p.get("variants", [])) for p in products),
            }

        return matrix

    def overlap_analysis(self, store_a_products, store_b_products):
        """找相似产品 (标题相似度>0.6) + 价差分析"""
        overlaps = []
        for pa in store_a_products:
            title_a = pa.get("title", "")
            price_a = float(pa.get("variants", [{}])[0].get("price", 0)) if pa.get("variants") else 0
            for pb in store_b_products:
                title_b = pb.get("title", "")
                sim = self._title_similarity(title_a, title_b)
                if sim > 0.6:
                    price_b = float(pb.get("variants", [{}])[0].get("price", 0)) if pb.get("variants") else 0
                    overlaps.append({
                        "title_a": title_a,
                        "title_b": title_b,
                        "similarity": round(sim, 2),
                        "price_a": price_a,
                        "price_b": price_b,
                        "price_diff": round(price_a - price_b, 2),
                        "price_diff_pct": round((price_a - price_b) / price_b * 100, 1) if price_b > 0 else 0,
                    })

        overlaps.sort(key=lambda x: x["similarity"], reverse=True)
        return {
            "overlap_count": len(overlaps),
            "overlaps": overlaps[:20],
            "avg_price_diff": round(_statistics.mean([o["price_diff"] for o in overlaps]), 2) if overlaps else 0,
        }

    def market_position(self, domain, all_stores):
        """判断市场定位: premium / mid / budget"""
        if domain not in all_stores:
            return {"error": f"{domain} 不在数据中"}

        my_products = all_stores[domain]
        my_prices = []
        for p in my_products:
            for v in p.get("variants", []):
                try:
                    pr = float(v.get("price", 0))
                    if pr > 0:
                        my_prices.append(pr)
                except:
                    pass

        all_prices = []
        for d, products in all_stores.items():
            for p in products:
                for v in p.get("variants", []):
                    try:
                        pr = float(v.get("price", 0))
                        if pr > 0:
                            all_prices.append(pr)
                    except:
                        pass

        if not my_prices or not all_prices:
            return {"position": "unknown"}

        my_median = _statistics.median(my_prices)
        all_median = _statistics.median(all_prices)
        all_q1 = sorted(all_prices)[len(all_prices) // 4]
        all_q3 = sorted(all_prices)[3 * len(all_prices) // 4]

        if my_median >= all_q3:
            position = "premium"
        elif my_median <= all_q1:
            position = "budget"
        else:
            position = "mid-range"

        my_categories = set(p.get("product_type", "") for p in my_products if p.get("product_type"))
        all_categories = set()
        for d, products in all_stores.items():
            if d != domain:
                for p in products:
                    if p.get("product_type"):
                        all_categories.add(p["product_type"])

        unique_cats = my_categories - all_categories

        return {
            "domain": domain,
            "position": position,
            "my_median_price": round(my_median, 2),
            "market_median_price": round(all_median, 2),
            "price_index": round(my_median / all_median * 100, 1),  # 100=市场均价
            "my_categories": len(my_categories),
            "unique_categories": sorted(unique_cats),
            "category_coverage": round(len(my_categories) / max(len(all_categories), 1) * 100, 1),
        }


class ProfitCalculator:
    """利润计算器 — 单品/批量利润率 + ROAS阈值"""

    @staticmethod
    def calculate(price, cost, shipping=0, ad_spend_ratio=0.3, platform_fee=0.029, transaction_fee=0.3):
        """单品利润计算"""
        revenue = price
        total_cost = cost + shipping
        ad_cost = price * ad_spend_ratio
        platform_cost = price * platform_fee + transaction_fee
        profit = revenue - total_cost - ad_cost - platform_cost
        margin = profit / revenue if revenue > 0 else 0

        return {
            "price": price,
            "cost": cost,
            "shipping": shipping,
            "ad_cost": round(ad_cost, 2),
            "platform_fee": round(platform_cost, 2),
            "total_expense": round(total_cost + ad_cost + platform_cost, 2),
            "profit": round(profit, 2),
            "margin": f"{margin*100:.1f}%",
            "markup": f"{(price/cost - 1)*100:.0f}%" if cost > 0 else "N/A",
            "roas": round(revenue / ad_cost, 2) if ad_cost > 0 else float("inf"),
            "breakeven_units": max(1, _math.ceil(1000 / profit)) if profit > 0 else "∞",
        }

    @staticmethod
    def batch_calculate(products, cost_map=None, default_cost_ratio=0.33):
        """批量利润计算"""
        results = []
        for p in products:
            title = p.get("title", "")
            variants = p.get("variants", [])
            if not variants:
                continue
            price = float(variants[0].get("price", 0))
            if price <= 0:
                continue

            # 成本: 从 cost_map 查找，否则用默认比例
            if cost_map and title in cost_map:
                cost = cost_map[title]
            else:
                cost = price * default_cost_ratio

            calc = ProfitCalculator.calculate(price, cost)
            calc["title"] = title
            calc["handle"] = p.get("handle", "")
            results.append(calc)

        # 按利润排序
        results.sort(key=lambda x: x["profit"], reverse=True)
        return results

    @staticmethod
    def roas_threshold(cost, ad_spend):
        """ROAS >= 3x 所需最低售价"""
        # ROAS = revenue / ad_spend >= 3
        # revenue >= 3 * ad_spend
        # price >= 3 * ad_spend (假设单次转化)
        min_price = 3 * ad_spend
        # 还要覆盖成本
        min_price_with_cost = cost + 3 * ad_spend
        return {
            "min_price_roas3x": round(min_price, 2),
            "min_price_profitable": round(min_price_with_cost, 2),
            "recommended_price": round(min_price_with_cost * 1.2, 2),  # 20% buffer
            "ad_spend": ad_spend,
            "cost": cost,
        }


class MarketAnalyzer:
    """市场分析器 — 品类热力图/趋势信号/缺口发现"""

    @staticmethod
    def category_heatmap(products):
        """品类分布热力图 — 产品数/均价/价格区间"""
        cats = _defaultdict(lambda: {"count": 0, "prices": [], "products": []})
        for p in products:
            cat = p.get("product_type", "未分类") or "未分类"
            cats[cat]["count"] += 1
            cats[cat]["products"].append(p.get("title", ""))
            for v in p.get("variants", []):
                try:
                    pr = float(v.get("price", 0))
                    if pr > 0:
                        cats[cat]["prices"].append(pr)
                except:
                    pass

        heatmap = []
        for cat, data in cats.items():
            prices = data["prices"]
            heatmap.append({
                "category": cat,
                "product_count": data["count"],
                "avg_price": round(_statistics.mean(prices), 2) if prices else 0,
                "min_price": min(prices) if prices else 0,
                "max_price": max(prices) if prices else 0,
                "revenue_potential": round((_statistics.mean(prices) if prices else 0) * data["count"], 0),
            })

        heatmap.sort(key=lambda x: x["revenue_potential"], reverse=True)
        return {"categories": heatmap, "total_categories": len(heatmap)}

    @staticmethod
    def trend_signals(snapshots, days=30):
        """趋势信号 — 新品速度/价格方向/品类扩张
        snapshots: [{"date": "2025-01-01", "products": [...]}]
        """
        if len(snapshots) < 2:
            return {"error": "需要至少2个快照才能分析趋势"}

        # 按日期排序
        snapshots.sort(key=lambda x: x.get("date", ""))
        first = snapshots[0]
        last = snapshots[-1]

        first_ids = set(p.get("id") or p.get("handle") for p in first.get("products", []))
        last_ids = set(p.get("id") or p.get("handle") for p in last.get("products", []))

        new_products = last_ids - first_ids
        removed_products = first_ids - last_ids

        # 价格方向
        first_prices = []
        last_prices = []
        for p in first.get("products", []):
            for v in p.get("variants", []):
                try:
                    first_prices.append(float(v.get("price", 0)))
                except:
                    pass
        for p in last.get("products", []):
            for v in p.get("variants", []):
                try:
                    last_prices.append(float(v.get("price", 0)))
                except:
                    pass

        first_avg = _statistics.mean(first_prices) if first_prices else 0
        last_avg = _statistics.mean(last_prices) if last_prices else 0
        price_direction = "上涨" if last_avg > first_avg * 1.02 else ("下跌" if last_avg < first_avg * 0.98 else "稳定")

        # 品类变化
        first_cats = set(p.get("product_type", "") for p in first.get("products", []) if p.get("product_type"))
        last_cats = set(p.get("product_type", "") for p in last.get("products", []) if p.get("product_type"))
        new_cats = last_cats - first_cats

        return {
            "period": f"{first.get('date', '?')} → {last.get('date', '?')}",
            "new_product_count": len(new_products),
            "removed_product_count": len(removed_products),
            "net_growth": len(new_products) - len(removed_products),
            "new_product_velocity": round(len(new_products) / max(days, 1), 2),  # 每天新品数
            "price_direction": price_direction,
            "avg_price_change": f"${first_avg:.0f} → ${last_avg:.0f}",
            "new_categories": sorted(new_cats),
            "category_expansion": len(new_cats),
        }

    @staticmethod
    def gap_finder(my_products, competitor_products):
        """缺口发现 — 竞品有而我没有的品类/价格带"""
        my_cats = set(p.get("product_type", "") for p in my_products if p.get("product_type"))
        comp_cats = set(p.get("product_type", "") for p in competitor_products if p.get("product_type"))

        missing_cats = comp_cats - my_cats

        # 价格带分析
        my_prices = []
        comp_prices = []
        for p in my_products:
            for v in p.get("variants", []):
                try:
                    my_prices.append(float(v.get("price", 0)))
                except:
                    pass
        for p in competitor_products:
            for v in p.get("variants", []):
                try:
                    comp_prices.append(float(v.get("price", 0)))
                except:
                    pass

        # 按50美元区间分桶
        my_bands = set(int(p // 50) * 50 for p in my_prices if p > 0)
        comp_bands = set(int(p // 50) * 50 for p in comp_prices if p > 0)
        missing_bands = comp_bands - my_bands

        # 竞品在缺失品类中的产品
        gap_products = []
        for p in competitor_products:
            if p.get("product_type") in missing_cats:
                price = float(p.get("variants", [{}])[0].get("price", 0)) if p.get("variants") else 0
                gap_products.append({
                    "title": p.get("title", ""),
                    "category": p.get("product_type", ""),
                    "price": price,
                })

        gap_products.sort(key=lambda x: x["price"], reverse=True)

        return {
            "missing_categories": sorted(missing_cats),
            "missing_category_count": len(missing_cats),
            "missing_price_bands": [f"${b}-${b+50}" for b in sorted(missing_bands)],
            "gap_products": gap_products[:15],
            "opportunity_score": min(100, len(missing_cats) * 10 + len(missing_bands) * 5),
        }


# ─── 便捷入口 ───
def run_business_analysis(my_domain, competitor_domains, data_dir=None, stores_data=None, cost=100):
    """一键运行全部商业分析 → 结构化报告
    
    Args:
        stores_data: 直接传入 {domain: [products]} 跳过快照加载
        cost: 产品成本（默认100美元）
    """
    if data_dir is None:
        data_dir = DATA_DIR

    # 加载数据
    if stores_data:
        stores = stores_data
    else:
        stores = {}
        for domain in [my_domain] + competitor_domains:
            snapshot_path = Path(data_dir) / f"{normalize_domain(domain)}_snapshot.json"
            if snapshot_path.exists():
                snap = json.loads(snapshot_path.read_text(encoding="utf-8"))
                products = snap.get("products", {})
                # snapshot 存的是 {id: parsed_product}，需要还原
                if isinstance(products, dict):
                    stores[domain] = list(products.values())
                elif isinstance(products, list):
                    stores[domain] = products
                else:
                    stores[domain] = []
            else:
                print(f"  ⚠️ 无快照: {domain}")
                stores[domain] = []

    if not any(stores.values()):
        return {"error": "无数据，请先运行 scrape 命令"}

    # 执行分析
    pricing = PricingStrategyEngine()
    matrix = CompetitorMatrix()
    profit = ProfitCalculator()
    market = MarketAnalyzer()

    report = {
        "generated_at": datetime.now().isoformat(),
        "my_domain": my_domain,
        "competitors": competitor_domains,
    }

    # 1. 定价分析
    all_comp_products = []
    for d in competitor_domains:
        all_comp_products.extend(stores.get(d, []))

    report["pricing"] = {
        "my_pricing": pricing.analyze_pricing(stores.get(my_domain, [])),
        "competitor_pricing": pricing.analyze_pricing(all_comp_products),
        "price_suggestion": pricing.suggest_price(
            cost=cost,  # 用户传入成本
            target_margin=0.67,
            competitor_prices=[
                float(v.get("price", 0))
                for p in all_comp_products
                for v in p.get("variants", [])
                if float(v.get("price", 0)) > 0
            ][:100]
        ),
        "anomalies": pricing.detect_price_anomalies(all_comp_products)[:10],
    }

    # 2. 竞品矩阵
    report["competitor_matrix"] = matrix.build_matrix(stores)

    # 3. 市场定位
    if my_domain in stores and stores[my_domain]:
        report["market_position"] = matrix.market_position(my_domain, stores)
    else:
        report["market_position"] = {"note": "我方无数据，跳过定位分析"}

    # 4. 利润计算
    my_products = stores.get(my_domain, [])
    if my_products:
        report["profit_analysis"] = {
            "top_profit_products": profit.batch_calculate(my_products)[:10],
            "roas_threshold": profit.roas_threshold(cost=cost, ad_spend=cost * 0.3),
        }
    else:
        report["profit_analysis"] = {"note": "我方无产品数据"}

    # 5. 市场分析
    report["market"] = {
        "category_heatmap": market.category_heatmap(all_comp_products),
        "gaps": market.gap_finder(my_products, all_comp_products) if my_products else {"note": "需要我方数据"},
    }

    return report


def format_business_report(report):
    """格式化商业分析报告为可读文本"""
    lines = []
    lines.append("=" * 60)
    lines.append("📊 Shopify 商业智能分析报告 v5.0")
    lines.append(f"   生成时间: {report.get('generated_at', 'N/A')}")
    lines.append(f"   我方: {report.get('my_domain', 'N/A')}")
    lines.append(f"   竞品: {', '.join(report.get('competitors', []))}")
    lines.append("=" * 60)

    # 定价
    pricing = report.get("pricing", {})
    if "my_pricing" in pricing:
        mp = pricing["my_pricing"]
        lines.append(f"\n💰 我方定价分布:")
        lines.append(f"   中位数: ${mp.get('median', 0)} | 均值: ${mp.get('mean', 0)}")
        lines.append(f"   区间: ${mp.get('min', 0)}-${mp.get('max', 0)} | σ={mp.get('stdev', 0)}")

    if "competitor_pricing" in pricing:
        cp = pricing["competitor_pricing"]
        lines.append(f"\n💰 竞品定价分布:")
        lines.append(f"   中位数: ${cp.get('median', 0)} | 均值: ${cp.get('mean', 0)}")
        lines.append(f"   众数带: {cp.get('mode_band', 'N/A')} ({cp.get('mode_band_count', 0)}个)")
        if cp.get("price_gaps"):
            gap_strs = [f"${g['from']}-${g['to']}" for g in cp['price_gaps'][:3]]
            lines.append(f"   价格缺口: {', '.join(gap_strs)}")

    if "price_suggestion" in pricing:
        ps = pricing["price_suggestion"]
        lines.append(f"\n🎯 定价建议 (成本${ps.get('cost', 100)}):")
        lines.append(f"   建议售价: ${ps.get('suggested_price', 0)}")
        lines.append(f"   底价: ${ps.get('floor_price', 0)} | 溢价: ${ps.get('premium_price', 0)}")
        lines.append(f"   ROAS 3x 最低价: ${ps.get('roas3x_min', 0)}")

    # 竞品矩阵
    cm = report.get("competitor_matrix", {})
    if cm:
        lines.append(f"\n📋 竞品矩阵:")
        for domain, info in cm.items():
            lines.append(f"   {domain}: {info.get('product_count', 0)}品 | "
                        f"均价${info.get('avg_price', 0)} | "
                        f"{info.get('category_count', 0)}品类")

    # 市场定位
    mp = report.get("market_position", {})
    if mp.get("position"):
        lines.append(f"\n🏷️ 市场定位: {mp['position'].upper()}")
        lines.append(f"   价格指数: {mp.get('price_index', 0)} (100=市场均价)")
        if mp.get("unique_categories"):
            lines.append(f"   独有品类: {', '.join(mp['unique_categories'][:5])}")

    # 利润
    pa = report.get("profit_analysis", {})
    if pa.get("top_profit_products"):
        lines.append(f"\n💵 TOP利润产品:")
        for i, p in enumerate(pa["top_profit_products"][:5], 1):
            lines.append(f"   {i}. {p.get('title', '')[:40]} — "
                        f"利润${p.get('profit', 0)} ({p.get('margin', '0%')})")

    if pa.get("roas_threshold"):
        rt = pa["roas_threshold"]
        lines.append(f"\n📈 ROAS 3x 阈值: 最低售价 ${rt.get('min_price_profitable', 0)}")

    # 市场缺口
    mkt = report.get("market", {})
    gaps = mkt.get("gaps", {})
    if gaps.get("missing_categories"):
        lines.append(f"\n🔍 市场缺口 ({gaps.get('opportunity_score', 0)}分):")
        lines.append(f"   缺失品类: {', '.join(gaps['missing_categories'][:5])}")
        if gaps.get("missing_price_bands"):
            lines.append(f"   缺失价格带: {', '.join(gaps['missing_price_bands'][:5])}")

    lines.append(f"\n{'=' * 60}")
    return "\n".join(lines)



# ─── CLI ───
def monitor_main():
    parser = argparse.ArgumentParser(description="Shopify 竞品监控")
    sub = parser.add_subparsers(dest="command")

    # scrape
    p_scrape = sub.add_parser("scrape", help="抓取全部产品")
    p_scrape.add_argument("domains", nargs="+", help="竞品域名")

    # check
    p_check = sub.add_parser("check", help="检测新品（对比快照）")
    p_check.add_argument("domains", nargs="+", help="竞品域名")

    # auto
    p_auto = sub.add_parser("auto", help="抓取 + 检测一步到位")
    p_auto.add_argument("domains", nargs="+", help="竞品域名")

    # export
    p_export = sub.add_parser("export", help="导出 Excel + HTML")
    p_export.add_argument("domains", nargs="+", help="竞品域名")

    # html
    p_html = sub.add_parser("html", help="导出专业 HTML 报告")
    p_html.add_argument("domains", nargs="+", help="竞品域名")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        print("\n💡 示例:")
        print("  python3 shopify_ultimate.py scrape competitor.com")
        print("  python3 shopify_ultimate.py check competitor.com")
        print("  python3 shopify_ultimate.py auto competitor.com competitor2.com")
        print("  python3 shopify_ultimate.py export competitor.com")
        sys.exit(0)

    domains = [normalize_domain(d) for d in args.domains]

    if not domains:
        if DEFAULT_COMPETITORS:
            domains = DEFAULT_COMPETITORS
            print(f"📋 使用默认竞品: {', '.join(domains)}")
        else:
            print("❌ 请提供竞品域名")
            sys.exit(1)

    session = get_session()
    # v3: Trajectory
    traj = reset_logger("shopify_monitor") if HAS_GUARD else None

    for domain in domains:
        print(f"\n🔍 {domain}")
        t0 = time.time()

        if args.command == "scrape":
            products = fetch_all_products(domain, session)
            if products:
                save_snapshot(domain, [parse_product(p) for p in products])
                print(f"  ✅ {len(products)} 个产品已保存快照")
                if traj:
                    traj.emit_span("scrape_done", {"domain": domain, "count": len(products), "duration": round(time.time()-t0, 2)})

        elif args.command == "check":
            products = fetch_all_products(domain, session)
            if products:
                changes = detect_changes(domain, products)
                print(format_change_report([changes]))
                save_snapshot(domain, [parse_product(p) for p in products])
                if traj:
                    traj.emit_span("check_done", {"domain": domain, "changes": len(changes.get("new", [])), "duration": round(time.time()-t0, 2)})

        elif args.command == "auto":
            products = fetch_all_products(domain, session)
            if products:
                changes = detect_changes(domain, products)
                parsed = [parse_product(p) for p in products]
                save_snapshot(domain, parsed)
                print(format_change_report([changes]))
                if traj:
                    traj.emit_span("auto_done", {"domain": domain, "products": len(products), "duration": round(time.time()-t0, 2)})

        elif args.command == "export":
            products = fetch_all_products(domain, session)
            if products:
                collections = fetch_collections(domain, session)
                shop_info = fetch_shop_info(domain, session)
                monitor_export_excel(domain, products, collections, shop_info)
                monitor_export_html(domain, products, collections, shop_info)
                if traj:
                    traj.emit_span("export_done", {"domain": domain, "products": len(products), "duration": round(time.time()-t0, 2)})

        elif args.command == "html":
            products = fetch_all_products(domain, session)
            if products:
                changes = detect_changes(domain, products)
                monitor_export_html(domain, products, changes=changes)
                save_snapshot(domain, [parse_product(p) for p in products])
                if traj:
                    traj.emit_span("html_done", {"domain": domain, "duration": round(time.time()-t0, 2)})

    # v3: 保存轨迹
    if traj:
        traj.set_meta(command=args.command, domains=len(domains))
        traj.save()



# ================================================================
#  Section 2: Product Discovery (from shopify_discovery)
# ================================================================

def search_products(keyword, max_products=30):
    """搜索 Shopify 产品，支持自动翻页"""
    all_products = []
    all_brands = []
    seen_upids = set()

    # 首次搜索
    print(f"🔍 搜索: {keyword}")
    r = requests.get(f"{SELOFY_BASE_URL}/search", params={"q": keyword}, headers=SELOFY_HEADERS, timeout=20)
    if r.status_code != 200:
        print(f"❌ 搜索失败: {r.status_code}")
        return [], []

    data = r.json()
    session_id = data.get("searchSessionId", "")

    for p in data.get("products", []):
        if p["upid"] not in seen_upids:
            seen_upids.add(p["upid"])
            all_products.append(p)

    for b in data.get("brands", []):
        if b not in all_brands:
            all_brands.append(b)

    meta = data.get("meta", {})
    print(f"   首页返回: {meta.get('totalProducts', 0)} 个产品")

    # 翻页加载更多
    if session_id and len(all_products) < max_products:
        for page in range(10):  # 最多翻10页
            more_state = data.get("moreState", {})
            can_more = more_state.get("canMore", False)
            status = more_state.get("status", "")

            if not can_more and status != "preparing":
                break

            time.sleep(random.uniform(1.5, 3.0))

            try:
                r2 = requests.post(f"{SELOFY_BASE_URL}/more", json={
                    "q": keyword,
                    "searchSessionId": session_id
                }, headers={**SELOFY_HEADERS, "Content-Type": "application/json"}, timeout=20)

                if r2.status_code != 200:
                    break

                data2 = r2.json()
                if not data2.get("ok"):
                    break

                new_products = data2.get("addedProducts", [])
                if not new_products:
                    # 检查是否还在准备中
                    if data2.get("moreState", {}).get("status") == "preparing":
                        time.sleep(2)
                        continue
                    break

                for p in new_products:
                    if p["upid"] not in seen_upids:
                        seen_upids.add(p["upid"])
                        all_products.append(p)

                for b in data2.get("brands", []):
                    if b not in all_brands:
                        all_brands.append(b)

                print(f"   翻页 {page+1}: 累计 {len(all_products)} 个产品")

                if len(all_products) >= max_products:
                    break

                # 更新 session 状态
                data = data2

            except Exception as e:
                print(f"   翻页异常: {e}")
                break

    # 只保留 max_products 个
    all_products = all_products[:max_products]
    print(f"✅ 共获取: {len(all_products)} 个产品, {len(all_brands)} 个品牌")
    return all_products, all_brands


def get_inspire():
    """获取随机搜索推荐词"""
    r = requests.get(f"{SELOFY_BASE_URL}/inspire", headers=SELOFY_HEADERS, timeout=10)
    if r.status_code == 200:
        return r.json().get("term", "")
    return ""


# ─── 数据处理 ───────────────────────────────────────────────
def flatten_product(p):
    """把产品数据扁平化为一行字典"""
    brand = p.get("brand", {})
    price = p.get("priceRange", {})
    rating = p.get("rating", {})

    # 提取属性
    attrs = {}
    for a in p.get("attributes", []):
        attrs[a["name"]] = ", ".join(a.get("values", []))

    return {
        "标题": p.get("title", ""),
        "品牌": brand.get("name", "") if isinstance(brand, dict) else str(brand),
        "品牌域名": brand.get("domain", "") if isinstance(brand, dict) else "",
        "描述": p.get("description", ""),
        "独特卖点": p.get("uniqueSellingPoint", ""),
        "卖点1": p.get("topFeatures", [""])[0] if len(p.get("topFeatures", [])) > 0 else "",
        "卖点2": p.get("topFeatures", [""])[1] if len(p.get("topFeatures", [])) > 1 else "",
        "卖点3": p.get("topFeatures", [""])[2] if len(p.get("topFeatures", [])) > 2 else "",
        "卖点4": p.get("topFeatures", [""])[3] if len(p.get("topFeatures", [])) > 3 else "",
        "卖点5": p.get("topFeatures", [""])[4] if len(p.get("topFeatures", [])) > 4 else "",
        "技术参数": " | ".join(p.get("techSpecs", [])),
        "最低价": price.get("min", {}).get("amount", "") if isinstance(price, dict) else "",
        "最高价": price.get("max", {}).get("amount", "") if isinstance(price, dict) else "",
        "币种": price.get("min", {}).get("currencyCode", "") if isinstance(price, dict) else "",
        "评分": rating.get("average", "") if isinstance(rating, dict) else "",
        "评分数": rating.get("count", "") if isinstance(rating, dict) else "",
        "兼容设备": attrs.get("Compatible device", ""),
        "电源类型": attrs.get("Power source", ""),
        "颜色": attrs.get("Color", ""),
        "连接方式": attrs.get("Connection type", ""),
        "产品图": "; ".join([m.get("url", "") for m in p.get("media", [])[:3]]),
        "UPID": p.get("upid", ""),
    }


# ─── 导出 Excel ────────────────────────────────────────────
def discovery_discovery_export_excel(products, brands, keyword, filepath):
    """导出为 Excel（两个 sheet：产品 + 品牌）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠️ openpyxl 未安装，跳过 Excel 导出")
        return

    wb = Workbook()

    # ── Sheet 1: 产品列表 ──
    ws1 = wb.active
    ws1.title = "产品列表"

    rows = [flatten_product(p) for p in products]
    if not rows:
        print("⚠️ 无产品数据")
        return

    headers = list(rows[0].keys())

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 写表头
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 写数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=str(row_data.get(h, "")))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 自动列宽
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # 冻结首行
    ws1.freeze_panes = "A2"

    # ── Sheet 2: 品牌列表 ──
    ws2 = wb.create_sheet("品牌列表")
    brand_headers = ["品牌名", "域名", "产品数"]
    for col, h in enumerate(brand_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, b in enumerate(brands, 2):
        ws2.cell(row=row_idx, column=1, value=b.get("name", ""))
        ws2.cell(row=row_idx, column=2, value=b.get("domain", ""))
        ws2.cell(row=row_idx, column=3, value=b.get("productCount", 0))

    wb.save(filepath)
    print(f"📊 Excel 已保存: {filepath}")


# ─── 导出 HTML ─────────────────────────────────────────────
def discovery_discovery_export_html(products, brands, keyword, filepath):
    """导出为专业 HTML 报告"""
    rows = [flatten_product(p) for p in products]
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 产品卡片 HTML
    cards_html = ""
    for i, p in enumerate(products):
        flat = rows[i]
        img_url = ""
        for m in p.get("media", []):
            if m.get("url"):
                img_url = m["url"]
                break

        price_str = ""
        if flat["最低价"]:
            price_str = f'{flat["币种"]} {flat["最低价"]}'
            if flat["最高价"] and flat["最高价"] != flat["最低价"]:
                price_str += f' - {flat["币种"]} {flat["最高价"]}'

        rating_str = ""
        if flat["评分"]:
            rating_str = f'⭐ {flat["评分"]} ({flat["评分数"]} reviews)'

        features_html = ""
        for fi in range(1, 6):
            feat = flat.get(f"卖点{fi}", "")
            if feat:
                features_html += f'<li>{feat}</li>\n'

        brand_link = ""
        if flat["品牌域名"]:
            brand_link = f'<a href="{flat["品牌域名"]}" target="_blank">{flat["品牌"]}</a>'
        else:
            brand_link = flat["品牌"]

        cards_html += f"""
        <div class="product-card" data-price="{flat['最低价']}" data-rating="{flat['评分']}">
          <div class="card-img">
            <img src="{img_url}" alt="{flat['标题']}" loading="lazy"
                 onerror="this.src='data:image/svg+xml,<svg xmlns=%22http://www.w3.org/2000/svg%22 viewBox=%220 0 200 200%22><rect fill=%22%23f0f0f0%22 width=%22200%22 height=%22200%22/><text fill=%22%23999%22 x=%2250%25%22 y=%2250%25%22 text-anchor=%22middle%22 dy=%22.3em%22 font-size=%2214%22>No Image</text></svg>'">
          </div>
          <div class="card-body">
            <h3 class="card-title">{flat['标题']}</h3>
            <div class="card-meta">
              <span class="brand">{brand_link}</span>
              <span class="price">{price_str}</span>
              <span class="rating">{rating_str}</span>
            </div>
            <p class="card-desc">{flat['独特卖点'] or flat['描述'][:120]}</p>
            <ul class="features">{features_html}</ul>
            <div class="card-specs">{flat['技术参数']}</div>
          </div>
        </div>
        """

    # 品牌列表 HTML
    brands_html = ""
    for b in brands:
        brands_html += f"""
        <tr>
          <td>{b.get('name', '')}</td>
          <td><a href="{b.get('domain', '')}" target="_blank">{b.get('domain', '')}</a></td>
          <td>{b.get('productCount', 0)}</td>
        </tr>
        """

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Shopify 选品发现 - {keyword}</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f7fa; color: #333; }}
  .header {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%); color: white; padding: 32px 40px; }}
  .header h1 {{ font-size: 28px; margin-bottom: 8px; }}
  .header .meta {{ color: #8892b0; font-size: 14px; }}
  .header .meta span {{ margin-right: 20px; }}
  .toolbar {{ background: white; padding: 16px 40px; border-bottom: 1px solid #e2e8f0; display: flex; gap: 12px; align-items: center; flex-wrap: wrap; }}
  .toolbar input {{ padding: 8px 16px; border: 1px solid #e2e8f0; border-radius: 8px; width: 300px; font-size: 14px; }}
  .toolbar select {{ padding: 8px 12px; border: 1px solid #e2e8f0; border-radius: 8px; font-size: 14px; }}
  .toolbar button {{ padding: 8px 20px; border: none; border-radius: 8px; cursor: pointer; font-size: 14px; }}
  .btn-filter {{ background: #e2e8f0; color: #333; }}
  .btn-filter.active {{ background: #2f5496; color: white; }}
  .section {{ padding: 24px 40px; }}
  .section h2 {{ font-size: 20px; margin-bottom: 16px; color: #1a1a2e; }}
  .products-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(380px, 1fr)); gap: 20px; }}
  .product-card {{ background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1); transition: transform 0.2s, box-shadow 0.2s; }}
  .product-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.15); }}
  .card-img {{ height: 240px; background: #f8f9fa; display: flex; align-items: center; justify-content: center; overflow: hidden; }}
  .card-img img {{ max-width: 100%; max-height: 100%; object-fit: contain; }}
  .card-body {{ padding: 20px; }}
  .card-title {{ font-size: 16px; font-weight: 600; margin-bottom: 8px; line-height: 1.4; }}
  .card-meta {{ display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 12px; font-size: 13px; }}
  .card-meta .brand {{ color: #2f5496; }}
  .card-meta .brand a {{ color: #2f5496; text-decoration: none; }}
  .card-meta .brand a:hover {{ text-decoration: underline; }}
  .card-meta .price {{ color: #e74c3c; font-weight: 600; }}
  .card-meta .rating {{ color: #f39c12; }}
  .card-desc {{ font-size: 13px; color: #666; margin-bottom: 12px; line-height: 1.5; }}
  .features {{ padding-left: 18px; margin-bottom: 12px; }}
  .features li {{ font-size: 12px; color: #555; margin-bottom: 4px; line-height: 1.4; }}
  .card-specs {{ font-size: 11px; color: #999; border-top: 1px solid #f0f0f0; padding-top: 10px; }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; }}
  th {{ background: #2f5496; color: white; padding: 12px 16px; text-align: left; font-size: 13px; }}
  td {{ padding: 10px 16px; border-bottom: 1px solid #f0f0f0; font-size: 13px; }}
  td a {{ color: #2f5496; text-decoration: none; }}
  tr:hover {{ background: #f8f9fa; }}
  .stats-bar {{ display: flex; gap: 24px; margin-bottom: 20px; }}
  .stat-card {{ background: white; padding: 16px 24px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  .stat-card .num {{ font-size: 28px; font-weight: 700; color: #2f5496; }}
  .stat-card .label {{ font-size: 12px; color: #999; margin-top: 4px; }}
</style>
</head>
<body>
<div class="header">
  <h1>🔍 Shopify 选品发现报告</h1>
  <div class="meta">
    <span>关键词: <strong>{keyword}</strong></span>
    <span>时间: {now}</span>
    <span>数据源: Shopify Official Catalog API</span>
  </div>
</div>

<div class="section">
  <div class="stats-bar">
    <div class="stat-card"><div class="num">{len(products)}</div><div class="label">产品数</div></div>
    <div class="stat-card"><div class="num">{len(brands)}</div><div class="label">品牌数</div></div>
    <div class="stat-card"><div class="num">{len([p for p in rows if p['最低价']])}</div><div class="label">有价格</div></div>
    <div class="stat-card"><div class="num">{len([p for p in rows if p['评分']])}</div><div class="label">有评分</div></div>
  </div>
</div>

<div class="toolbar">
  <input type="text" id="searchInput" placeholder="🔍 筛选产品..." oninput="filterProducts()">
  <select id="sortSelect" onchange="sortProducts()">
    <option value="">默认排序</option>
    <option value="price-asc">价格 ↑</option>
    <option value="price-desc">价格 ↓</option>
    <option value="rating-desc">评分 ↓</option>
  </select>
</div>

<div class="section">
  <h2>📦 产品列表</h2>
  <div class="products-grid" id="productsGrid">
    {cards_html}
  </div>
</div>

<div class="section">
  <h2>🏢 品牌列表</h2>
  <table>
    <thead><tr><th>品牌名</th><th>独立站</th><th>产品数</th></tr></thead>
    <tbody>{brands_html}</tbody>
  </table>
</div>

<script>
function filterProducts() {{
  const q = document.getElementById('searchInput').value.toLowerCase();
  document.querySelectorAll('.product-card').forEach(card => {{
    card.style.display = card.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}
function sortProducts() {{
  const v = document.getElementById('sortSelect').value;
  const grid = document.getElementById('productsGrid');
  const cards = [...grid.querySelectorAll('.product-card')];
  cards.sort((a, b) => {{
    if (v === 'price-asc') return (parseFloat(a.dataset.price)||999) - (parseFloat(b.dataset.price)||999);
    if (v === 'price-desc') return (parseFloat(b.dataset.price)||0) - (parseFloat(a.dataset.price)||0);
    if (v === 'rating-desc') return (parseFloat(b.dataset.rating)||0) - (parseFloat(a.dataset.rating)||0);
    return 0;
  }});
  cards.forEach(c => grid.appendChild(c));
}}
</script>
</body>
</html>"""

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"📄 HTML 已保存: {filepath}")


# ─── Shopify 验证（选品日报集成）────────────────────────────
def verify_shopify(keyword, max_products=15):
    """
    验证一个关键词在 Shopify 上的竞争情况。
    返回结构化字典，可直接写入选品报告 Excel。
    """
    products, brands = search_products(keyword, max_products)

    if not products:
        return {
            "shopify_competitors": 0,
            "shopify_price_min": 0.0,
            "shopify_price_max": 0.0,
            "shopify_price_avg": 0.0,
            "shopify_stores": "",
            "shopify_brands": "",
            "shopify_product_count": 0,
            "shopify_top_products": "",
            "shopify_verify_status": "无结果",
        }

    prices = []
    store_links = []
    brand_names = []
    top_titles = []

    for p in products:
        flat = flatten_product(p)
        # 价格
        if flat["最低价"]:
            try:
                prices.append(float(flat["最低价"]))
            except (ValueError, TypeError):
                pass
        # 品牌域名
        if flat["品牌域名"]:
            store_links.append(flat["品牌域名"])
        # 品牌名
        if flat["品牌"]:
            brand_names.append(flat["品牌"])
        # 产品标题
        if flat["标题"]:
            top_titles.append(flat["标题"][:60])

    price_min = round(min(prices), 2) if prices else 0.0
    price_max = round(max(prices), 2) if prices else 0.0
    price_avg = round(sum(prices) / len(prices), 2) if prices else 0.0

    # 去重
    store_links = list(dict.fromkeys(store_links))
    brand_names = list(dict.fromkeys(brand_names))

    return {
        "shopify_competitors": len(products),
        "shopify_price_min": price_min,
        "shopify_price_max": price_max,
        "shopify_price_avg": price_avg,
        "shopify_stores": " | ".join(store_links[:10]),
        "shopify_brands": " | ".join(brand_names[:10]),
        "shopify_product_count": len(products),
        "shopify_top_products": " | ".join(top_titles[:5]),
        "shopify_verify_status": "✅ 已验证",
    }


def print_verify_result(keyword, result):
    """打印验证结果"""
    c = result["shopify_competitors"]
    if c == 0:
        print(f"  ❌ '{keyword}' → Shopify 无竞品（蓝海？）")
        return

    status = "🔴 红海" if c >= 10 else "🟡 有竞争" if c >= 5 else "🟢 蓝海机会"
    price_range = ""
    if result["shopify_price_min"] > 0:
        price_range = f"${result['shopify_price_min']}~${result['shopify_price_max']} (均${result['shopify_price_avg']})"

    print(f"  {status} '{keyword}' → {c} 个竞品 | {price_range}")
    if result["shopify_brands"]:
        brands = result["shopify_brands"].split(" | ")[:5]
        print(f"      品牌: {', '.join(brands)}")
    if result["shopify_stores"]:
        stores = result["shopify_stores"].split(" | ")[:3]
        for s in stores:
            print(f"      🔗 {s}")


def verify_batch_xlsx(input_path, top_n=0):
    """
    批量验证 Excel 文件中的产品。
    读取选品报告，逐行用产品标题/关键词搜索 Shopify，追加验证列到新文件。
    """
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("❌ 需要 openpyxl: pip3 install openpyxl")
        sys.exit(1)

    if not os.path.exists(input_path):
        print(f"❌ 文件不存在: {input_path}")
        sys.exit(1)

    print(f"📂 读取: {input_path}")
    wb = load_workbook(input_path)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers) if h}

    # 找标题列（优先中文列名）
    title_col = None
    for name in ["标题", "产品标题", "Product", "Title", "Product Title", "商品名称", "关键词", "Keyword", "Product Name"]:
        if name in header_map:
            title_col = header_map[name]
            break

    if title_col is None:
        print(f"❌ 找不到标题列。现有列: {headers[:15]}...")
        sys.exit(1)

    # 新列名
    new_cols = [
        "Shopify搜索词", "Shopify竞品数", "Shopify最低价", "Shopify最高价", "Shopify均价",
        "Shopify品牌", "Shopify店铺", "Shopify验证状态"
    ]

    # 写入位置：已有列用旧位置覆盖，新列追加到末尾
    existing_new = [c for c in new_cols if c in header_map]
    if existing_new:
        print(f"⚠️ 已有验证列: {existing_new}，直接覆盖旧数据")

    new_col_indices = {}
    append_col = len(headers) + 1
    for col_name in new_cols:
        if col_name in header_map:
            # 已存在 → 用已有列（header_map 是 0-based，+1 转 Excel 列号）
            new_col_indices[col_name] = header_map[col_name] + 1
        else:
            # 新列 → 追加
            cell = ws.cell(row=1, column=append_col, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            new_col_indices[col_name] = append_col
            append_col += 1

    # 遍历数据行
    total_rows = ws.max_row - 1
    if top_n > 0:
        total_rows = min(total_rows, top_n)

    print(f"📊 共 {total_rows} 行待验证\n")

    verified = 0
    errors = 0

    for row_idx in range(2, 2 + total_rows):
        title_cell = ws.cell(row=row_idx, column=title_col + 1)
        title = str(title_cell.value or "").strip()

        if not title or title == "None":
            continue

        # 从标题提取搜索词（跳过品牌名，取产品品类词）
        words = title.split()
        search_words = []
        skip_words = {
            "for", "with", "and", "the", "a", "an", "in", "on", "at", "to", "of", "by",
            "-", "&", "|", "®", "™", "plus", "pro", "max", "mini", "lite", "gen",
            "new", "upgraded", "2024", "2025", "2026", "pack", "set"
        }
        # 跳过前1-2个词（通常是品牌名），从第3个词开始取
        start = min(2, len(words) // 3)
        for w in words[start:]:
            if len(search_words) >= 4:
                break
            clean = w.strip("(),.!?\"':;®™#").lower()
            if clean and clean not in skip_words and len(clean) > 1:
                search_words.append(w.strip("(),.!?\"':;®™#"))
        keyword = " ".join(search_words)

        if not keyword:
            # 回退：取全部有意义词
            keyword = " ".join(w.strip("(),.!?\"':;®™#") for w in words[:5] if len(w) > 2)

        row_num = row_idx
        print(f"  [{row_idx - 1}/{total_rows + 1}] {keyword[:50]}...", end=" ")

        try:
            # 带重试的验证（429 限流自动重试）
            result = None
            for attempt in range(3):
                result = verify_shopify(keyword, max_products=10)
                if result["shopify_competitors"] > 0 or attempt == 2:
                    break
                wait = random.uniform(5, 10)
                print(f"⏳ 可能限流，等待 {wait:.0f}s 重试...")
                time.sleep(wait)

            # 写入数据
            ws.cell(row=row_num, column=new_col_indices["Shopify搜索词"], value=keyword)
            ws.cell(row=row_num, column=new_col_indices["Shopify竞品数"], value=result["shopify_competitors"])
            ws.cell(row=row_num, column=new_col_indices["Shopify最低价"], value=result["shopify_price_min"])
            ws.cell(row=row_num, column=new_col_indices["Shopify最高价"], value=result["shopify_price_max"])
            ws.cell(row=row_num, column=new_col_indices["Shopify均价"], value=result["shopify_price_avg"])
            ws.cell(row=row_num, column=new_col_indices["Shopify品牌"], value=result["shopify_brands"])
            ws.cell(row=row_num, column=new_col_indices["Shopify店铺"], value=result["shopify_stores"])
            ws.cell(row=row_num, column=new_col_indices["Shopify验证状态"], value=result["shopify_verify_status"])

            c = result["shopify_competitors"]
            tag = "🔴" if c >= 10 else "🟡" if c >= 5 else "🟢" if c > 0 else "⚪"
            print(f"{tag} {c}竞品")
            verified += 1

        except Exception as e:
            print(f"❌ 错误: {e}")
            ws.cell(row=row_num, column=new_col_indices["Shopify验证状态"], value=f"❌ {str(e)[:30]}")
            errors += 1

        # 防封延迟
        time.sleep(random.uniform(1.5, 3.0))

    # 保存到新文件
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(DISCOVERY_OUTPUT_DIR, f"{base_name}_shopify_verified_{ts}.xlsx")
    os.makedirs(DISCOVERY_OUTPUT_DIR, exist_ok=True)
    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"✅ 验证完成")
    print(f"   验证: {verified} 行")
    print(f"   错误: {errors} 行")
    print(f"   输出: {output_path}")
    print(f"{'='*60}")

    return output_path


# ─── 主入口 ────────────────────────────────────────────────
def discovery_main():
    os.makedirs(DISCOVERY_OUTPUT_DIR, exist_ok=True)

    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    cmd = sys.argv[1]

    if cmd == "inspire":
        for _ in range(5):
            word = get_inspire()
            if word:
                print(f"  💡 {word}")
            time.sleep(1)

    elif cmd == "search":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py search '关键词' [--max N] [--export]")
            sys.exit(1)

        keyword = sys.argv[2]
        max_products = 30
        do_export = False

        for i, arg in enumerate(sys.argv):
            if arg == "--max" and i + 1 < len(sys.argv):
                max_products = int(sys.argv[i + 1])
            if arg == "--export":
                do_export = True

        products, brands = search_products(keyword, max_products)

        if not products:
            print("❌ 未找到产品")
            sys.exit(1)

        # 打印摘要
        print(f"\n{'='*60}")
        print(f"📊 搜索结果: '{keyword}'")
        print(f"{'='*60}")
        for i, p in enumerate(products[:10], 1):
            flat = flatten_product(p)
            price = f"{flat['币种']} {flat['最低价']}" if flat['最低价'] else "询价"
            print(f"  {i:2d}. {flat['标题'][:50]}")
            print(f"      品牌: {flat['品牌']} | 价格: {price} | 评分: {flat['评分'] or 'N/A'}")
            if flat['品牌域名']:
                print(f"      🔗 {flat['品牌域名']}")

        if len(products) > 10:
            print(f"  ... 还有 {len(products) - 10} 个产品")

        # 品牌汇总
        print(f"\n🏢 品牌 ({len(brands)}):")
        for b in brands[:10]:
            print(f"  • {b['name']} → {b.get('domain', 'N/A')}")

        # 导出
        if do_export:
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            safe_kw = keyword.replace(" ", "_")[:30]
            xlsx_path = os.path.join(DISCOVERY_OUTPUT_DIR, f"shopify_{safe_kw}_{ts}.xlsx")
            html_path = os.path.join(DISCOVERY_OUTPUT_DIR, f"shopify_{safe_kw}_{ts}.html")
            discovery_export_excel(products, brands, keyword, xlsx_path)
            discovery_export_html(products, brands, keyword, html_path)

    elif cmd == "batch":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py batch 'kw1,kw2,kw3' [--max N] [--export]")
            sys.exit(1)

        keywords = [k.strip() for k in sys.argv[2].split(",")]
        max_per_kw = 20
        do_export = True

        for i, arg in enumerate(sys.argv):
            if arg == "--max" and i + 1 < len(sys.argv):
                max_per_kw = int(sys.argv[i + 1])

        all_products = []
        all_brands = []
        seen_upids = set()

        for kw in keywords:
            products, brands = search_products(kw, max_per_kw)
            for p in products:
                if p["upid"] not in seen_upids:
                    seen_upids.add(p["upid"])
                    all_products.append(p)
            for b in brands:
                if b not in all_brands:
                    all_brands.append(b)
            time.sleep(random.uniform(2, 4))

        print(f"\n📊 批量搜索完成: {len(keywords)} 个关键词, {len(all_products)} 个去重产品, {len(all_brands)} 个品牌")

        if do_export:
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            xlsx_path = os.path.join(DISCOVERY_OUTPUT_DIR, f"shopify_batch_{ts}.xlsx")
            html_path = os.path.join(DISCOVERY_OUTPUT_DIR, f"shopify_batch_{ts}.html")
            export_excel(all_products, all_brands, f"批量: {', '.join(keywords[:3])}", xlsx_path)
            export_html(all_products, all_brands, f"批量: {', '.join(keywords[:3])}", html_path)

    elif cmd == "verify":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py verify '产品关键词'")
            sys.exit(1)
        keyword = sys.argv[2]
        result = verify_shopify(keyword)
        print(f"\n{'='*60}")
        print(f"🔍 Shopify 验证: '{keyword}'")
        print(f"{'='*60}")
        print_verify_result(keyword, result)
        print(f"\n📋 完整数据:")
        for k, v in result.items():
            print(f"  {k}: {v}")

    elif cmd == "verify-batch":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py verify-batch input.xlsx [--top N]")
            sys.exit(1)
        input_path = sys.argv[2]
        top_n = 0
        for i, arg in enumerate(sys.argv):
            if arg == "--top" and i + 1 < len(sys.argv):
                top_n = int(sys.argv[i + 1])
        verify_batch_xlsx(input_path, top_n)

    else:
        print(f"未知命令: {cmd}")
        print(__doc__)
        sys.exit(1)



# ================================================================
#  Section 3: Store Manager (from shopify_manager)
# ================================================================

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            return json.load(f)
    return {}

def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

def run_shopify(args, timeout=60):
    """运行 shopify CLI 命令"""
    cmd = [SHOPIFY_CLI] + args
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return r.returncode, r.stdout, r.stderr
    except subprocess.TimeoutExpired:
        return -1, "", "timeout"
    except Exception as e:
        return -1, "", str(e)

def run_graphql(store, query, variables=None, mutations=False):
    """执行 GraphQL 查询"""
    args = ["store", "execute", "-s", store, "-q", query, "-j"]
    if mutations:
        args.append("--allow-mutations")
    if variables:
        args += ["-v", json.dumps(variables)]
    code, out, err = run_shopify(args)
    if code != 0:
        return {"error": err or out}
    try:
        return json.loads(out)
    except json.JSONDecodeError:
        return {"raw": out, "error": err}


# ─── AUTH ──────────────────────────────────────────────────
def cmd_auth(store):
    """认证 Shopify 店铺"""
    print(f"🔐 认证店铺: {store}")
    print("   将打开浏览器进行 OAuth 授权...")
    code, out, err = run_shopify(["store", "auth", "-s", store], timeout=120)
    if code == 0:
        cfg = load_config()
        cfg["store"] = store
        cfg["authenticated_at"] = datetime.now().isoformat()
        save_config(cfg)
        print(f"✅ 认证成功，已保存到 {CONFIG_FILE}")
    else:
        print(f"❌ 认证失败: {err or out}")
    return code == 0


# ─── INFO ──────────────────────────────────────────────────
def cmd_info():
    """查看店铺信息"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺。先运行: python3 shopify_ultimate.py auth your-store.myshopify.com")
        return

    query = """
    {
      shop {
        name
        email
        url
        myshopifyDomain
        plan { displayName }
        currencyCode
        timezoneAbbreviation
        shipsToCountries
      }
    }
    """
    result = run_graphql(store, query)
    shop = result.get("data", {}).get("shop", {})
    if shop:
        print(f"\n🏪 店铺信息:")
        print(f"   名称: {shop.get('name')}")
        print(f"   域名: {shop.get('myshopifyDomain')}")
        print(f"   邮箱: {shop.get('email')}")
        print(f"   套餐: {shop.get('plan', {}).get('displayName')}")
        print(f"   货币: {shop.get('currencyCode')}")
        print(f"   时区: {shop.get('timezoneAbbreviation')}")
        print(f"   配送国家: {', '.join(shop.get('shipsToCountries', []))}")
    else:
        print(f"❌ 查询失败: {result}")


# ─── PRODUCTS ──────────────────────────────────────────────
def cmd_products(limit=10):
    """列出产品"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    query = f"""
    {{
      products(first: {limit}) {{
        edges {{
          node {{
            id
            title
            status
            vendor
            productType
            createdAt
            variants(first: 1) {{
              edges {{
                node {{
                  price
                  compareAtPrice
                }}
              }}
            }}
            images(first: 1) {{
              edges {{
                node {{ url }}
              }}
            }}
          }}
        }}
      }}
    }}
    """
    result = run_graphql(store, query)
    products = result.get("data", {}).get("products", {}).get("edges", [])

    print(f"\n📦 产品列表 ({len(products)}):")
    print(f"{'─'*70}")
    for i, edge in enumerate(products, 1):
        p = edge["node"]
        variants = p.get("variants", {}).get("edges", [])
        price = variants[0]["node"]["price"] if variants else "N/A"
        compare = variants[0]["node"].get("compareAtPrice", "") if variants else ""
        status = p.get("status", "UNKNOWN")
        icon = "🟢" if status == "ACTIVE" else "🔴" if status == "ARCHIVED" else "⚪"
        price_str = f"${price}"
        if compare:
            price_str += f" (原价 ${compare})"
        print(f"  {icon} {i}. {p['title'][:50]}")
        print(f"      ID: {p['id']} | {price_str} | {p.get('vendor', '')} | {status}")
    print()


def cmd_create_product_from_json(json_path):
    """从 JSON 文件创建产品"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    with open(json_path) as f:
        data = json.load(f)

    title = data.get("title", "Untitled")
    body_html = data.get("description", data.get("body_html", ""))
    vendor = data.get("vendor", data.get("brand", ""))
    product_type = data.get("product_type", data.get("type", ""))
    price = str(data.get("price", "0.00"))
    compare_price = data.get("compare_at_price", data.get("comparePrice", ""))
    images = data.get("images", [])
    tags = data.get("tags", [])
    variants = data.get("variants", [])

    # 构建 variants JSON
    variants_input = []
    if variants:
        for v in variants:
            variants_input.append({
                "price": str(v.get("price", price)),
                "compareAtPrice": str(v.get("compare_at_price", compare_price)) if v.get("compare_at_price") else None,
                "sku": v.get("sku", ""),
                "inventoryQuantity": v.get("inventory_quantity", 0),
            })
    else:
        variants_input.append({
            "price": price,
            "compareAtPrice": str(compare_price) if compare_price else None,
        })

    # GraphQL mutation
    mutation = """
    mutation productCreate($input: ProductInput!) {
      productCreate(input: $input) {
        product {
          id
          title
          status
          variants(first: 1) {
            edges {
              node {
                price
                compareAtPrice
              }
            }
          }
        }
        userErrors {
          field
          message
        }
      }
    }
    """

    input_data = {
        "title": title,
        "descriptionHtml": body_html,
        "vendor": vendor,
        "productType": product_type,
        "tags": tags if isinstance(tags, list) else [t.strip() for t in str(tags).split(",")],
        "variants": variants_input,
    }

    if images:
        input_data["images"] = [{"src": url} for url in images[:10]]

    variables = {"input": input_data}
    result = run_graphql(store, mutation, variables, mutations=True)

    product = result.get("data", {}).get("productCreate", {})
    errors = product.get("userErrors", [])
    created = product.get("product", {})

    if errors:
        print(f"❌ 创建失败:")
        for e in errors:
            print(f"   {e.get('field')}: {e.get('message')}")
    elif created:
        vid = created.get("id", "")
        variants = created.get("variants", {}).get("edges", [])
        price = variants[0]["node"]["price"] if variants else "?"
        print(f"✅ 产品已创建:")
        print(f"   标题: {created.get('title')}")
        print(f"   ID: {vid}")
        print(f"   价格: ${price}")
        print(f"   状态: {created.get('status')}")
    else:
        print(f"❌ 未知错误: {result}")

    return created


def cmd_create_product_from_xlsx(xlsx_path, row=2):
    """从 Excel 行创建产品"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 需要 openpyxl")
        return

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    values = [ws.cell(row=row, column=i+1).value for i in range(len(headers))]

    data = dict(zip(headers, values))

    # 映射常见列名
    title_keys = ["标题", "Product", "Title", "Product Title", "title"]
    desc_keys = ["描述", "Description", "description", "独特卖点"]
    price_keys = ["Shopify售价", "shopify_sell_price", "Price", "售价"]
    brand_keys = ["品牌", "Brand", "brand", "vendor"]
    img_keys = ["产品图", "Image", "image_url", "main_image"]

    def find_val(keys):
        for k in keys:
            if k in data and data[k]:
                return data[k]
        return ""

    json_data = {
        "title": str(find_val(title_keys)),
        "description": str(find_val(desc_keys)),
        "price": float(find_val(price_keys) or 0),
        "vendor": str(find_val(brand_keys)),
        "images": [str(find_val(img_keys))] if find_val(img_keys) else [],
    }

    print(f"📋 从 Excel 提取产品数据:")
    print(f"   标题: {json_data['title'][:60]}")
    print(f"   价格: ${json_data['price']}")
    print(f"   品牌: {json_data['vendor']}")

    # 保存临时 JSON
    tmp_path = "/tmp/shopify_product_tmp.json"
    with open(tmp_path, "w") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    return cmd_create_product_from_json(tmp_path)


def cmd_update_product(product_id, json_path):
    """更新产品"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    with open(json_path) as f:
        data = json.load(f)

    mutation = """
    mutation productUpdate($input: ProductInput!) {
      productUpdate(input: $input) {
        product { id title status }
        userErrors { field message }
      }
    }
    """

    input_data = {"id": product_id}
    if "title" in data:
        input_data["title"] = data["title"]
    if "description" in data:
        input_data["descriptionHtml"] = data["description"]
    if "price" in data:
        input_data["variants"] = [{"price": str(data["price"])}]
    if "tags" in data:
        input_data["tags"] = data["tags"]

    result = run_graphql(store, mutation, {"input": input_data}, mutations=True)
    upd = result.get("data", {}).get("productUpdate", {})
    errors = upd.get("userErrors", [])

    if errors:
        for e in errors:
            print(f"❌ {e.get('field')}: {e.get('message')}")
    else:
        p = upd.get("product", {})
        print(f"✅ 已更新: {p.get('title')} ({p.get('id')})")


def cmd_delete_product(product_id):
    """删除产品"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    confirm = input(f"⚠️ 确认删除 {product_id}? (y/N): ").strip().lower()
    if confirm != "y":
        print("已取消")
        return

    mutation = """
    mutation productDelete($input: ProductDeleteInput!) {
      productDelete(input: $input) {
        deletedProductId
        userErrors { field message }
      }
    }
    """

    result = run_graphql(store, mutation, {"input": {"id": product_id}}, mutations=True)
    d = result.get("data", {}).get("productDelete", {})
    errors = d.get("userErrors", [])

    if errors:
        for e in errors:
            print(f"❌ {e.get('field')}: {e.get('message')}")
    else:
        print(f"✅ 已删除: {d.get('deletedProductId')}")


# ─── THEMES ────────────────────────────────────────────────
def cmd_themes():
    """列出主题"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    code, out, err = run_shopify(["theme", "list", "-s", store])
    if code == 0:
        print(f"\n🎨 主题列表:\n{out}")
    else:
        print(f"❌ {err or out}")


def cmd_theme_dev():
    """本地开发主题"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    print(f"🚀 启动开发主题模式...")
    print(f"   将打开浏览器，实时预览修改")
    os.system(f'{SHOPIFY_CLI} theme dev -s {store}')


def cmd_theme_push(theme_dir):
    """推送主题"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    print(f"📤 推送主题: {theme_dir}")
    os.system(f'{SHOPIFY_CLI} theme push -s {store} {theme_dir}')


def cmd_theme_pull(theme_dir):
    """拉取主题"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    print(f"📥 拉取主题到: {theme_dir}")
    os.system(f'{SHOPIFY_CLI} theme pull -s {store} -d {theme_dir}')


# ─── COLLECTIONS ───────────────────────────────────────────
def cmd_collections(limit=10):
    """列出集合"""
    cfg = load_config()
    store = cfg.get("store")
    if not store:
        print("❌ 未配置店铺")
        return

    query = f"""
    {{
      collections(first: {limit}) {{
        edges {{
          node {{
            id
            title
            handle
            description
            productsCount {{ count }}
            updatedAt
          }}
        }}
      }}
    }}
    """
    result = run_graphql(store, query)
    collections = result.get("data", {}).get("collections", {}).get("edges", [])

    print(f"\n📁 集合列表 ({len(collections)}):")
    print(f"{'─'*50}")
    for i, edge in enumerate(collections, 1):
        c = edge["node"]
        count = c.get("productsCount", {}).get("count", 0)
        print(f"  {i}. {c['title']} ({count} 个产品) — {c['handle']}")
        print(f"     ID: {c['id']}")
    print()


# ─── BULK IMPORT ───────────────────────────────────────────
def cmd_bulk_import(xlsx_path, limit=0):
    """批量导入产品"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 需要 openpyxl")
        return

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    total = ws.max_row - 1
    if limit > 0:
        total = min(total, limit)

    print(f"📦 批量导入: {xlsx_path}")
    print(f"   共 {total} 行待导入\n")

    success = 0
    failed = 0

    for row_idx in range(2, 2 + total):
        values = [ws.cell(row=row_idx, column=i+1).value for i in range(len(headers))]
        data = dict(zip(headers, values))

        title_keys = ["标题", "Product", "Title"]
        desc_keys = ["描述", "Description", "独特卖点"]
        price_keys = ["Shopify售价", "shopify_sell_price", "Price"]
        brand_keys = ["品牌", "Brand", "vendor"]

        def find_val(keys):
            for k in keys:
                if k in data and data[k]:
                    return data[k]
            return ""

        title = str(find_val(title_keys))
        if not title or title == "None":
            continue

        print(f"  [{row_idx-1}/{total+1}] {title[:50]}...", end=" ")

        tmp = {
            "title": title,
            "description": str(find_val(desc_keys)),
            "price": float(find_val(price_keys) or 0),
            "vendor": str(find_val(brand_keys)),
        }

        tmp_path = "/tmp/shopify_bulk_tmp.json"
        with open(tmp_path, "w") as f:
            json.dump(tmp, f, ensure_ascii=False)

        created = cmd_create_product_from_json(tmp_path)
        if created:
            success += 1
        else:
            failed += 1

        time.sleep(1)

    print(f"\n{'='*50}")
    print(f"✅ 成功: {success} | ❌ 失败: {failed}")
    print(f"{'='*50}")


# ─── 主入口 ────────────────────────────────────────────────
def manager_main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    cmd = sys.argv[1]

    if cmd == "auth":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py auth your-store.myshopify.com")
            sys.exit(1)
        cmd_auth(sys.argv[2])

    elif cmd == "info":
        cmd_info()

    elif cmd == "products":
        limit = 10
        for i, a in enumerate(sys.argv):
            if a == "--limit" and i+1 < len(sys.argv):
                limit = int(sys.argv[i+1])
        cmd_products(limit)

    elif cmd == "create-product":
        json_path = None
        xlsx_path = None
        row = 2
        for i, a in enumerate(sys.argv):
            if a == "--json" and i+1 < len(sys.argv):
                json_path = sys.argv[i+1]
            if a == "--from-xlsx" and i+1 < len(sys.argv):
                xlsx_path = sys.argv[i+1]
            if a == "--row" and i+1 < len(sys.argv):
                row = int(sys.argv[i+1])
        if json_path:
            cmd_create_product_from_json(json_path)
        elif xlsx_path:
            cmd_create_product_from_xlsx(xlsx_path, row)
        else:
            print("用法: python3 shopify_ultimate.py create-product --json data.json")
            print("   或: python3 shopify_ultimate.py create-product --from-xlsx report.xlsx --row 2")

    elif cmd == "update-product":
        if len(sys.argv) < 4:
            print("用法: python3 shopify_ultimate.py update-product <product-id> --json data.json")
            sys.exit(1)
        product_id = sys.argv[2]
        json_path = None
        for i, a in enumerate(sys.argv):
            if a == "--json" and i+1 < len(sys.argv):
                json_path = sys.argv[i+1]
        if json_path:
            cmd_update_product(product_id, json_path)
        else:
            print("需要 --json 参数")

    elif cmd == "delete-product":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py delete-product <product-id>")
            sys.exit(1)
        cmd_delete_product(sys.argv[2])

    elif cmd == "themes":
        cmd_themes()

    elif cmd == "theme-dev":
        cmd_theme_dev()

    elif cmd == "theme-push":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py theme-push ./theme-dir")
            sys.exit(1)
        cmd_theme_push(sys.argv[2])

    elif cmd == "theme-pull":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py theme-pull ./theme-dir")
            sys.exit(1)
        cmd_theme_pull(sys.argv[2])

    elif cmd == "collections":
        cmd_collections()

    elif cmd == "bulk-import":
        if len(sys.argv) < 3:
            print("用法: python3 shopify_ultimate.py bulk-import report.xlsx [--limit N]")
            sys.exit(1)
        limit = 0
        for i, a in enumerate(sys.argv):
            if a == "--limit" and i+1 < len(sys.argv):
                limit = int(sys.argv[i+1])
        cmd_bulk_import(sys.argv[2], limit)

    else:
        print(f"未知命令: {cmd}")
        print(__doc__)
        sys.exit(1)



# ================================================================
#  Unified CLI
# ================================================================

def main():
    # Determine which sub-tool to use based on first argument
    if len(sys.argv) < 2:
        print("Shopify Ultimate v5.0")
        print("Usage: python3 shopify_ultimate.py <command> [args]")
        print("")
        print("Monitor commands: scrape, check, auto, export")
        print("Discovery commands: search, batch, inspire, verify, verify-batch")
        print("Manager commands: auth, info, products, create-product, update-product,")
        print("  delete-product, themes, theme-dev, theme-push, theme-pull,")
        print("  collections, create-collection, bulk-import")
        print("Analysis commands: analyze <domain1> [domain2 ...] [--cost 100] [--export]")
        return

    cmd = sys.argv[1]

    # Monitor commands
    if cmd in ('scrape', 'check', 'auto', 'export'):
        monitor_main()
    # Discovery commands
    elif cmd in ('search', 'batch', 'inspire', 'verify', 'verify-batch'):
        discovery_main()
    # Manager commands
    elif cmd in ('auth', 'info', 'products', 'create-product', 'update-product',
                 'delete-product', 'themes', 'theme-dev', 'theme-push', 'theme-pull',
                 'collections', 'create-collection', 'bulk-import'):
        manager_main()
    # v5.0 Business Intelligence
    elif cmd == 'analyze':
        domains = []
        cost = 100.0
        do_export = False
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == '--cost' and i + 1 < len(sys.argv):
                cost = float(sys.argv[i + 1])
                i += 2
            elif sys.argv[i] == '--export':
                do_export = True
                i += 1
            else:
                domains.append(sys.argv[i])
                i += 1

        if not domains:
            print("用法: python3 shopify_ultimate.py analyze domain1.com [domain2.com] [--cost 100] [--export]")
            sys.exit(1)

        print(f"🧠 v5.0 商业智能分析: {', '.join(domains)} (成本=${cost})")
        session = get_session()
        stores = {}
        for d in domains:
            d = normalize_domain(d)
            print(f"\n  📦 抓取 {d}...")
            products = fetch_all_products(d, session)
            if products:
                stores[d] = products
                print(f"  ✅ {len(products)} 个产品")
            else:
                print(f"  ⚠️ 无数据")

        if not stores:
            print("❌ 无有效数据，退出")
            sys.exit(1)

        report = run_business_analysis(
            my_domain=domains[0],
            competitor_domains=domains[1:] if len(domains) > 1 else [],
            data_dir=str(DATA_DIR),
            stores_data=stores,
            cost=cost
        )

        # 打印文本报告
        text = format_business_report(report)
        print(text)

        # 导出
        if do_export:
            out_path = DATA_DIR / f"biz_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(report, f, ensure_ascii=False, indent=2, default=str)
            print(f"\n📁 JSON 报告: {out_path}")

    else:
        print(f"Unknown command: {cmd}")
        print("Use: scrape/check/auto/export | search/batch/inspire/verify/verify-batch | analyze | auth/info/products/...")


if __name__ == "__main__":
    main()
