#!/usr/bin/env python3
"""
ecommerce_rank_monitor.py — 有价值电商榜单机会雷达

目标：统一监控高价值榜单源，输出每日 Top 产品机会。
优先级清单：
1 Amazon Movers & Shakers
2 Amazon New Releases
3 Amazon Best Sellers
4 TikTok Shop Trending Products
5 TikTok Creative Center Top Ads
6 Facebook Ad Library active ads
7 Shopify 竞品新品
8 Etsy Bestsellers
9 Temu Best Sellers
10 AliExpress Top Ranking
11 Target Best Sellers
12 Wayfair Top Sellers
13 Walmart Best Sellers

当前实现：
- 先稳定接入本地已有高价值数据源：Amazon Movers 历史 Excel、Shopify 新品报告、TikTok/V7 输出、FB JSON 输出。
- 生成统一 JSON / Excel / Markdown 日报。
- 未接入/当天无数据的平台会进入 pending_sources，不伪造数据。

用法：
  python3 ecommerce_rank_monitor.py
  python3 ecommerce_rank_monitor.py --top 20
  python3 ecommerce_rank_monitor.py --run-shopify --shopify-limit 100
  python3 ecommerce_rank_monitor.py --run-v7
"""
from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

HOME = Path(os.environ.get("HOME") or "/Users/doi")
if not (HOME / "Desktop").exists():
    HOME = Path("/Users/doi")

BASE = HOME / "Desktop" / "amazon"
SEL = HOME / "Desktop" / "Selection"
OUT = BASE / "output" / "ecommerce_rank_monitor"
OUT.mkdir(parents=True, exist_ok=True)

WATCHLIST = BASE / "output" / "shopify_monitor" / "competitors_watchlist.json"
SHOPIFY_NEW_DIR = BASE / "output" / "shopify_monitor" / "new_products"
SHOPIFY_REPORT_DIR = BASE / "output" / "shopify_monitor" / "reports"
FB_RESULTS_DIRS = [
    HOME / ".claude" / "skills" / "fb-ad-library-results",
    HOME / ".hermes" / "skills" / "fb-ad-library-results",
    HOME / "Desktop" / "reports" / "fb-ad-library-results",
]
TIKTOK_DIRS = [
    HOME / "Downloads" / "tiktok-ads-scraper" / "output",
    SEL / "V7" / "output",
    SEL / "output",
]

PRIORITY_SOURCES = [
    {"priority": 1, "name": "Amazon Movers & Shakers", "status": "implemented", "cadence": "daily"},
    {"priority": 2, "name": "Amazon New Releases", "status": "pending_connector", "cadence": "daily"},
    {"priority": 3, "name": "Amazon Best Sellers", "status": "pending_connector", "cadence": "daily"},
    {"priority": 4, "name": "TikTok Shop Trending Products", "status": "implemented_if_data_exists", "cadence": "daily"},
    {"priority": 5, "name": "TikTok Creative Center Top Ads", "status": "implemented_if_data_exists", "cadence": "daily"},
    {"priority": 6, "name": "Facebook Ad Library active ads", "status": "implemented_if_data_exists", "cadence": "daily"},
    {"priority": 7, "name": "Shopify 竞品新品", "status": "implemented", "cadence": "daily"},
    {"priority": 8, "name": "Etsy Bestsellers", "status": "pending_connector", "cadence": "weekly"},
    {"priority": 9, "name": "Temu Best Sellers", "status": "pending_connector", "cadence": "weekly"},
    {"priority": 10, "name": "AliExpress Top Ranking", "status": "pending_connector", "cadence": "weekly"},
    {"priority": 11, "name": "Target Best Sellers", "status": "pending_connector", "cadence": "weekly"},
    {"priority": 12, "name": "Wayfair Top Sellers", "status": "pending_connector", "cadence": "weekly"},
    {"priority": 13, "name": "Walmart Best Sellers", "status": "pending_connector", "cadence": "weekly"},
]

TARGET_KEYWORDS = {
    "home": 10, "kitchen": 8, "sleep": 10, "relief": 10, "pain": 8, "massage": 10,
    "air": 8, "sound": 8, "quiet": 10, "calm": 10, "wellness": 10, "therapy": 8,
    "pet": 7, "baby": 5, "beauty": 6, "women": 8, "gift": 7, "organizer": 5,
    "pillow": 8, "blanket": 7, "humidifier": 8, "purifier": 9, "lamp": 6,
    "portable": 4, "wireless": 4, "heated": 6, "cooling": 6, "ergonomic": 8,
}
BAD_KEYWORDS = {
    "iphone", "samsung", "nike", "adidas", "lego", "disney", "pokemon", "toyota",
    "book", "kindle", "dvd", "game", "xbox", "playstation", "shirt", "dress", "shoe",
}

@dataclass
class Opportunity:
    source: str
    platform: str
    title: str
    url: str = ""
    price: Optional[float] = None
    rank: Optional[str] = None
    change: Optional[str] = None
    rating: Optional[float] = None
    reviews: Optional[int] = None
    category: str = ""
    brand: str = ""
    image: str = ""
    evidence: str = ""
    data_date: str = ""
    data_file: str = ""
    data_freshness_days: Optional[int] = None
    trend_metrics: str = ""
    source_transparency: str = ""
    score: float = 0.0
    opportunity_type: str = ""
    suggested_price: str = ""
    sellable_forms: str = ""
    risk: str = ""
    decision: str = ""
    decision_reason: str = ""
    product_angle: str = ""
    next_action: str = ""
    raw: Optional[Dict[str, Any]] = None


def norm_text(x: Any) -> str:
    return re.sub(r"\s+", " ", str(x or "")).strip()


def parse_price(x: Any) -> Optional[float]:
    if x is None:
        return None
    s = str(x)
    m = re.search(r"\$?\s*([0-9]+(?:\.[0-9]{1,2})?)", s.replace(",", ""))
    return float(m.group(1)) if m else None


def parse_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    s = str(x).replace(",", "")
    m = re.search(r"([0-9]+)", s)
    return int(m.group(1)) if m else None


def newest_file(patterns: Iterable[str]) -> Optional[Path]:
    files: List[str] = []
    for p in patterns:
        files.extend(glob.glob(p, recursive=True))
    files = [f for f in files if Path(f).is_file()]
    if not files:
        return None
    return Path(max(files, key=lambda f: Path(f).stat().st_mtime))


def file_age_days(path: Path) -> Optional[int]:
    try:
        return max(0, int((datetime.now() - datetime.fromtimestamp(path.stat().st_mtime)).total_seconds() // 86400))
    except Exception:
        return None


def date_from_path(path: Path) -> str:
    m = re.search(r"(20\d{2})[-_]?([01]\d)[-_]?([0-3]\d)", path.name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")
    except Exception:
        return ""


def with_source_meta(o: Opportunity, path: Path, detail: str = "") -> Opportunity:
    o.data_date = date_from_path(path)
    o.data_file = path.name
    o.data_freshness_days = file_age_days(path)
    fresh = f"{o.data_freshness_days}d old" if o.data_freshness_days is not None else "unknown age"
    o.source_transparency = f"file={path.name}; date={o.data_date}; freshness={fresh}"
    if detail:
        o.source_transparency += f"; {detail}"
    if not o.trend_metrics:
        o.trend_metrics = o.change or o.evidence or o.source_transparency
    if o.trend_metrics and o.trend_metrics not in o.evidence:
        o.evidence = (o.evidence + "; " + o.trend_metrics).strip("; ")
    return o


def row_get(d: Dict[str, Any], *names: str) -> Any:
    lower = {str(k).strip().lower(): v for k, v in d.items()}
    for name in names:
        if name in d and d.get(name) not in (None, ""):
            return d.get(name)
        v = lower.get(str(name).strip().lower())
        if v not in (None, ""):
            return v
    return None


def trend_from_fields(d: Dict[str, Any], fields: Iterable[str]) -> str:
    parts = []
    for f in fields:
        v = row_get(d, f)
        if v not in (None, ""):
            parts.append(f"{f}={norm_text(v)}")
    return "; ".join(parts[:8])


def score_item(o: Opportunity) -> Opportunity:
    title = o.title.lower()
    score = 0.0
    # 来源权重
    source_weight = {
        "Amazon Movers & Shakers": 22,
        "Amazon New Releases": 20,
        "Amazon Best Sellers": 18,
        "TikTok Shop Trending Products": 20,
        "TikTok Creative Center Top Ads": 18,
        "Facebook Ad Library active ads": 22,
        "Shopify 竞品新品": 18,
        "Amazon Dual-signal FB+BSR": 24,
        "FB Keyword Market Radar": 20,
        "Smart Selection Keyword Radar": 18,
        "Local historical marketplace exports": 12,
    }.get(o.source, 10)
    score += source_weight

    # 价格与利润空间
    if o.price is not None:
        if 80 <= o.price <= 500:
            score += 22
        elif 50 <= o.price < 80:
            score += 14
        elif 20 <= o.price < 50:
            score += 6
        elif o.price > 500:
            score += 8
        else:
            score -= 10
    else:
        score += 4

    # 评论/评分适中
    if o.rating and o.rating >= 4.2:
        score += 8
    if o.reviews is not None:
        if 50 <= o.reviews <= 3000:
            score += 10
        elif 1 <= o.reviews < 50:
            score += 6
        elif o.reviews > 10000:
            score -= 5

    # 关键词匹配
    for kw, pts in TARGET_KEYWORDS.items():
        if kw in title:
            score += pts
    for kw in BAD_KEYWORDS:
        if kw in title:
            score -= 10

    # 排名变化/广告活跃证据
    ev = (o.evidence + " " + str(o.change or "")).lower()
    if "%" in ev or "change" in ev or "mover" in ev:
        score += 8
    if "active" in ev or "ad" in ev or "shop now" in ev:
        score += 8
    if "new" in ev or "新品" in ev:
        score += 8

    if o.data_freshness_days is not None:
        if o.data_freshness_days <= 3:
            score += 6
        elif o.data_freshness_days <= 14:
            score += 2
        elif o.data_freshness_days > 30:
            score -= 8

    score = max(0, min(100, round(score, 1)))
    o.score = score

    if score >= 80:
        o.opportunity_type = "爆发品 / 广告验证"
    elif score >= 65:
        o.opportunity_type = "高客单机会 / 新品起量"
    elif score >= 50:
        o.opportunity_type = "观察池"
    else:
        o.opportunity_type = "低优先级"

    if o.price:
        if o.price < 80:
            o.suggested_price = f"${max(89, int(o.price * 2.2))}-${max(129, int(o.price * 3.2))}"
        else:
            o.suggested_price = f"${int(o.price * 1.3)}-${int(o.price * 2.2)}"
    else:
        o.suggested_price = "$89-$299"

    o.sellable_forms = "标准版 / 高端版 / 礼盒套装 / 配件组合"
    risks = []
    if o.price and o.price < 30:
        risks.append("低客单，需升级包装")
    if any(k in title for k in BAD_KEYWORDS):
        risks.append("品牌/IP/标品风险")
    if not o.url:
        risks.append("缺少链接，需二次验证")
    o.risk = "；".join(risks) if risks else "需验证供应链、广告合规、退货率"
    return o


def pro_judge(o: Opportunity) -> Opportunity:
    """30年选品师口径：不是榜单热就能做，先判能不能卖、能不能投、能不能赚钱。"""
    title = o.title.lower()
    text = " ".join([title, (o.category or "").lower(), (o.brand or "").lower()])
    no_reasons: List[str] = []
    maybe_reasons: List[str] = []
    go_reasons: List[str] = []

    hard_no = {
        "药品/OTC/医疗宣称，广告和合规风险高": ["pepcid", "famotidine", "heartburn", "acid reducer", "otc medicine", "tablet", "teeth cleaning spray", "bad breath in pets", "spray formula"],
        "大牌/强品牌词，独立站复刻空间小": ["amazon basics", "q-tips", "la roche", "cerave", "clinique", "revlon", "nyx", "ordinary", "sun bum", "hempz", "nécessaire", "necessaire", "earth breeze"],
        "安全认证/责任风险，不适合冷启动独立站": ["helmet", "sunscreen", "spf", "baby care", "light therapy", "red light", "led mask"],
        "服饰尺码/退货率/库存复杂": ["shirt", "dress", "shoe", "activewear", "leggings", "clothing", "hair bundles", "bundles🎀", "hairstore"],
        "极低客单标品，难以打到 $300 AOV": ["cotton swab", "hair ties", "loofah", "bath sponge", "cleaning gloves", "pimple patches"],
        "B2B/批发/设备类广告，不适合当前独立站选品": ["wholesale factory", "packaging bag", "supplier", "industrial", "factory#"],
        "内容/软件/测评服务，不是可直接上架SKU": ["hair score", "ai-powered", "stabilizer", "creation", "home of possibilities"],
    }
    for reason, kws in hard_no.items():
        if any(k in text for k in kws):
            no_reasons.append(reason)

    if any(k in text for k in ["mask", "retinol", "glycolic", "moisturizer", "balm", "makeup", "eyeliner", "brow", "skincare"]):
        maybe_reasons.append("美妆护肤竞争强/功效宣称敏感，除非有差异化配方和素材")
    if any(k in text for k in ["hair dryer", "1875w", "ionic"]):
        maybe_reasons.append("电器可做高客单，但要认证、退货率和差异化外观")
    if any(k in text for k in ["weight bench", "workout bench"]):
        maybe_reasons.append("高客单成立，但体积/运费/售后重，适合美国仓而非普通 dropshipping")
    if any(k in text for k in ["resistance bands", "exercise bands"]):
        maybe_reasons.append("健身小器械可组合套装，但同质化强，需要女性场景内容")
    if any(k in text for k in ["shower steamers", "aromatherapy shower", "shower aromatherapy"]):
        go_reasons.append("礼品+香薰+女性场景，适合做套装和订阅，素材容易")
    if any(k in text for k in ["laundry detergent sheets", "eco laundry"]):
        maybe_reasons.append("环保耗材+复购，但客单偏低且 Earth Breeze 品牌强，只能做订阅/组合装验证")
    if any(k in text for k in ["camping pillow", "lumbar support", "neck"]):
        maybe_reasons.append("旅行/露营细分可做，但客单偏低，需要多件套提升 AOV")
    if any(k in text for k in ["body scrub", "sugar scrub", "african net"]):
        maybe_reasons.append("沐浴护理有内容场景，但低客单，必须套装化")

    if o.source == "Amazon Movers & Shakers":
        maybe_reasons.append("目前只有 Amazon Movers 单信号，不能直接当爆品，需要二次验证 FB/TK/GT")
    if o.source in {"Facebook Ad Library active ads", "TikTok Creative Center Top Ads", "TikTok Shop Trending Products"}:
        if any(k in text for k in ["shower", "massager", "massage", "pillow", "organizer", "kitchen", "cleaning tool", "pet bed", "humidifier", "water filter", "hair dryer", "hair tool", "beauty device"]):
            go_reasons.append("已有广告/内容侧信号，且落在明确可上架消费品方向")
        else:
            maybe_reasons.append("有广告信号，但标题不是明确产品，需要先抽象成可卖 SKU")
    if o.price is not None and o.price < 35:
        maybe_reasons.append("当前客单低，必须靠 bundle/礼盒/订阅提升毛利")

    if no_reasons:
        o.decision = "NO-GO"
        o.decision_reason = "；".join(dict.fromkeys(no_reasons))
        o.product_angle = "不建议做原品；最多提取需求方向，不碰同款"
        o.next_action = "丢弃原品；若保留，只提炼关键词进入新品创意池"
        o.score = min(o.score, 39)
    elif go_reasons and len(maybe_reasons) <= 1:
        o.decision = "GO"
        o.decision_reason = "；".join(dict.fromkeys(go_reasons + maybe_reasons[:1]))
        if "shower steamers" in text or "aromatherapy" in text:
            o.product_angle = "女性礼品香薰淋浴套装：睡眠/减压/节日礼盒/订阅补充装"
            o.suggested_price = "$49-$129 起；礼盒/多套装冲 $150+，不强行 $300"
            o.next_action = "找 1688/Alibaba 香薰片源头；验证 FB/TK 礼品广告；做 3 套 Hook"
        elif "laundry" in text:
            o.product_angle = "环保洗衣耗材：旅行装+家庭装+订阅补充装，主打省空间/无塑料"
            o.suggested_price = "$39-$99；靠订阅/LTV，不按 $300 单品逻辑"
            o.next_action = "查竞品落地页和复购模型；找无品牌白标供应链；验证 CPC"
        else:
            o.product_angle = "可测试，但必须做差异化包装/套装/内容角度"
            o.next_action = "补 Amazon 详情+FB/TK 广告验证+供应链报价"
    else:
        o.decision = "MAYBE"
        rs = maybe_reasons or ["有榜单信号，但缺少价格/评论/广告/供应链验证"]
        o.decision_reason = "；".join(dict.fromkeys(rs))
        if "hair dryer" in text:
            o.product_angle = "卷发/速干/低噪高颜值吹风机，不做白牌同款，做外观和套装差异"
            o.suggested_price = "$89-$199"
            o.next_action = "先查认证/FBA退货/竞品广告；无差异化就放弃"
        elif "weight bench" in text:
            o.product_angle = "女性家庭健身角落套装：凳子+弹力带+课程内容"
            o.suggested_price = "$149-$299"
            o.next_action = "只看美国仓/低破损供应商；核算运费后再决定"
        elif "resistance bands" in text:
            o.product_angle = "女性臀腿训练套装：踏板弹力带+训练卡+收纳包"
            o.suggested_price = "$39-$89"
            o.next_action = "验证 TikTok 内容量和 CPA；能做素材再进测试"
        else:
            o.product_angle = "只保留需求方向，寻找更高客单/更强差异化替代品"
            o.next_action = "补三维验证：Amazon 评论痛点、FB/TK 广告、Google Trends"
        o.score = min(o.score, 69)
    return o


def load_amazon_movers(limit: int = 80) -> Tuple[List[Opportunity], str]:
    path = newest_file([str(SEL / "output" / "amazon_movers_shakers_*.xlsx")])
    if not path:
        return [], "未找到 amazon_movers_shakers_*.xlsx"
    try:
        import openpyxl
    except Exception:
        return [], "openpyxl 未安装"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], f"空文件 {path}"
    headers = [norm_text(h) for h in rows[0]]
    out: List[Opportunity] = []
    for r in rows[1:limit + 1]:
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        title = norm_text(d.get("产品名称") or d.get("Product Title") or d.get("title") or d.get("Title"))
        if not title:
            continue
        o = Opportunity(
            source="Amazon Movers & Shakers",
            platform="Amazon",
            title=title,
            url=norm_text(d.get("链接") or d.get("URL") or d.get("Link")),
            price=parse_price(d.get("价格") or d.get("Price")),
            rank=norm_text(d.get("排名") or d.get("Rank")),
            change=norm_text(d.get("变化%") or d.get("Change")),
            rating=parse_price(d.get("评分") or d.get("Rating")),
            reviews=parse_int(d.get("评论数") or d.get("Reviews")),
            category=norm_text(d.get("品类") or d.get("Category")),
            evidence=f"Amazon 24h Movers 榜单；文件 {path.name}",
            raw=d,
        )
        out.append(score_item(with_source_meta(o, path, f"rows<= {limit}")))
    return out, str(path)


def load_shopify_new(limit: int = 100) -> Tuple[List[Opportunity], str]:
    files = list(SHOPIFY_NEW_DIR.glob("*.xlsx")) + list(SHOPIFY_NEW_DIR.glob("*.json")) + list(SHOPIFY_REPORT_DIR.glob("*.json"))
    if not files:
        return [], "未找到 Shopify 新品输出"
    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:8]
    out: List[Opportunity] = []
    used = []
    for path in files:
        used.append(path.name)
        if path.suffix.lower() == ".json":
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            candidates = []
            if isinstance(data, list):
                candidates = data
            elif isinstance(data, dict):
                for k in ["new_products", "products", "items", "results"]:
                    v = data.get(k)
                    if isinstance(v, list):
                        candidates.extend(v)
                # batch report only has counts; skip if no product detail
            for item in candidates[:limit]:
                if not isinstance(item, dict):
                    continue
                title = norm_text(item.get("title") or item.get("name") or item.get("product_title"))
                if not title:
                    continue
                o = Opportunity(
                    source="Shopify 竞品新品",
                    platform="Shopify/DTC",
                    title=title,
                    url=norm_text(item.get("url") or item.get("link") or item.get("product_url")),
                    price=parse_price(item.get("price") or item.get("variants_price") or item.get("min_price")),
                    category=norm_text(item.get("product_type") or item.get("category") or item.get("vendor")),
                    brand=norm_text(item.get("domain") or item.get("vendor") or item.get("brand")),
                    image=norm_text(item.get("image") or item.get("image_url")),
                    evidence=f"Shopify 竞品新品；文件 {path.name}",
                    raw=item,
                )
                out.append(score_item(with_source_meta(o, path, f"shopify_json_candidates<= {limit}")))
        elif path.suffix.lower() == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [norm_text(h) for h in rows[0]]
                for r in rows[1:limit+1]:
                    d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
                    title = norm_text(d.get("标题") or d.get("产品名") or d.get("title") or d.get("Title"))
                    if not title:
                        continue
                    o = Opportunity(
                        source="Shopify 竞品新品",
                        platform="Shopify/DTC",
                        title=title,
                        url=norm_text(d.get("链接") or d.get("url") or d.get("URL")),
                        price=parse_price(d.get("价格") or d.get("price") or d.get("Price")),
                        category=norm_text(d.get("品类") or d.get("category") or d.get("Category")),
                        evidence=f"Shopify 竞品新品；文件 {path.name}",
                        raw=d,
                    )
                    out.append(score_item(with_source_meta(o, path, f"shopify_xlsx_rows<= {limit}")))
            except Exception:
                continue
    return out[:limit], ", ".join(used)


def load_fb_ads(limit: int = 120) -> Tuple[List[Opportunity], str]:
    files: List[Path] = []
    for d in FB_RESULTS_DIRS:
        if d.exists():
            files.extend(d.glob("*.json"))
    if not files:
        return [], "未找到 FB Ad Library JSON"
    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:30]
    out: List[Opportunity] = []
    for path in files:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        keyword = norm_text(data.get("keyword") if isinstance(data, dict) else "")
        rows = data.get("results") or data.get("ads") or [] if isinstance(data, dict) else []
        if not isinstance(rows, list):
            continue
        for item in rows[:20]:
            if not isinstance(item, dict):
                continue
            title = norm_text(item.get("title") or item.get("page_name") or keyword)
            body = norm_text(item.get("body") or item.get("text") or item.get("link_description"))
            link = norm_text(item.get("link_url") or item.get("url"))
            if not title and not body:
                continue
            o = Opportunity(
                source="Facebook Ad Library active ads",
                platform="Facebook/Meta",
                title=title or body[:80],
                url=link,
                brand=norm_text(item.get("page_name")),
                evidence=f"active ad; keyword={keyword}; file={path.name}; {body[:120]}",
                raw=item,
            )
            out.append(score_item(with_source_meta(o, path, f"keyword={keyword}")))
            if len(out) >= limit:
                break
        if len(out) >= limit:
            break
    return out, f"{len(files)} JSON files"


def load_tiktok(limit: int = 120) -> Tuple[List[Opportunity], str]:
    patterns = []
    for d in TIKTOK_DIRS:
        patterns += [str(d / "**" / "*.csv"), str(d / "**" / "*.json")]
    files = [Path(f) for f in glob.glob(str(patterns[0]), recursive=True)] if patterns else []
    files = []
    for p in patterns:
        files += [Path(f) for f in glob.glob(p, recursive=True)]
    files = [f for f in files if f.is_file() and any(s in f.name.lower() for s in ["top", "tiktok", "ads", "product", "viral"])]
    if not files:
        return [], "未找到 TikTok 输出"
    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[:20]
    out: List[Opportunity] = []
    used = []
    for path in files:
        used.append(path.name)
        try:
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", errors="ignore") as f:
                    for row in csv.DictReader(f):
                        title = norm_text(row.get("ad_title") or row.get("title") or row.get("product_name") or row.get("url_title") or row.get("brand_name"))
                        if not title:
                            continue
                        source = "TikTok Creative Center Top Ads"
                        if "product" in path.name.lower():
                            source = "TikTok Shop Trending Products"
                        o = Opportunity(
                            source=source,
                            platform="TikTok",
                            title=title,
                            url=norm_text(row.get("landing_page") or row.get("tt_url") or row.get("url")),
                            category=norm_text(row.get("industry_name") or row.get("category") or row.get("first_ecom_category")),
                            image=norm_text(row.get("video_cover") or row.get("cover_url")),
                            evidence=f"TikTok data; file={path.name}; ctr={row.get('ctr','')}; likes={row.get('like','')}",
                            raw=dict(row),
                        )
                        out.append(score_item(with_source_meta(o, path, "csv")))
                        if len(out) >= limit:
                            break
            elif path.suffix.lower() == ".json":
                data = json.loads(path.read_text(encoding="utf-8"))
                rows = data if isinstance(data, list) else data.get("results") or data.get("items") or data.get("data") or []
                if isinstance(rows, dict):
                    rows = rows.get("list") or rows.get("items") or []
                if not isinstance(rows, list):
                    continue
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    title = norm_text(row.get("ad_title") or row.get("title") or row.get("product_name") or row.get("url_title") or row.get("name"))
                    if not title:
                        continue
                    source = "TikTok Shop Trending Products" if "product" in path.name.lower() else "TikTok Creative Center Top Ads"
                    o = Opportunity(
                        source=source,
                        platform="TikTok",
                        title=title,
                        url=norm_text(row.get("landing_page") or row.get("tt_url") or row.get("url")),
                        category=norm_text(row.get("industry_name") or row.get("category")),
                        image=norm_text(row.get("video_cover") or row.get("cover_url")),
                        evidence=f"TikTok data; file={path.name}",
                        raw=row,
                    )
                    out.append(score_item(with_source_meta(o, path, "json")))
                    if len(out) >= limit:
                        break
        except Exception:
            continue
        if len(out) >= limit:
            break
    return out, ", ".join(used[:5])


def load_dual_signal(limit: int = 100) -> Tuple[List[Opportunity], str]:
    path = newest_file([str(SEL / "output" / "dual_signal*.xlsx"), str(SEL / "output" / "dual_signal_candidates*.xlsx")])
    if not path:
        return [], "未找到 dual_signal*.xlsx"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [], f"读取失败 {e}"
    out: List[Opportunity] = []
    used_sheets = []
    for sheet in wb.sheetnames[:3]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm_text(h) for h in rows[0]]
        if not any(h in headers for h in ["产品标题", "ASIN", "综合得分"]):
            continue
        used_sheets.append(sheet)
        for r in rows[1:limit + 1]:
            d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
            title = norm_text(row_get(d, "产品标题", "Product Title", "title", "关键词"))
            if not title:
                continue
            asin = norm_text(row_get(d, "ASIN"))
            link = norm_text(row_get(d, "链接", "URL")) or (f"https://www.amazon.com/dp/{asin}" if asin else "")
            trend = trend_from_fields(d, ["趋势变化", "匹配FB关键词", "关键词数", "FB广告总量", "活跃广告数", "Amazon分(30%)", "FB广告分(45%)", "相关性分(25%)", "综合得分"])
            o = Opportunity(
                source="Amazon Dual-signal FB+BSR",
                platform="Amazon+Facebook",
                title=title,
                url=link,
                price=parse_price(row_get(d, "价格", "Price")),
                rank=norm_text(row_get(d, "Amazon排名", "排名", "Rank")),
                change=norm_text(row_get(d, "趋势变化", "Change")),
                rating=parse_price(row_get(d, "评分", "Rating")),
                category=norm_text(row_get(d, "品类", "Category")),
                evidence=f"Amazon榜单 + FB广告双信号；sheet={sheet}; ASIN={asin}",
                trend_metrics=trend,
                raw=d,
            )
            out.append(score_item(with_source_meta(o, path, f"sheet={sheet}")))
            if len(out) >= limit:
                return out, f"{path.name}; sheets={','.join(used_sheets)}"
    return out, f"{path.name}; sheets={','.join(used_sheets)}"


def load_fb_keyword_radar(limit: int = 120) -> Tuple[List[Opportunity], str]:
    path = newest_file([str(SEL / "output" / "FB广告扫描_*.xlsx")])
    if not path:
        return [], "未找到 FB广告扫描_*.xlsx"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["赛道总览"] if "赛道总览" in wb.sheetnames else wb[wb.sheetnames[0]]
    except Exception as e:
        return [], f"读取失败 {e}"
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], f"空文件 {path.name}"
    headers = [norm_text(h) for h in rows[0]]
    out: List[Opportunity] = []
    for r in rows[1:limit + 1]:
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        kw = norm_text(row_get(d, "关键词", "keyword"))
        if not kw:
            continue
        trend = trend_from_fields(d, ["活跃广告", "总广告", "市场热度", "热度分", "长跑广告(>30天)", "长跑比例%", "利润信号", "视频比例%", "Shopify数", "独立域名数", "头部域名", "品牌壁垒", "选品师判定"])
        o = Opportunity(
            source="FB Keyword Market Radar",
            platform="Facebook/Meta",
            title=kw,
            rank=norm_text(row_get(d, "排名")),
            evidence="FB关键词赛道总览；active ads only",
            trend_metrics=trend,
            raw=d,
        )
        out.append(score_item(with_source_meta(o, path, "sheet=赛道总览")))
    return out, str(path)


def load_smart_selection(limit: int = 120) -> Tuple[List[Opportunity], str]:
    path = newest_file([str(SEL / "output" / "智能选品_*.xlsx")])
    if not path:
        return [], "未找到 智能选品_*.xlsx"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["选品Dashboard"] if "选品Dashboard" in wb.sheetnames else wb[wb.sheetnames[0]]
    except Exception as e:
        return [], f"读取失败 {e}"
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], f"空文件 {path.name}"
    headers = [norm_text(h) for h in rows[0]]
    out: List[Opportunity] = []
    for r in rows[1:limit + 1]:
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        kw = norm_text(row_get(d, "关键词", "keyword"))
        if not kw:
            continue
        trend = trend_from_fields(d, ["综合分", "需求验证", "利润信号", "竞争窗口", "素材复制", "入场门槛", "供应链", "独立站适配", "广告数", "品牌数", "季节性", "供应基地", "选品师判定", "关键理由"])
        o = Opportunity(
            source="Smart Selection Keyword Radar",
            platform="FB+Amazon+SupplyChain",
            title=kw,
            rank=norm_text(row_get(d, "排名")),
            category=norm_text(row_get(d, "品类", "品类(中文)")),
            evidence="本地智能选品Dashboard；多维评分",
            trend_metrics=trend,
            raw=d,
        )
        out.append(score_item(with_source_meta(o, path, "sheet=选品Dashboard")))
    return out, str(path)


def dedupe(items: List[Opportunity]) -> List[Opportunity]:
    seen = {}
    for o in items:
        key = re.sub(r"[^a-z0-9]+", " ", o.title.lower()).strip()[:90]
        if not key:
            continue
        if key not in seen or o.score > seen[key].score:
            seen[key] = o
    return list(seen.values())


def write_excel(items: List[Opportunity], path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Opportunities"
    headers = ["decision", "score", "source", "platform", "title", "product_angle", "suggested_price", "decision_reason", "next_action", "price", "rank", "change", "rating", "reviews", "category", "brand", "url", "trend_metrics", "source_transparency", "data_date", "data_freshness_days", "evidence", "sellable_forms", "risk"]
    ws.append(headers)
    for o in items:
        ws.append([getattr(o, h) for h in headers])
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        ws.column_dimensions[get_column_letter(c)].width = min(55, max(12, len(headers[c-1]) + 2))
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        if row[0].value == "GO":
            row[0].fill = PatternFill("solid", fgColor="C6EFCE")
        elif row[0].value == "MAYBE":
            row[0].fill = PatternFill("solid", fgColor="FFEB9C")
        elif row[0].value == "NO-GO":
            row[0].fill = PatternFill("solid", fgColor="F4CCCC")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(path)
    return True


def markdown_report(items: List[Opportunity], meta: Dict[str, Any], top: int) -> str:
    lines = []
    go = [o for o in items if o.decision == "GO"]
    maybe = [o for o in items if o.decision == "MAYBE"]
    nogo = [o for o in items if o.decision == "NO-GO"]
    lines.append(f"# 今日电商榜单机会雷达 — {meta['date']}")
    lines.append("")
    lines.append(f"数据源命中：{meta['sources_with_data']} / {len(PRIORITY_SOURCES)}")
    lines.append(f"候选数：{meta['raw_count']} → 去重后：{meta['dedup_count']} → GO {len(go)} / MAYBE {len(maybe)} / NO-GO {len(nogo)}")
    lines.append("数据源透明度：每条记录输出 data_file / data_date / freshness；无数据源不伪造。")
    lines.append("")
    lines.append("## 结论：真正可以做的")
    if go:
        for i, o in enumerate(go[:10], 1):
            lines.append(f"### GO {i}：{o.title}")
            lines.append(f"- 判断：可以测试，不是直接抄同款")
            lines.append(f"- 产品化方向：{o.product_angle}")
            lines.append(f"- 建议售价：{o.suggested_price}")
            lines.append(f"- 为什么能做：{o.decision_reason}")
            lines.append(f"- 下一步：{o.next_action}")
            lines.append(f"- 证据：{o.platform} / {o.source}；{o.evidence}")
            lines.append(f"- 数据源：{o.source_transparency}")
            if o.url:
                lines.append(f"- 链接：{o.url}")
            lines.append("")
    else:
        lines.append("- 今天没有直接 GO 的产品。不要硬上；只保留 MAYBE 做二次验证。")
        lines.append("")

    lines.append("## 值得二次验证 / 可做变体")
    for i, o in enumerate(maybe[:max(top, 15)], 1):
        lines.append(f"### MAYBE {i}：{o.title}")
        lines.append(f"- 不是原品照抄：{o.product_angle}")
        lines.append(f"- 卡点：{o.decision_reason}")
        lines.append(f"- 下一步：{o.next_action}")
        lines.append(f"- 售价：{o.suggested_price}；分数：{o.score}/100")
        lines.append(f"- 趋势/来源：{o.trend_metrics or o.evidence}；{o.source_transparency}")
        if o.url:
            lines.append(f"- 链接：{o.url}")
        lines.append("")

    lines.append("## 直接淘汰：不要浪费时间")
    for i, o in enumerate(nogo[:20], 1):
        lines.append(f"- NO-GO {i}：{o.title}")
        lines.append(f"  - 原因：{o.decision_reason}")
        if o.url:
            lines.append(f"  - 链接：{o.url}")
    lines.append("")
    lines.append("## 专业筛选原则")
    lines.append("- 单一 Amazon Movers 只代表短期波动，不等于可投爆品。")
    lines.append("- OTC/医疗、防晒、头盔、强品牌、极低客单标品直接淘汰。")
    lines.append("- 能做的是：女性场景强、可套装化、能讲痛点、能找白标源头、广告素材容易的方向。")
    lines.append("- 下一步只验证 GO + 前 5 个 MAYBE，避免把时间花在垃圾信号上。")
    lines.append("")
    lines.append("## 待接入/无数据榜单")
    for s in meta["pending_sources"]:
        lines.append(f"- P{s['priority']} {s['name']}：{s['status']} / {s['cadence']}")
    return "\n".join(lines)


def run_cmd(cmd: List[str], cwd: Path, timeout: int = 600):
    print("RUN:", " ".join(cmd))
    return subprocess.run(cmd, cwd=str(cwd), text=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=timeout)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=20)
    ap.add_argument("--run-shopify", action="store_true", help="先运行 Shopify 新品监控，再汇总")
    ap.add_argument("--shopify-limit", type=int, default=100)
    ap.add_argument("--run-v7", action="store_true", help="先运行 V7 tier2/tier3 扫描，再汇总")
    args = ap.parse_args()

    logs = []
    if args.run_shopify:
        cmd = [sys.executable, "monitor_new_products.py", "--limit", str(args.shopify_limit), "--workers", "20", "--no-trend"]
        try:
            res = run_cmd(cmd, BASE, timeout=1800)
            logs.append({"cmd": " ".join(cmd), "returncode": res.returncode, "output_tail": res.stdout[-3000:]})
        except Exception as e:
            logs.append({"cmd": " ".join(cmd), "error": str(e)})

    if args.run_v7:
        cmd = [sys.executable, "run.py", "--tier2"]
        try:
            res = run_cmd(cmd, SEL / "V7", timeout=1800)
            logs.append({"cmd": " ".join(cmd), "returncode": res.returncode, "output_tail": res.stdout[-3000:]})
        except Exception as e:
            logs.append({"cmd": " ".join(cmd), "error": str(e)})

    all_items: List[Opportunity] = []
    source_notes = {}
    loaders = [load_amazon_movers, load_dual_signal, load_fb_keyword_radar, load_smart_selection, load_tiktok, load_fb_ads, load_shopify_new]
    for loader in loaders:
        items, note = loader()
        all_items.extend(items)
        source_notes[loader.__name__] = {"count": len(items), "note": note}

    raw_count = len(all_items)
    items = [pro_judge(o) for o in dedupe(all_items)]
    decision_rank = {"GO": 0, "MAYBE": 1, "NO-GO": 2}
    items.sort(key=lambda x: (decision_rank.get(x.decision, 9), -x.score))

    sources_with_data = len(set(o.source for o in items))
    implemented_sources = set(o.source for o in items)
    pending = []
    for s in PRIORITY_SOURCES:
        if s["name"] not in implemented_sources:
            pending.append(s)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    date = datetime.now().strftime("%Y-%m-%d %H:%M")
    json_path = OUT / f"rank_monitor_{ts}.json"
    md_path = OUT / f"rank_monitor_{ts}.md"
    xlsx_path = OUT / f"rank_monitor_{ts}.xlsx"
    latest_md = OUT / "latest.md"
    latest_json = OUT / "latest.json"
    latest_xlsx = OUT / "latest.xlsx"

    meta = {
        "date": date,
        "raw_count": raw_count,
        "dedup_count": len(items),
        "sources_with_data": sources_with_data,
        "source_notes": source_notes,
        "pending_sources": pending,
        "logs": logs,
    }
    payload = {"meta": meta, "items": [asdict(o) for o in items]}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    latest_json.write_text(json_path.read_text(encoding="utf-8"), encoding="utf-8")

    md = markdown_report(items, meta, args.top)
    md_path.write_text(md, encoding="utf-8")
    latest_md.write_text(md, encoding="utf-8")
    excel_ok = write_excel(items[: max(args.top, 50)], xlsx_path)
    if excel_ok:
        try:
            latest_xlsx.unlink(missing_ok=True)
        except Exception:
            pass
        import shutil
        shutil.copy2(xlsx_path, latest_xlsx)

    print(f"✅ 完成：候选 {raw_count}，去重 {len(items)}，来源 {sources_with_data}")
    print(f"MD: {md_path}")
    print(f"JSON: {json_path}")
    print(f"XLSX: {xlsx_path if excel_ok else 'openpyxl unavailable'}")
    print("\n" + "="*60 + "\n")
    print("\n".join(md.splitlines()[:80]))


if __name__ == "__main__":
    main()
